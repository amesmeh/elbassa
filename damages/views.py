from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Avg
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
import pandas as pd
import io
from .models import Damage
from .forms import DamageForm, DamageSearchForm, DamageImportForm
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def damages_list(request):
    """عرض قائمة الأضرار مع البحث والتصفية"""
    damages = Damage.objects.all().order_by('-created_at')
    search_form = DamageSearchForm(request.GET)
    
    # تطبيق البحث والتصفية
    if search_form.is_valid():
        if search_form.cleaned_data['search_query']:
            query = search_form.cleaned_data['search_query']
            damages = damages.filter(
                Q(guardian_name__icontains=query) |
                Q(guardian_national_id__icontains=query) |
                Q(notes__icontains=query)
            )
        
        if search_form.cleaned_data['damage_type']:
            damages = damages.filter(damage_type=search_form.cleaned_data['damage_type'])
        
        if search_form.cleaned_data['housing_type']:
            damages = damages.filter(housing_type=search_form.cleaned_data['housing_type'])
        
        if search_form.cleaned_data['district']:
            damages = damages.filter(district=search_form.cleaned_data['district'])
        
        if search_form.cleaned_data['date_from']:
            damages = damages.filter(damage_date__gte=search_form.cleaned_data['date_from'])
        
        if search_form.cleaned_data['date_to']:
            damages = damages.filter(damage_date__lte=search_form.cleaned_data['date_to'])
        
        if search_form.cleaned_data['cost_ils_min']:
            damages = damages.filter(estimated_cost_ils__gte=search_form.cleaned_data['cost_ils_min'])
        
        if search_form.cleaned_data['cost_ils_max']:
            damages = damages.filter(estimated_cost_ils__lte=search_form.cleaned_data['cost_ils_max'])
    
    # إحصائيات
    total_damages = damages.count()
    avg_damage_percentage = damages.aggregate(avg=Avg('damage_percentage'))['avg'] or 0
    total_cost_ils = damages.aggregate(total=Sum('estimated_cost_ils'))['total'] or 0
    total_cost_usd = damages.aggregate(total=Sum('estimated_cost_usd'))['total'] or 0
    
    # التصفح
    paginator = Paginator(damages, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_count': total_damages,
        'avg_damage_percentage': avg_damage_percentage,
        'total_cost_ils': total_cost_ils,
        'total_cost_usd': total_cost_usd,
        'title': 'الأضرار'
    }
    return render(request, 'damages/damages_list.html', context)

def damage_detail(request, pk):
    """عرض تفاصيل ضرر"""
    damage = get_object_or_404(Damage, pk=pk)
    context = {
        'damage': damage,
        'title': f'تفاصيل الضرر - {damage.guardian_name}'
    }
    return render(request, 'damages/damage_detail.html', context)

def damage_add(request):
    """إضافة ضرر جديد"""
    if request.method == 'POST':
        form = DamageForm(request.POST)
        if form.is_valid():
            damage = form.save()
            messages.success(request, f'تم إضافة الضرر لـ {damage.guardian_name} بنجاح.')
            return redirect('damages:damage_detail', pk=damage.pk)
    else:
        form = DamageForm()
    
    context = {
        'form': form,
        'title': 'إضافة ضرر جديد'
    }
    return render(request, 'damages/damage_form.html', context)

def damage_edit(request, pk):
    """تعديل ضرر"""
    damage = get_object_or_404(Damage, pk=pk)
    
    if request.method == 'POST':
        form = DamageForm(request.POST, instance=damage)
        if form.is_valid():
            damage = form.save()
            messages.success(request, f'تم تحديث بيانات الضرر لـ {damage.guardian_name} بنجاح.')
            return redirect('damages:damage_detail', pk=damage.pk)
    else:
        form = DamageForm(instance=damage)
    
    context = {
        'form': form,
        'damage': damage,
        'title': f'تعديل الضرر - {damage.guardian_name}'
    }
    return render(request, 'damages/damage_form.html', context)

@require_http_methods(["DELETE"])
def damage_delete(request, pk):
    """حذف ضرر"""
    damage = get_object_or_404(Damage, pk=pk)
    guardian_name = damage.guardian_name
    damage.delete()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'تم حذف الضرر لـ {guardian_name} بنجاح.'})
    
    messages.success(request, f'تم حذف الضرر لـ {guardian_name} بنجاح.')
    return redirect('damages:damages_list')

def damage_import(request):
    """استيراد الأضرار من ملف Excel"""
    if request.method == 'POST':
        form = DamageImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                
                # قراءة ملف Excel
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['guardian_national_id', 'guardian_name', 'housing_type', 'damage_type', 'damage_percentage']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'damages/damage_import.html', {'form': form})
                
                # استيراد البيانات
                imported_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # التحقق من عدم وجود الضرر مسبقاً
                        if Damage.objects.filter(
                            guardian_national_id=row['guardian_national_id'],
                            housing_type=row.get('housing_type', 'apartment')
                        ).exists():
                            errors.append(f'الصف {index + 2}: الضرر موجود مسبقاً')
                            continue
                        
                        # إنشاء الضرر
                        damage = Damage(
                            guardian_national_id=row['guardian_national_id'],
                            guardian_name=row['guardian_name'],
                            phone_number=row.get('phone_number', ''),
                            ownership_type=row.get('ownership_type', 'owned'),
                            housing_type=row['housing_type'],
                            housing_condition=row.get('housing_condition', 'good'),
                            construction_type=row.get('construction_type', 'concrete'),
                            floors_count=row.get('floors_count', 1),
                            damage_type=row['damage_type'],
                            damage_percentage=row['damage_percentage'],
                            damage_date=pd.to_datetime(row['damage_date']).date() if pd.notna(row.get('damage_date')) else None,
                            estimated_cost_ils=row.get('estimated_cost_ils', 0) if pd.notna(row.get('estimated_cost_ils')) else None,
                            estimated_cost_usd=row.get('estimated_cost_usd', 0) if pd.notna(row.get('estimated_cost_usd')) else None,
                            notes=row.get('notes', '')
                        )
                        
                        damage.full_clean()
                        damage.save()
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                
                # عرض النتائج
                if imported_count > 0:
                    messages.success(request, f'تم استيراد {imported_count} ضرر بنجاح.')
                
                if errors:
                    error_message = f'حدثت أخطاء في {len(errors)} صف:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_message += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_message)
                
                if imported_count > 0:
                    return redirect('damages:damages_list')
                    
            except Exception as e:
                messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    else:
        form = DamageImportForm()
    
    context = {
        'form': form,
        'title': 'استيراد الأضرار'
    }
    return render(request, 'damages/damage_import.html', context)

def damage_search_ajax(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    damages = Damage.objects.filter(
        Q(property_owner_name__icontains=query) |
        Q(property_owner_id__icontains=query) |
        Q(guardian_name__icontains=query) |
        Q(damage_description__icontains=query)
    )[:10]
    
    results = []
    for damage in damages:
        results.append({
            'id': damage.id,
            'text': f'{damage.property_owner_name} - {damage.get_damage_type_display()} ({damage.damage_date})',
            'property_owner_name': damage.property_owner_name,
            'property_owner_id': damage.property_owner_id,
            'damage_type': damage.damage_type,
            'estimated_cost': str(damage.estimated_cost)
        })
    
    return JsonResponse({'results': results})

def export_damages_excel(request):
    """تصدير بيانات الأضرار إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    search_form = DamageSearchForm(request.GET)
    damages = Damage.objects.all().order_by('-created_at')
    
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search_query')
        if search_query:
            damages = damages.filter(
                Q(property_owner_name__icontains=search_query) |
                Q(property_owner_id__icontains=search_query) |
                Q(damage_description__icontains=search_query)
            )
        
        damage_type = search_form.cleaned_data.get('damage_type')
        if damage_type:
            damages = damages.filter(damage_type=damage_type)
            
        housing_type = search_form.cleaned_data.get('housing_type')
        if housing_type:
            damages = damages.filter(housing_type=housing_type)
            
        date_from = search_form.cleaned_data.get('date_from')
        if date_from:
            damages = damages.filter(damage_date__gte=date_from)
            
        date_to = search_form.cleaned_data.get('date_to')
        if date_to:
            damages = damages.filter(damage_date__lte=date_to)
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الأضرار"
    
    # العناوين
    headers = [
        'صاحب العقار', 'رقم الهوية', 'نوع السكن', 'نوع الضرر', 'نسبة الضرر',
        'التكلفة (شيكل)', 'التكلفة (دولار)', 'تاريخ الضرر', 'مكان الضرر',
        'رقم الهاتف', 'وصف الضرر', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, damage in enumerate(damages, 2):
        ws.cell(row=row, column=1, value=damage.property_owner_name or "")
        ws.cell(row=row, column=2, value=damage.property_owner_id or "")
        ws.cell(row=row, column=3, value=damage.get_housing_type_display())
        ws.cell(row=row, column=4, value=damage.get_damage_type_display())
        ws.cell(row=row, column=5, value=f"{damage.damage_percentage or 0}%")
        ws.cell(row=row, column=6, value=float(damage.estimated_cost_ils or 0))
        ws.cell(row=row, column=7, value=float(damage.estimated_cost_usd or 0))
        ws.cell(row=row, column=8, value=damage.damage_date.strftime('%Y-%m-%d') if damage.damage_date else "")
        ws.cell(row=row, column=9, value=damage.damage_location or "")
        ws.cell(row=row, column=10, value=damage.phone_number or "")
        ws.cell(row=row, column=11, value=damage.damage_description or "")
        ws.cell(row=row, column=12, value=damage.notes or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # إضافة ملاحظات توضيحية
    ws.cell(row=4, column=1, value="ملاحظات:")
    ws.cell(row=5, column=1, value="• نوع السكن: apartment, house, commercial, warehouse")
    ws.cell(row=6, column=1, value="• نوع الضرر: minor, moderate, major, total")
    ws.cell(row=7, column=1, value="• التاريخ: استخدم تنسيق YYYY-MM-DD")
    ws.cell(row=8, column=1, value="• رقم الهوية: 9 أرقام بالضبط")
    
    # حفظ الملف في response
    filename = f"الأضرار_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_damages_excel(request):
    """استيراد بيانات الأضرار من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات المطلوبة
                    if pd.isna(row.get('صاحب العقار')) or pd.isna(row.get('رقم الهوية')):
                        errors.append(f"الصف {index + 2}: صاحب العقار ورقم الهوية مطلوبان")
                        error_count += 1
                        continue
                    
                    # إنشاء الضرر
                    damage = Damage.objects.create(
                        property_owner_name=str(row['صاحب العقار']).strip(),
                        property_owner_id=str(row['رقم الهوية']).strip(),
                        housing_type=str(row.get('نوع السكن', 'apartment')).strip().lower() if not pd.isna(row.get('نوع السكن')) else 'apartment',
                        damage_type=str(row.get('نوع الضرر', 'minor')).strip().lower() if not pd.isna(row.get('نوع الضرر')) else 'minor',
                        damage_percentage=float(row.get('نسبة الضرر', 0)) if not pd.isna(row.get('نسبة الضرر')) else 0,
                        estimated_cost_ils=float(row.get('التكلفة (شيكل)', 0)) if not pd.isna(row.get('التكلفة (شيكل)')) else 0,
                        estimated_cost_usd=float(row.get('التكلفة (دولار)', 0)) if not pd.isna(row.get('التكلفة (دولار)')) else 0,
                        damage_date=pd.to_datetime(row['تاريخ الضرر']).date() if not pd.isna(row.get('تاريخ الضرر')) else None,
                        damage_location=str(row.get('مكان الضرر', '')).strip() if not pd.isna(row.get('مكان الضرر')) else '',
                        phone_number=str(row.get('رقم الهاتف', '')).strip() if not pd.isna(row.get('رقم الهاتف')) else '',
                        damage_description=str(row.get('وصف الضرر', '')).strip() if not pd.isna(row.get('وصف الضرر')) else '',
                        notes=str(row.get('ملاحظات', '')).strip() if not pd.isna(row.get('ملاحظات')) else '',
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} ضرر بنجاح")
            
            if error_count > 0:
                messages.warning(request, f"فشل في استيراد {error_count} صف. الأخطاء: {'; '.join(errors[:5])}")
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('damages:damages_list')

def download_damages_template(request):
    """تحميل نموذج Excel فارغ للأضرار"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج الأضرار"
    
    # العناوين
    headers = [
        'صاحب العقار', 'رقم الهوية', 'نوع السكن', 'نوع الضرر', 'نسبة الضرر',
        'التكلفة (شيكل)', 'التكلفة (دولار)', 'تاريخ الضرر', 'مكان الضرر',
        'رقم الهاتف', 'وصف الضرر', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'أحمد محمد', '123456789', 'apartment', 'major', '75',
        '50000', '15000', '2024-01-15', 'رفح',
        '0591234567', 'ضرر في السقف والجدران', 'يحتاج إعادة بناء'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # إضافة ملاحظات توضيحية
    ws.cell(row=4, column=1, value="ملاحظات:")
    ws.cell(row=5, column=1, value="• نوع السكن: apartment, house, commercial, warehouse")
    ws.cell(row=6, column=1, value="• نوع الضرر: minor, moderate, major, total")
    ws.cell(row=7, column=1, value="• التاريخ: استخدم تنسيق YYYY-MM-DD")
    ws.cell(row=8, column=1, value="• رقم الهوية: 9 أرقام بالضبط")
    
    # حفظ الملف في response
    filename = "نموذج_الأضرار.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
