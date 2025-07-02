from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.urls import reverse
from .models import MedicalRecord
from basic_data.models import District, Guardian
from .forms import MedicalRecordForm, MedicalRecordSearchForm, MedicalRecordImportForm
from .utils import search_medical_records
# import pandas as pd  # تعطيل مؤقتاً للتوافق مع Render
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def medical_list(request):
    """عرض قائمة البيانات المرضية مع البحث والفلترة"""
    form = MedicalRecordSearchForm(request.GET)
    medical_records = MedicalRecord.objects.all().select_related('district', 'guardian')
    
    if form.is_valid():
        # البحث المرن
        search_query = form.cleaned_data.get('search')
        if search_query:
            medical_records = search_medical_records(search_query)
        
        # فلترة حسب الجنس
        gender = form.cleaned_data.get('gender')
        if gender:
            medical_records = medical_records.filter(gender=gender)
        
        # فلترة حسب الحي
        district = form.cleaned_data.get('district')
        if district:
            medical_records = medical_records.filter(district=district)
        
        # فلترة حسب نوع المرض
        disease_type = form.cleaned_data.get('disease_type')
        if disease_type:
            medical_records = medical_records.filter(disease_type=disease_type)
        
        # فلترة حسب نوع الإعاقة
        disability_type = form.cleaned_data.get('disability_type')
        if disability_type:
            medical_records = medical_records.filter(disability_type=disability_type)
    
    # ترتيب النتائج
    medical_records = medical_records.order_by('-created_at', 'name')
    
    # التقسيم إلى صفحات
    paginator = Paginator(medical_records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_count': medical_records.count(),
    }
    
    return render(request, 'medical/medical_list.html', context)

def medical_add(request):
    """إضافة سجل مرضي جديد"""
    if request.method == 'POST':
        form = MedicalRecordForm(request.POST)
        if form.is_valid():
            medical_record = form.save()
            messages.success(request, f'تم إضافة السجل المرضي لـ {medical_record.name} بنجاح')
            return redirect('medical:medical_detail', pk=medical_record.pk)
    else:
        form = MedicalRecordForm()
    
    return render(request, 'medical/medical_form.html', {
        'form': form,
        'title': 'إضافة سجل مرضي جديد'
    })

def medical_detail(request, pk):
    """عرض تفاصيل سجل مرضي"""
    medical_record = get_object_or_404(MedicalRecord, pk=pk)
    return render(request, 'medical/medical_detail.html', {
        'medical_record': medical_record
    })

def medical_edit(request, pk):
    """تعديل بيانات سجل مرضي"""
    medical_record = get_object_or_404(MedicalRecord, pk=pk)
    
    if request.method == 'POST':
        form = MedicalRecordForm(request.POST, instance=medical_record)
        if form.is_valid():
            form.save()
            messages.success(request, f'تم تحديث السجل المرضي لـ {medical_record.name} بنجاح')
            return redirect('medical:medical_detail', pk=pk)
    else:
        form = MedicalRecordForm(instance=medical_record)
    
    return render(request, 'medical/medical_form.html', {
        'form': form,
        'medical_record': medical_record,
        'title': f'تعديل السجل المرضي: {medical_record.name}'
    })

def medical_delete(request, pk):
    """حذف سجل مرضي"""
    medical_record = get_object_or_404(MedicalRecord, pk=pk)
    
    if request.method == 'POST':
        patient_name = medical_record.name
        medical_record.delete()
        messages.success(request, f'تم حذف السجل المرضي لـ {patient_name} بنجاح')
        return redirect('medical:medical_list')
    
    return render(request, 'medical/medical_delete.html', {
        'medical_record': medical_record
    })

def medical_import(request):
    """استيراد البيانات المرضية من ملف Excel"""
    if request.method == 'POST':
        form = MedicalRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['excel_file']
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['name', 'national_id', 'gender', 'guardian_national_id', 
                                  'disease_type']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'medical/medical_import.html', {'form': form})
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # البحث عن ولي الأمر
                        guardian = None
                        if pd.notna(row['guardian_national_id']):
                            try:
                                guardian = Guardian.objects.get(national_id=str(row['guardian_national_id']).zfill(9))
                            except Guardian.DoesNotExist:
                                pass
                        
                        medical_data = {
                            'name': row['name'],
                            'national_id': str(row['national_id']).zfill(9),
                            'gender': row['gender'],
                            'guardian_national_id': str(row['guardian_national_id']).zfill(9),
                            'disease_type': row['disease_type'],
                            'guardian': guardian,
                            'guardian_name': guardian.name if guardian else '',
                            'phone_number': row.get('phone_number', guardian.phone_number if guardian else ''),
                            'disease_description': row.get('disease_description', ''),
                            'disability_type': row.get('disability_type', 'none'),
                            'disability_percentage': row.get('disability_percentage', None),
                            'district': guardian.district if guardian else None,
                            'notes': row.get('notes', '')
                        }
                        
                        # التحقق من عدم وجود رقم الهوية مسبقاً
                        if not MedicalRecord.objects.filter(national_id=medical_data['national_id']).exists():
                            MedicalRecord.objects.create(**medical_data)
                            success_count += 1
                        else:
                            errors.append(f'الصف {index + 2}: رقم الهوية {medical_data["national_id"]} موجود مسبقاً')
                            error_count += 1
                            
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                        error_count += 1
                
                if success_count > 0:
                    messages.success(request, f'تم استيراد {success_count} سجل مرضي بنجاح')
                
                if error_count > 0:
                    error_msg = f'فشل في استيراد {error_count} سجل. الأخطاء:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_msg)
                
                return redirect('medical:medical_list')
                
            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء قراءة الملف: {str(e)}')
    else:
        form = MedicalRecordImportForm()
    
    return render(request, 'medical/medical_import.html', {'form': form})

def medical_ajax_search(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) >= 2:
        medical_records = search_medical_records(query)[:10]
        results = [
            {
                'id': record.id,
                'name': record.name,
                'national_id': record.national_id,
                'disease_type': record.get_disease_type_display(),
                'district': record.district.name if record.district else ''
            }
            for record in medical_records
        ]
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})

def guardian_lookup_ajax(request):
    """البحث عن ولي الأمر عبر AJAX"""
    national_id = request.GET.get('national_id', '')
    if len(national_id) == 9:
        try:
            guardian = Guardian.objects.get(national_id=national_id)
            return JsonResponse({
                'found': True,
                'name': guardian.name,
                'phone': guardian.phone_number or '',
                'district_id': guardian.district.id if guardian.district else None,
                'district_name': guardian.district.name if guardian.district else ''
            })
        except Guardian.DoesNotExist:
            return JsonResponse({'found': False})
    return JsonResponse({'found': False})

def export_medical_excel(request):
    """تصدير البيانات الطبية إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة  
    form = MedicalRecordSearchForm(request.GET)
    records = MedicalRecord.objects.all().select_related('guardian', 'guardian__district')
    
    if form.is_valid():
        search_query = form.cleaned_data.get('search')
        if search_query:
            records = search_medical_records(search_query)
        
        disease_type = form.cleaned_data.get('disease_type')
        if disease_type:
            records = records.filter(disease_type=disease_type)
            
        disability_type = form.cleaned_data.get('disability_type') 
        if disability_type:
            records = records.filter(disability_type=disability_type)
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "البيانات الطبية"
    
    # العناوين
    headers = [
        'اسم المريض', 'رقم الهوية', 'الجنس', 'تاريخ الميلاد', 'اسم ولي الأمر',
        'رقم هوية ولي الأمر', 'نوع المرض', 'وصف المرض', 'نوع الإعاقة',
        'نسبة الإعاقة', 'الأدوية', 'الحي', 'رقم الهاتف', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.name)
        ws.cell(row=row, column=2, value=record.national_id)
        ws.cell(row=row, column=3, value=record.get_gender_display())
        ws.cell(row=row, column=4, value=record.birth_date.strftime('%Y-%m-%d') if record.birth_date else "")
        ws.cell(row=row, column=5, value=record.guardian.name if record.guardian else record.guardian_name)
        ws.cell(row=row, column=6, value=record.guardian.national_id if record.guardian else record.guardian_national_id)
        ws.cell(row=row, column=7, value=record.get_disease_type_display())
        ws.cell(row=row, column=8, value=record.disease_description or "")
        ws.cell(row=row, column=9, value=record.get_disability_type_display())
        ws.cell(row=row, column=10, value=f"{record.disability_percentage or 0}%")
        ws.cell(row=row, column=11, value=record.medications or "")
        ws.cell(row=row, column=12, value=record.guardian.district.name if record.guardian and record.guardian.district else "")
        ws.cell(row=row, column=13, value=record.phone_number or "")
        ws.cell(row=row, column=14, value=record.notes or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = f"البيانات_الطبية_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_medical_excel(request):
    """استيراد البيانات الطبية من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # تحديد الأعمدة التي يجب قراءتها كنص لتجنب مشكلة الأرقام العشرية
            dtype_dict = {
                'رقم الهوية': str,
                'رقم هوية ولي الأمر': str,
                'رقم الهاتف': str
            }
            
            df = pd.read_excel(excel_file, dtype=dtype_dict)
            
            def clean_national_id(value):
                """تنظيف رقم الهوية وإضافة الأصفار إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                if cleaned.isdigit() and len(cleaned) < 9:
                    cleaned = cleaned.zfill(9)
                
                return cleaned
            
            def clean_phone_number(value):
                """تنظيف رقم الجوال وإضافة الصفر إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                if cleaned.isdigit() and len(cleaned) == 9 and cleaned.startswith('5'):
                    cleaned = '0' + cleaned
                
                return cleaned
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات المطلوبة
                    if pd.isna(row.get('اسم المريض')) or pd.isna(row.get('رقم الهوية')):
                        errors.append(f"الصف {index + 2}: اسم المريض ورقم الهوية مطلوبان")
                        error_count += 1
                        continue
                    
                    # البحث عن ولي الأمر إذا كان موجود
                    guardian = None
                    if not pd.isna(row.get('رقم هوية ولي الأمر')):
                        try:
                            guardian = Guardian.objects.get(national_id=str(row['رقم هوية ولي الأمر']).strip())
                        except Guardian.DoesNotExist:
                            pass
                    
                    # إنشاء السجل الطبي
                    record = MedicalRecord.objects.create(
                        name=str(row['اسم المريض']).strip(),
                        national_id=str(row['رقم الهوية']).strip(),
                        gender=str(row.get('الجنس', 'M')).strip().upper() if not pd.isna(row.get('الجنس')) else 'M',
                        birth_date=pd.to_datetime(row['تاريخ الميلاد']).date() if not pd.isna(row.get('تاريخ الميلاد')) else None,
                        guardian=guardian,
                        guardian_name=str(row.get('اسم ولي الأمر', '')).strip() if not pd.isna(row.get('اسم ولي الأمر')) else '',
                        guardian_national_id=str(row.get('رقم هوية ولي الأمر', '')).strip() if not pd.isna(row.get('رقم هوية ولي الأمر')) else '',
                        disease_type=str(row.get('نوع المرض', 'chronic')).strip().lower() if not pd.isna(row.get('نوع المرض')) else 'chronic',
                        disease_description=str(row.get('وصف المرض', '')).strip() if not pd.isna(row.get('وصف المرض')) else '',
                        disability_type=str(row.get('نوع الإعاقة', 'none')).strip().lower() if not pd.isna(row.get('نوع الإعاقة')) else 'none',
                        disability_percentage=float(row.get('نسبة الإعاقة', 0)) if not pd.isna(row.get('نسبة الإعاقة')) else 0,
                        medications=str(row.get('الأدوية', '')).strip() if not pd.isna(row.get('الأدوية')) else '',
                        phone_number=str(row.get('رقم الهاتف', '')).strip() if not pd.isna(row.get('رقم الهاتف')) else '',
                        notes=str(row.get('ملاحظات', '')).strip() if not pd.isna(row.get('ملاحظات')) else '',
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} سجل طبي بنجاح")
            
            if error_count > 0:
                messages.warning(request, f"فشل في استيراد {error_count} صف. الأخطاء: {'; '.join(errors[:5])}")
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('medical:medical_list')

def download_medical_template(request):
    """تحميل نموذج Excel فارغ للبيانات الطبية"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج البيانات الطبية"
    
    # العناوين
    headers = [
        'اسم المريض', 'رقم الهوية', 'الجنس', 'تاريخ الميلاد', 'اسم ولي الأمر',
        'رقم هوية ولي الأمر', 'نوع المرض', 'وصف المرض', 'نوع الإعاقة',
        'نسبة الإعاقة', 'الأدوية', 'رقم الهاتف', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'أحمد محمد', '123456789', 'M', '2000-01-15', 'محمد أحمد',
        '987654321', 'chronic', 'السكري', 'physical', '25',
        'الأنسولين', '0591234567', 'يحتاج متابعة'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = "نموذج_البيانات_الطبية.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
