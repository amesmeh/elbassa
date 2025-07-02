from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, permission_required
import json
from .models import District, Guardian, Wife, Child, Representative
from .utils import search_guardians, search_children, search_wives, search_districts
from utils.excel_utils_simple import (
    export_guardians_to_excel, export_children_to_excel, export_districts_to_excel, export_wives_to_excel
)
from django.utils import timezone
from django.db import models
from .forms import DistrictForm, GuardianForm, WifeForm, ChildForm, GuardianSearchForm
from .permissions import get_user_district, user_has_district_permission

@login_required
def home(request):
    """الصفحة الرئيسية مع الإحصائيات"""
    # Filter statistics based on user's district if not superuser
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = Guardian.objects.filter(district=request.user.representative.district)
        children = Child.objects.filter(guardian__district=request.user.representative.district)
        districts = District.objects.filter(id=request.user.representative.district.id)
    else:
        guardians = Guardian.objects.all()
        children = Child.objects.all()
        districts = District.objects.all()

    context = {
        'total_guardians': guardians.count(),
        'total_children': children.count(),
        'total_districts': districts.count(),
        'total_families': guardians.aggregate(
            total=Sum('family_members_count')
        )['total'] or 0,
        'recent_guardians': guardians.order_by('-created_at')[:6]
    }
    return render(request, 'home.html', context)

@login_required
def basic_data_home(request):
    """الصفحة الرئيسية للبيانات الأساسية"""
    # Filter data based on user's district if not superuser
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = Guardian.objects.filter(district=request.user.representative.district)
        children = Child.objects.filter(guardian__district=request.user.representative.district)
        wives = Wife.objects.filter(guardian__district=request.user.representative.district)
        districts = District.objects.filter(id=request.user.representative.district.id)
    else:
        guardians = Guardian.objects.all()
        children = Child.objects.all()
        wives = Wife.objects.all()
        districts = District.objects.all()

    context = {
        'total_guardians': guardians.count(),
        'total_children': children.count(),
        'total_wives': wives.count(),
        'total_districts': districts.count(),
        'recent_guardians': guardians.order_by('-created_at')[:6],
        'districts_stats': districts.annotate(
            guardians_count=Count('guardian')
        ).order_by('-guardians_count')[:5]
    }
    return render(request, 'basic_data/basic_data_home.html', context)

@login_required
def guardian_list(request):
    """عرض قائمة أولياء الأمور مع البحث والتصفية"""
    guardians = Guardian.objects.all().order_by('-created_at')
    
    # تصفية أولياء الأمور حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_user_district(request.user)
        if district:
            guardians = guardians.filter(
                Q(district=district) |  # أولياء الأمور في الحي
                Q(district__isnull=True)  # أولياء الأمور غير المحددين
            )
    
    # البحث
    search_query = request.GET.get('q', '').strip()
    search_type = request.GET.get('search_type', 'name')
    
    if search_query:
        if search_type == 'id':
            guardians = guardians.filter(national_id__icontains=search_query)
        elif search_type == 'phone':
            guardians = guardians.filter(phone_number__icontains=search_query)
        else:  # name
            # تقسيم النص إلى كلمات للبحث
            terms = search_query.split()
            query = Q()
            
            # إنشاء query يتطلب وجود كل الكلمات (AND)
            for term in terms:
                query &= Q(name__icontains=term)
            
            guardians = guardians.filter(query)
    
    # فلترة حسب حالة الإقامة
    selected_residence_status = request.GET.get('residence_status', '')
    if selected_residence_status:
        guardians = guardians.filter(residence_status=selected_residence_status)
    
    # فلاتر إضافية من GET parameters
    wives_count_min = request.GET.get('wives_count_min', '')
    wives_count_max = request.GET.get('wives_count_max', '')
    
    if wives_count_min:
        try:
            guardians = guardians.filter(wives_count__gte=int(wives_count_min))
        except ValueError:
            pass
    
    if wives_count_max:
        try:
            guardians = guardians.filter(wives_count__lte=int(wives_count_max))
        except ValueError:
            pass
    
    # الفلاتر المتقدمة
    selected_district = request.GET.get('district', '')
    selected_gender = request.GET.get('gender', '')
    family_size_min = request.GET.get('family_size_min', '')
    family_size_max = request.GET.get('family_size_max', '')
    selected_marital_status = request.GET.get('marital_status', '')
    selected_housing_type = request.GET.get('housing_type', '')
    
    # تطبيق فلتر الحي
    if selected_district:
        try:
            district_id = int(selected_district)
            guardians = guardians.filter(district_id=district_id)
        except ValueError:
            pass
    
    # تطبيق فلتر الجنس
    if selected_gender:
        guardians = guardians.filter(gender=selected_gender)
    
    # تطبيق فلتر عدد أفراد الأسرة
    if family_size_min:
        try:
            guardians = guardians.filter(family_members_count__gte=int(family_size_min))
        except ValueError:
            pass
    
    if family_size_max:
        try:
            guardians = guardians.filter(family_members_count__lte=int(family_size_max))
        except ValueError:
            pass
    
    # تطبيق فلتر الحالة الاجتماعية
    if selected_marital_status:
        guardians = guardians.filter(marital_status=selected_marital_status)
    
    # تطبيق فلتر نوع السكن
    if selected_housing_type:
        guardians = guardians.filter(housing_type=selected_housing_type)
    
    # إحصائيات
    try:
        total_records = guardians.count()
        total_guardians = Guardian.objects.count()
        
        # حساب إجمالي أفراد العائلات
        total_family_members = Guardian.objects.aggregate(
            total=models.Sum('family_members_count')
        )['total'] or 0
        
        # عدد المقيمين والنازحين
        residents_count = Guardian.objects.filter(residence_status='مقيم').count()
        displaced_count = Guardian.objects.filter(residence_status='نازح').count()
        
    except Exception as e:
        total_records = 0
        total_guardians = 0
        total_family_members = 0
        residents_count = 0
        displaced_count = 0
        messages.error(request, f'خطأ في تحميل الإحصائيات: {str(e)}')
    
    # الحصول على قائمة الأحياء المتاحة
    districts = District.objects.all().order_by('name')
    
    # التصفح
    paginator = Paginator(guardians, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_records': total_records,
        'total_guardians': total_guardians,
        'total_family_members': total_family_members,
        'residents_count': residents_count,
        'displaced_count': displaced_count,
        'districts': districts,
        'title': 'أولياء الأمور',
        'can_edit': request.user.is_superuser,
        'can_delete': request.user.is_superuser,
        'search_query': search_query,
        'search_type': search_type,
        'wives_count_min': wives_count_min,
        'wives_count_max': wives_count_max,
        'total_count': total_guardians,  # إضافة لاستخدامه في نوافذ الحذف
        # قيم الفلاتر المتقدمة
        'selected_district': selected_district,
        'selected_gender': selected_gender,
        'family_size_min': family_size_min,
        'family_size_max': family_size_max, 
        'selected_marital_status': selected_marital_status,
        'selected_housing_type': selected_housing_type,
        'selected_residence_status': selected_residence_status,
    }
    return render(request, 'basic_data/guardian_list.html', context)

@login_required
def guardian_detail(request, pk):
    """عرض تفاصيل ولي أمر"""
    guardian = get_object_or_404(Guardian, pk=pk)
    
    # التحقق من صلاحية الوصول للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_user_district(request.user)
        if district and guardian.district and guardian.district != district:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذا الولي')
            return redirect('basic_data:guardians_list')
    
    context = {
        'guardian': guardian,
        'title': f'تفاصيل ولي الأمر - {guardian.name}',
        'can_edit': request.user.is_superuser,
        'can_delete': request.user.is_superuser,
    }
    return render(request, 'basic_data/guardian_detail.html', context)

@login_required
@permission_required('basic_data.add_guardian', raise_exception=True)
def guardian_create(request):
    """إضافة ولي أمر جديد - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بإضافة أولياء أمور جدد')
        return redirect('basic_data:guardians_list')
    
    if request.method == 'POST':
        form = GuardianForm(request.POST)
        if form.is_valid():
            guardian = form.save()
            messages.success(request, f'تم إضافة ولي الأمر "{guardian.name}" بنجاح')
            return redirect('basic_data:guardian_detail', pk=guardian.pk)
    else:
        form = GuardianForm()
    
    # الحصول على خيارات الحالة الاجتماعية من النموذج
    marital_choices = Guardian.MARITAL_STATUS_CHOICES
    districts = District.objects.all().order_by('name')
    
    context = {
        'form': form,
        'title': 'إضافة ولي أمر جديد',
        'marital_choices': marital_choices,
        'districts': districts,
    }
    return render(request, 'basic_data/guardian_form.html', context)

@login_required
@permission_required('basic_data.change_guardian', raise_exception=True)
def guardian_edit(request, pk):
    """تعديل ولي أمر - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بتعديل أولياء الأمور')
        return redirect('basic_data:guardians_list')
    
    guardian = get_object_or_404(Guardian, pk=pk)
    
    if request.method == 'POST':
        form = GuardianForm(request.POST, instance=guardian)
        if form.is_valid():
            guardian = form.save()
            messages.success(request, f'تم تحديث بيانات ولي الأمر "{guardian.name}" بنجاح')
            return redirect('basic_data:guardian_detail', pk=guardian.pk)
    else:
        form = GuardianForm(instance=guardian)
    
    # الحصول على خيارات الحالة الاجتماعية من النموذج
    marital_choices = Guardian.MARITAL_STATUS_CHOICES
    districts = District.objects.all().order_by('name')
    
    # الحصول على الزوجات والأبناء المربوطين
    current_wives = guardian.wives.all()
    current_children = guardian.children.all()
    
    context = {
        'form': form,
        'title': 'تعديل ولي أمر',
        'guardian': guardian,
        'marital_choices': marital_choices,
        'districts': districts,
        'current_wives': current_wives,
        'current_children': current_children,
        'is_edit': True,
    }
    return render(request, 'basic_data/guardian_form.html', context)

@login_required
@permission_required('basic_data.delete_guardian', raise_exception=True)
def guardian_delete(request, pk):
    """حذف ولي أمر - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بحذف أولياء الأمور')
        return redirect('basic_data:guardians_list')
    
    guardian = get_object_or_404(Guardian, pk=pk)
    
    if request.method == 'POST':
        guardian_name = guardian.name
        guardian.delete()
        messages.success(request, f'تم حذف ولي الأمر "{guardian_name}" بنجاح')
        return redirect('basic_data:guardians_list')
    
    context = {
        'guardian': guardian,
        'title': 'حذف ولي أمر'
    }
    return render(request, 'basic_data/guardian_confirm_delete.html', context)

@login_required
def district_create(request):
    """إضافة حي جديد"""
    # فقط المشرفين يمكنهم إضافة أحياء جديدة
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بإضافة أحياء جديدة')
        return redirect('basic_data:districts_list')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        representative_name = request.POST.get('representative_name')
        representative_phone = request.POST.get('representative_phone')
        
        if name and representative_name and representative_phone:
            try:
                district = District.objects.create(
                    name=name,
                    representative_name=representative_name,
                    representative_phone=representative_phone
                )
                messages.success(request, f'تم إضافة الحي "{district.name}" بنجاح')
                return redirect('basic_data:districts_list')
            except Exception as e:
                messages.error(request, f'حدث خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى ملء جميع الحقول المطلوبة')
    
    return render(request, 'basic_data/district_form.html')

@login_required
def district_detail(request, pk):
    """عرض تفاصيل الحي"""
    # Check if user has access to this district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_object_or_404(District, pk=pk, id=request.user.representative.district.id)
    else:
        district = get_object_or_404(District, pk=pk)
    
    guardians = Guardian.objects.filter(district=district)
    
    context = {
        'district': district,
        'guardians': guardians,
        'guardians_count': guardians.count()
    }
    return render(request, 'basic_data/district_detail.html', context)

@login_required
def district_edit(request, pk):
    """تعديل الحي"""
    # Check if user has access to this district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_object_or_404(District, pk=pk, id=request.user.representative.district.id)
    else:
        district = get_object_or_404(District, pk=pk)
    
    if request.method == 'POST':
        district.name = request.POST.get('name', district.name)
        district.representative_name = request.POST.get('representative_name', district.representative_name)
        district.representative_phone = request.POST.get('representative_phone', district.representative_phone)
        
        try:
            district.save()
            messages.success(request, f'تم تحديث الحي "{district.name}" بنجاح')
            return redirect('basic_data:district_detail', pk=district.pk)
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    
    context = {'district': district}
    return render(request, 'basic_data/district_form.html', context)

@login_required
def district_delete(request, pk):
    """حذف الحي"""
    # فقط المشرفين يمكنهم حذف الأحياء
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بحذف الأحياء')
        return redirect('basic_data:districts_list')
    
    district = get_object_or_404(District, pk=pk)
    
    if request.method == 'POST':
        district_name = district.name
        district.delete()
        messages.success(request, f'تم حذف الحي "{district_name}" بنجاح')
        return redirect('basic_data:districts_list')
    
    context = {'district': district}
    return render(request, 'basic_data/district_confirm_delete.html', context)

@login_required
def wives_list(request):
    """عرض قائمة الزوجات"""
    search_query = request.GET.get('search', '')
    search_type = request.GET.get('search_type', 'exact')
    
    # استخدام البحث المرن مع نوع البحث
    if search_query:
        wives = search_wives(search_query, search_type)
    else:
        wives = Wife.objects.select_related('guardian').all()
    
    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        wives = wives.filter(guardian__district=request.user.representative.district)
    
    # Pagination
    paginator = Paginator(wives, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'search_type': search_type,
        'total_count': wives.count()
    }
    return render(request, 'basic_data/wives_list.html', context)

@login_required
def wife_detail(request, wife_id):
    """عرض تفاصيل زوجة"""
    # Check if user has access to this wife's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        wife = get_object_or_404(Wife, id=wife_id, guardian__district=request.user.representative.district)
    else:
        wife = get_object_or_404(Wife, id=wife_id)
    
    return render(request, 'basic_data/wife_detail.html', {'wife': wife})

@login_required
def wife_create(request):
    """إضافة زوجة جديدة"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            national_id = request.POST.get('national_id') if request.POST.get('national_id') else None
            guardian_id = request.POST.get('guardian_id')
            
            if not name or not guardian_id:
                messages.error(request, 'يرجى إدخال اسم الزوجة واختيار ولي الأمر')
                return render(request, 'basic_data/wife_form.html')
            
            # Check if user has access to this guardian's district
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                guardian = get_object_or_404(Guardian, pk=guardian_id, district=request.user.representative.district)
            else:
                guardian = get_object_or_404(Guardian, pk=guardian_id)
            
            # التحقق من عدم تكرار رقم الهوية
            if national_id and Wife.objects.filter(national_id=national_id).exists():
                messages.error(request, 'رقم الهوية مسجل مسبقاً لزوجة أخرى')
                return render(request, 'basic_data/wife_form.html')
            
            wife = Wife.objects.create(
                name=name,
                national_id=national_id,
                guardian=guardian
            )
            
            messages.success(request, f'تم إضافة الزوجة "{wife.name}" بنجاح')
            return redirect('basic_data:wife_detail', wife_id=wife.pk)
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء إضافة الزوجة: {str(e)}')
            return render(request, 'basic_data/wife_form.html')
    
    return render(request, 'basic_data/wife_form.html')

@login_required
def wife_edit(request, wife_id):
    """تعديل بيانات زوجة"""
    # Check if user has access to this wife's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        wife = get_object_or_404(Wife, id=wife_id, guardian__district=request.user.representative.district)
    else:
        wife = get_object_or_404(Wife, id=wife_id)
    
    if request.method == 'POST':
        form = WifeForm(request.POST, instance=wife)
        if form.is_valid():
            wife = form.save(commit=False)
            # Check if user has access to the selected guardian's district
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                if wife.guardian.district != request.user.representative.district:
                    messages.error(request, 'لا يمكنك نقل الزوجة إلى ولي أمر من حي آخر')
                    return redirect('basic_data:wife_detail', wife_id=wife.id)
            wife.save()
            messages.success(request, f'تم تحديث بيانات الزوجة "{wife.name}" بنجاح')
            return redirect('basic_data:wife_detail', wife_id=wife.id)
    else:
        form = WifeForm(instance=wife)
        
        # Filter guardians by district for representatives
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            form.fields['guardian'].queryset = Guardian.objects.filter(district=request.user.representative.district)
    
    context = {
        'form': form,
        'wife': wife
    }
    return render(request, 'basic_data/wife_form.html', context)

@login_required
def wife_delete(request, wife_id):
    """حذف زوجة"""
    # Check if user has access to this wife's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        wife = get_object_or_404(Wife, id=wife_id, guardian__district=request.user.representative.district)
    else:
        wife = get_object_or_404(Wife, id=wife_id)
    
    if request.method == 'POST':
        wife_name = wife.name
        wife.delete()
        messages.success(request, f'تم حذف الزوجة "{wife_name}" بنجاح')
        return redirect('basic_data:wives_list')
    
    context = {'wife': wife}
    return render(request, 'basic_data/wife_confirm_delete.html', context)

@login_required
def children_list(request):
    """عرض قائمة الأبناء"""
    search_query = request.GET.get('search', '')
    search_type = request.GET.get('search_type', 'exact')
    
    # استخدام البحث المرن مع نوع البحث
    if search_query:
        children = search_children(search_query, search_type)
    else:
        children = Child.objects.select_related('guardian').all()
    
    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        children = children.filter(guardian__district=request.user.representative.district)
    
    # Pagination
    paginator = Paginator(children, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'search_type': search_type,
        'total_count': children.count()
    }
    return render(request, 'basic_data/children_list.html', context)

@login_required
def child_detail(request, child_id):
    """عرض تفاصيل طفل"""
    # Check if user has access to this child's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        child = get_object_or_404(Child, id=child_id, guardian__district=request.user.representative.district)
    else:
        child = get_object_or_404(Child, id=child_id)
    
    return render(request, 'basic_data/child_detail.html', {'child': child})

@login_required
def child_create(request):
    """إضافة طفل جديد"""
    context = {'title': 'إضافة طفل جديد'}
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            birth_date = request.POST.get('birth_date')
            gender = request.POST.get('gender')
            national_id = request.POST.get('national_id') if request.POST.get('national_id') else None
            guardian_id = request.POST.get('guardian_id')
            
            if not all([name, birth_date, gender, guardian_id]):
                messages.error(request, 'يرجى إدخال جميع البيانات المطلوبة')
                return render(request, 'basic_data/child_form.html', context)
            
            # Check if user has access to this guardian's district
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                guardian = get_object_or_404(Guardian, pk=guardian_id, district=request.user.representative.district)
            else:
                guardian = get_object_or_404(Guardian, pk=guardian_id)
            
            # التحقق من عدم تكرار رقم الهوية
            if national_id and Child.objects.filter(national_id=national_id).exists():
                messages.error(request, 'رقم الهوية مسجل مسبقاً لطفل آخر')
                return render(request, 'basic_data/child_form.html', context)
            
            child = Child.objects.create(
                name=name,
                birth_date=birth_date,
                gender=gender,
                national_id=national_id,
                guardian=guardian
            )
            
            messages.success(request, f'تم إضافة الطفل "{child.name}" بنجاح')
            return redirect('basic_data:child_detail', child_id=child.pk)
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء إضافة الطفل: {str(e)}')
            return render(request, 'basic_data/child_form.html', context)
    
    return render(request, 'basic_data/child_form.html', context)

@login_required
def child_edit(request, child_id):
    """تعديل بيانات طفل"""
    # Check if user has access to this child's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        child = get_object_or_404(Child, id=child_id, guardian__district=request.user.representative.district)
    else:
        child = get_object_or_404(Child, id=child_id)
    
    if request.method == 'POST':
        form = ChildForm(request.POST, instance=child)
        if form.is_valid():
            child = form.save(commit=False)
            # Check if user has access to the selected guardian's district
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                if child.guardian.district != request.user.representative.district:
                    messages.error(request, 'لا يمكنك نقل الطفل إلى ولي أمر من حي آخر')
                    return redirect('basic_data:child_detail', child_id=child.id)
            child.save()
            messages.success(request, f'تم تحديث بيانات الطفل "{child.name}" بنجاح')
            return redirect('basic_data:child_detail', child_id=child.id)
    else:
        form = ChildForm(instance=child)
        
        # Filter guardians by district for representatives
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            form.fields['guardian'].queryset = Guardian.objects.filter(district=request.user.representative.district)
    
    context = {
        'form': form,
        'child': child
    }
    return render(request, 'basic_data/child_form.html', context)

@login_required
def child_delete(request, child_id):
    """حذف طفل"""
    # Check if user has access to this child's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        child = get_object_or_404(Child, id=child_id, guardian__district=request.user.representative.district)
    else:
        child = get_object_or_404(Child, id=child_id)
    
    if request.method == 'POST':
        child_name = child.name
        child.delete()
        messages.success(request, f'تم حذف الطفل "{child_name}" بنجاح')
        return redirect('basic_data:children_list')
    
    context = {'child': child}
    return render(request, 'basic_data/child_confirm_delete.html', context)

@login_required
def export_guardians_excel(request):
    """تصدير أولياء الأمور إلى Excel مع تطبيق الفلاتر"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    search_query = request.GET.get('q', '')
    search_type = request.GET.get('search_type', 'name')
    district_filter = request.GET.get('district', '')
    gender_filter = request.GET.get('gender', '')
    marital_status_filter = request.GET.get('marital_status', '')
    residence_status_filter = request.GET.get('residence_status', '')
    family_size_min = request.GET.get('family_size_min', '')
    family_size_max = request.GET.get('family_size_max', '')
    
    # تطبيق نفس منطق الفلترة المستخدم في guardians_list
    if search_query:
        guardians = search_guardians(search_query, search_type)
    else:
        guardians = Guardian.objects.select_related('district').prefetch_related('wives').all()
    
    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = guardians.filter(district=request.user.representative.district)
    
    # الفلاتر الأساسية
    if district_filter and request.user.is_superuser:
        try:
            guardians = guardians.filter(district_id=int(district_filter))
        except (ValueError, TypeError):
            pass
    
    if gender_filter:
        guardians = guardians.filter(gender=gender_filter)
    
    if marital_status_filter:
        guardians = guardians.filter(marital_status=marital_status_filter)
    
    if residence_status_filter:
        guardians = guardians.filter(residence_status=residence_status_filter)
    
    # فلتر عدد أفراد العائلة
    if family_size_min:
        try:
            guardians = guardians.filter(family_members_count__gte=int(family_size_min))
        except (ValueError, TypeError):
            pass
    
    if family_size_max:
        try:
            guardians = guardians.filter(family_members_count__lte=int(family_size_max))
        except (ValueError, TypeError):
            pass
    
    guardians = guardians.order_by('name')
    
    # استخدام وظيفة التصدير المحسنة من utils
    try:
        from utils.excel_utils import export_guardians_to_excel
        return export_guardians_to_excel(guardians)
    except ImportError:
        # نظام بديل في حالة عدم وجود utils
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        from django.utils import timezone
        
        # إعداد البيانات للتصدير
        data = []
        for guardian in guardians:
            # تجميع أسماء الزوجات مع أرقام هوياتهن
            wives_info = []
            for wife in guardian.wives.all():
                wife_text = wife.name
                if wife.national_id:
                    wife_text += f" ({wife.national_id})"
                wives_info.append(wife_text)
            
            wives_text = " - ".join(wives_info) if wives_info else "لا توجد"
            
            # تحويل القيم للعربية
            gender_arabic = "ذكر" if guardian.gender == "ذكر" else "أنثى"
            marital_status_arabic = guardian.marital_status
            residence_status_arabic = guardian.residence_status
            
            row_data = {
                'اسم ولي الأمر': guardian.name,
                'رقم الهوية': guardian.national_id,
                'الجنس': gender_arabic,
                'الوظيفة': guardian.current_job or '',
                'رقم الجوال': guardian.phone_number,
                'الحالة الاجتماعية': marital_status_arabic,
                'عدد الأبناء': guardian.children_count,
                'عدد الزوجات': guardian.wives_count,
                'أسماء الزوجات وأرقام هوياتهن': wives_text,
                'عدد أفراد العائلة': guardian.family_members_count,
                'حالة الإقامة': residence_status_arabic,
                'الحي': guardian.district.name if guardian.district else '',
                'المحافظة الأصلية': guardian.original_governorate or '',
                'المدينة الأصلية': guardian.original_city or '',
                'عنوان النزوح': guardian.displacement_address or '',
                'تاريخ الإنشاء': guardian.created_at.strftime('%Y-%m-%d %H:%M:%S') if guardian.created_at else '',
            }
            data.append(row_data)
        
        df = pd.DataFrame(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='أولياء الأمور', index=False)
        
        output.seek(0)
        
        # إنشاء الاستجابة
        filename = f"guardians_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

@login_required
def guardian_aids_summary(request, pk):
    """عرض ملخص المساعدات التي تلقاها ولي الأمر"""
    # Check if user has access to this guardian's district
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardian = get_object_or_404(Guardian, pk=pk, district=request.user.representative.district)
    else:
        guardian = get_object_or_404(Guardian, pk=pk)
    
    # معلومات تشخيصية تفصيلية
    debug_info = {
        'guardian_pk': guardian.pk,
        'guardian_name': guardian.name,
        'guardian_national_id': guardian.national_id,
        'aids_available': False,
        'error_message': None,
        'query_results': []
    }
    
    # استيراد نماذج المساعدات
    try:
        from aids.models import AidDistribution, AidBeneficiary
        debug_info['aids_available'] = True
        
        # التحقق من أسماء الحقول في النماذج
        aid_distribution_fields = [f.name for f in AidDistribution._meta.get_fields()]
        aid_beneficiary_fields = [f.name for f in AidBeneficiary._meta.get_fields()]
        
        debug_info['aid_distribution_fields'] = aid_distribution_fields
        debug_info['aid_beneficiary_fields'] = aid_beneficiary_fields
        
        # البحث المباشر باستخدام pk
        distributions_by_pk = AidDistribution.objects.filter(guardian_id=guardian.pk).select_related('aid_type')
        beneficiaries_by_pk = AidBeneficiary.objects.filter(guardian_id=guardian.pk).select_related('aid_type')
        
        debug_info['distributions_by_pk_count'] = distributions_by_pk.count()
        debug_info['beneficiaries_by_pk_count'] = beneficiaries_by_pk.count()
        
        # البحث باستخدام الكائن - هذا هو الطريقة الصحيحة
        distributions_by_obj = AidDistribution.objects.filter(guardian=guardian).select_related('aid_type')
        beneficiaries_by_obj = AidBeneficiary.objects.filter(guardian=guardian).select_related('aid_type')
        
        debug_info['distributions_by_obj_count'] = distributions_by_obj.count()
        debug_info['beneficiaries_by_obj_count'] = beneficiaries_by_obj.count()
        
        # البحث بالاسم ورقم الهوية (للمقارنة فقط)
        distributions_by_name = AidDistribution.objects.filter(
            guardian__national_id=guardian.national_id
        ).select_related('aid_type', 'guardian')
        
        beneficiaries_by_name = AidBeneficiary.objects.filter(
            guardian__national_id=guardian.national_id
        ).select_related('aid_type', 'guardian')
        
        debug_info['distributions_by_name_count'] = distributions_by_name.count()
        debug_info['beneficiaries_by_name_count'] = beneficiaries_by_name.count()
        
        # عرض تفاصيل المساعدات المكتشفة بالاسم/رقم الهوية (للتشخيص)
        for dist in distributions_by_name:
            debug_info['query_results'].append({
                'type': 'distribution',
                'aid_type': dist.aid_type.name,
                'guardian_in_record': {
                    'pk': dist.guardian.pk,
                    'name': dist.guardian.name,
                    'national_id': dist.guardian.national_id
                },
                'date': dist.distribution_date,
                'matches_current_guardian': dist.guardian.pk == guardian.pk
            })
        
        for ben in beneficiaries_by_name:
            debug_info['query_results'].append({
                'type': 'beneficiary',
                'aid_type': ben.aid_type.name,
                'guardian_in_record': {
                    'pk': ben.guardian.pk,
                    'name': ben.guardian.name,
                    'national_id': ben.guardian.national_id
                },
                'date': ben.aid_date,
                'matches_current_guardian': ben.guardian.pk == guardian.pk
            })
        
        # استخدام النتائج الصحيحة - البحث بالكائن المباشر
        distributions = distributions_by_obj.order_by('-distribution_date')
        beneficiaries = beneficiaries_by_obj.order_by('-aid_date')
        
        # جمع جميع المساعدات في قائمة واحدة
        all_aids = []
        
        # إضافة التوزيعات
        for dist in distributions:
            all_aids.append({
                'type': 'distribution',
                'aid_type': dist.aid_type.name,
                'aid_category': dist.aid_type.get_category_display(),
                'date': dist.distribution_date,
                'notes': dist.notes or '',
                'is_transferred': dist.is_transferred
            })
        
        # إضافة المستفيدين
        for ben in beneficiaries:
            all_aids.append({
                'type': 'beneficiary',
                'aid_type': ben.aid_type.name,
                'aid_category': ben.aid_type.get_category_display(),
                'date': ben.aid_date,
                'notes': ben.notes or '',
                'is_transferred': True
            })
        
        # ترتيب حسب التاريخ (الأحدث أولاً)
        all_aids.sort(key=lambda x: x['date'], reverse=True)
        
        # إحصائيات
        total_aids_count = len(all_aids)
        distributions_count = distributions.count()
        beneficiaries_count = beneficiaries.count()
        
        # آخر مساعدة
        last_aid_date = all_aids[0]['date'] if all_aids else None
        
        # عدد مرات الاستفادة حسب نوع المساعدة
        aid_types_stats = {}
        for aid in all_aids:
            aid_type = aid['aid_type']
            if aid_type not in aid_types_stats:
                aid_types_stats[aid_type] = {
                    'count': 0,
                    'last_date': None,
                    'category': aid['aid_category']
                }
            aid_types_stats[aid_type]['count'] += 1
            if not aid_types_stats[aid_type]['last_date'] or aid['date'] > aid_types_stats[aid_type]['last_date']:
                aid_types_stats[aid_type]['last_date'] = aid['date']
        
        # تحويل إلى قائمة مرتبة
        aid_types_list = []
        for aid_type, stats in aid_types_stats.items():
            aid_types_list.append({
                'name': aid_type,
                'category': stats['category'],
                'count': stats['count'],
                'last_date': stats['last_date']
            })
        aid_types_list.sort(key=lambda x: x['last_date'], reverse=True)
        
        aids_available = True
        
    except ImportError as e:
        debug_info['error_message'] = f"ImportError: {str(e)}"
        all_aids = []
        total_aids_count = 0
        distributions_count = 0
        beneficiaries_count = 0
        last_aid_date = None
        aid_types_list = []
        aids_available = False
    except Exception as e:
        debug_info['error_message'] = f"Exception: {str(e)}"
        all_aids = []
        total_aids_count = 0
        distributions_count = 0
        beneficiaries_count = 0
        last_aid_date = None
        aid_types_list = []
        aids_available = False
    
    context = {
        'guardian': guardian,
        'all_aids': all_aids,
        'total_aids_count': total_aids_count,
        'distributions_count': distributions_count,
        'beneficiaries_count': beneficiaries_count,
        'last_aid_date': last_aid_date,
        'aid_types_list': aid_types_list,
        'aids_available': aids_available,
        'debug_info': debug_info  # معلومات التشخيص
    }
    
    return render(request, 'basic_data/guardian_aids_summary.html', context)

def check_aids_data_integrity(request):
    """فحص سلامة بيانات المساعدات للتأكد من عدم وجود تضارب"""
    if not request.user.is_staff:
        messages.error(request, 'غير مصرح لك بالوصول لهذه الصفحة')
        return redirect('basic_data:basic_data_home')
    
    try:
        from aids.models import AidDistribution, AidBeneficiary
        
        # البحث عن تكرارات في البيانات
        issues = []
        
        # التحقق من وجود توزيعات لأولياء أمور غير موجودين
        invalid_distributions = AidDistribution.objects.select_related('guardian').filter(guardian__isnull=True)
        if invalid_distributions.exists():
            issues.append(f"وجد {invalid_distributions.count()} توزيعات بدون ولي أمر محدد")
        
        # التحقق من وجود مستفيدين لأولياء أمور غير موجودين
        invalid_beneficiaries = AidBeneficiary.objects.select_related('guardian').filter(guardian__isnull=True)
        if invalid_beneficiaries.exists():
            issues.append(f"وجد {invalid_beneficiaries.count()} مستفيدين بدون ولي أمر محدد")
        
        # البحث عن البيانات المتضاربة (نفس الشخص بأسماء مختلفة)
        guardians_with_aids = Guardian.objects.filter(
            models.Q(distributions__isnull=False) | models.Q(aid_benefits__isnull=False)
        ).distinct()
        
        # فحص كل ولي أمر
        for guardian in guardians_with_aids:
            distributions = AidDistribution.objects.filter(guardian=guardian)
            beneficiaries = AidBeneficiary.objects.filter(guardian=guardian)
            
            # فحص التوزيعات
            for dist in distributions:
                if dist.guardian.name != guardian.name:
                    issues.append(f"تضارب في الاسم: التوزيع {dist.pk} يشير لـ {dist.guardian.name} لكن ولي الأمر {guardian.pk} اسمه {guardian.name}")
                
                if dist.guardian.national_id != guardian.national_id:
                    issues.append(f"تضارب في رقم الهوية: التوزيع {dist.pk} يشير لرقم {dist.guardian.national_id} لكن ولي الأمر {guardian.pk} رقمه {guardian.national_id}")
            
            # فحص المستفيدين
            for ben in beneficiaries:
                if ben.guardian.name != guardian.name:
                    issues.append(f"تضارب في الاسم: المستفيد {ben.pk} يشير لـ {ben.guardian.name} لكن ولي الأمر {guardian.pk} اسمه {guardian.name}")
                
                if ben.guardian.national_id != guardian.national_id:
                    issues.append(f"تضارب في رقم الهوية: المستفيد {ben.pk} يشير لرقم {ben.guardian.national_id} لكن ولي الأمر {guardian.pk} رقمه {guardian.national_id}")
        
        # البحث عن أولياء أمور لهم نفس رقم الهوية
        from django.db.models import Count
        duplicate_ids = Guardian.objects.values('national_id').annotate(
            count=Count('id')
        ).filter(count__gt=1)
        
        for dup in duplicate_ids:
            guardians_with_same_id = Guardian.objects.filter(national_id=dup['national_id'])
            names = [g.name for g in guardians_with_same_id]
            issues.append(f"رقم الهوية {dup['national_id']} مكرر لدى: {', '.join(names)}")
        
        # عرض النتائج
        if issues:
            for issue in issues:
                messages.warning(request, issue)
            messages.error(request, f"تم العثور على {len(issues)} مشكلة في بيانات المساعدات")
        else:
            messages.success(request, "لا توجد مشاكل في بيانات المساعدات")
            
    except ImportError:
        messages.error(request, "تطبيق المساعدات غير متوفر")
    except Exception as e:
        messages.error(request, f"حدث خطأ أثناء فحص البيانات: {str(e)}")
    
    return redirect('basic_data:basic_data_home')

@login_required
def search_guardian_aids(request):
    """البحث عن ولي أمر ومعرفة مساعداته - للتشخيص"""
    if request.method == 'GET':
        search_query = request.GET.get('q', '').strip()
        
        if not search_query:
            messages.warning(request, 'يرجى إدخال اسم ولي الأمر أو رقم الهوية للبحث')
            return redirect('basic_data:basic_data_home')
        
        # البحث عن ولي الأمر
        guardians = Guardian.objects.filter(
            models.Q(name__icontains=search_query) | 
            models.Q(national_id__icontains=search_query)
        )
        
        # تصفية حسب الحي للمندوبين
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            guardians = guardians.filter(district=request.user.representative.district)
        
        if not guardians.exists():
            messages.error(request, f'لم يتم العثور على ولي أمر باسم أو رقم هوية: "{search_query}"')
            return redirect('basic_data:basic_data_home')
        
        if guardians.count() == 1:
            # إذا وُجد ولي أمر واحد فقط، انتقل مباشرة لصفحة مساعداته
            guardian = guardians.first()
            return redirect('basic_data:guardian_aids_summary', pk=guardian.pk)
        else:
            # إذا وُجد أكثر من ولي أمر، اعرض قائمة للاختيار
            context = {
                'guardians': guardians,
                'search_query': search_query,
                'total_found': guardians.count()
            }
            return render(request, 'basic_data/search_guardian_results.html', context)
    
    return redirect('basic_data:basic_data_home')

@login_required
def search_aids_by_name(request):
    """البحث المباشر في جداول المساعدات بالاسم أو رقم الهوية"""
    if not request.user.is_staff:
        messages.error(request, 'غير مصرح لك بالوصول لهذه الصفحة')
        return redirect('basic_data:basic_data_home')
    
    search_query = request.GET.get('q', '').strip()
    results = []
    
    if search_query:
        try:
            from aids.models import AidDistribution, AidBeneficiary
            
            # البحث في جدول التوزيعات
            distributions = AidDistribution.objects.select_related('guardian', 'aid_type').filter(
                models.Q(guardian__name__icontains=search_query) |
                models.Q(guardian__national_id__icontains=search_query)
            )
            
            # البحث في جدول المستفيدين
            beneficiaries = AidBeneficiary.objects.select_related('guardian', 'aid_type').filter(
                models.Q(guardian__name__icontains=search_query) |
                models.Q(guardian__national_id__icontains=search_query)
            )
            
            # تصفية حسب الحي للمندوبين
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                distributions = distributions.filter(guardian__district=request.user.representative.district)
                beneficiaries = beneficiaries.filter(guardian__district=request.user.representative.district)
            
            distributions = distributions.order_by('-distribution_date')
            beneficiaries = beneficiaries.order_by('-aid_date')
            
            # تجميع النتائج
            for dist in distributions:
                results.append({
                    'type': 'distribution',
                    'aid_type': dist.aid_type.name,
                    'guardian_name': dist.guardian.name,
                    'guardian_id': dist.guardian.national_id,
                    'guardian_pk': dist.guardian.pk,
                    'date': dist.distribution_date,
                    'notes': dist.notes or '',
                    'is_transferred': dist.is_transferred,
                    'record_id': dist.pk
                })
            
            for ben in beneficiaries:
                results.append({
                    'type': 'beneficiary',
                    'aid_type': ben.aid_type.name,
                    'guardian_name': ben.guardian.name,
                    'guardian_id': ben.guardian.national_id,
                    'guardian_pk': ben.guardian.pk,
                    'date': ben.aid_date,
                    'notes': ben.notes or '',
                    'is_transferred': True,
                    'record_id': ben.pk
                })
            
            # ترتيب حسب التاريخ
            results.sort(key=lambda x: x['date'], reverse=True)
            
            # إحصائيات
            total_distributions = distributions.count()
            total_beneficiaries = beneficiaries.count()
            
            # البحث عن أولياء أمور بنفس الاسم للمقارنة
            similar_guardians = Guardian.objects.filter(
                models.Q(name__icontains=search_query) |
                models.Q(national_id__icontains=search_query)
            ).order_by('name')
            
            context = {
                'search_query': search_query,
                'results': results,
                'total_results': len(results),
                'total_distributions': total_distributions,
                'total_beneficiaries': total_beneficiaries,
                'similar_guardians': similar_guardians,
            }
            
            return render(request, 'basic_data/aids_search_results.html', context)
            
        except ImportError:
            messages.error(request, 'تطبيق المساعدات غير متوفر')
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء البحث: {str(e)}')
    
    return render(request, 'basic_data/aids_search_form.html', {'search_query': search_query})

@require_POST
@login_required
def delete_selected_guardians(request):
    """حذف أولياء الأمور المحددين"""
    # التحقق من صلاحية المشرف فقط
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'غير مصرح لك بحذف أولياء الأمور'})
    
    try:
        import json
        data = json.loads(request.body)
        guardian_ids = data.get('guardian_ids', [])
        
        if not guardian_ids:
            return JsonResponse({'success': False, 'error': 'لم يتم تحديد أي عناصر للحذف'})
        
        # التحقق من وجود أولياء الأمور
        guardians = Guardian.objects.filter(pk__in=guardian_ids)
        deleted_count = guardians.count()
        
        if deleted_count == 0:
            return JsonResponse({'success': False, 'error': 'لم يتم العثور على العناصر المحددة'})
        
        # حذف أولياء الأمور (سيتم حذف البيانات المرتبطة تلقائياً)
        guardians.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'تم حذف {deleted_count} من أولياء الأمور بنجاح'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'خطأ في تنسيق البيانات'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})

@require_POST
@login_required
def delete_all_guardians(request):
    """حذف جميع أولياء الأمور"""
    # التحقق من صلاحية المشرف فقط
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'غير مصرح لك بحذف أولياء الأمور'})
    
    try:
        # عدد أولياء الأمور قبل الحذف
        total_count = Guardian.objects.count()
        
        if total_count == 0:
            return JsonResponse({'success': False, 'error': 'لا توجد بيانات للحذف'})
        
        # حذف جميع أولياء الأمور
        Guardian.objects.all().delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'تم حذف جميع أولياء الأمور ({total_count}) بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})

@require_POST
@login_required
def delete_selected_children(request):
    """حذف الأطفال المحددين"""
    try:
        child_ids = request.POST.getlist('child_ids')
        if not child_ids:
            return JsonResponse({
                'success': False,
                'message': 'لم يتم تحديد أي أطفال للحذف'
            })
        
        # الحصول على الأطفال المحددين
        children = Child.objects.filter(id__in=child_ids)
        
        # تصفية حسب الحي للمندوبين
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            children = children.filter(guardian__district=request.user.representative.district)
        
        deleted_count = children.count()
        
        if deleted_count == 0:
            return JsonResponse({
                'success': False,
                'message': 'لم يتم العثور على الأطفال المحددين'
            })
        
        # جمع معرفات أولياء الأمور لتحديث العدادات
        guardian_ids = set(children.values_list('guardian_id', flat=True))
        
        # حذف الأطفال
        children.delete()
        
        # لا نحدث العدادات - نحتفظ بقيم Excel
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'تم حذف {deleted_count} من الأطفال بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@require_POST
@login_required
def delete_all_children(request):
    """حذف جميع الأطفال"""
    try:
        # عد جميع الأطفال
        children = Child.objects.all()
        
        # تصفية حسب الحي للمندوبين
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            children = children.filter(guardian__district=request.user.representative.district)
        
        total_count = children.count()
        
        if total_count == 0:
            return JsonResponse({
                'success': False,
                'message': 'لا يوجد أطفال للحذف'
            })
        
        # جمع معرفات أولياء الأمور لتحديث العدادات
        guardian_ids = set(children.values_list('guardian_id', flat=True))
        
        # حذف جميع الأطفال
        children.delete()
        
        # لا نحدث العدادات - نحتفظ بقيم Excel
        
        return JsonResponse({
            'success': True,
            'deleted_count': total_count,
            'message': f'تم حذف جميع الأطفال ({total_count}) بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@login_required
def fix_guardian_counters(request):
    """إصلاح عدادات أولياء الأمور"""
    if request.method == 'GET':
        # عرض صفحة التأكيد
        guardians = Guardian.objects.all()
        
        # تصفية حسب الحي للمندوبين
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            guardians = guardians.filter(district=request.user.representative.district)
        
        guardians_count = guardians.count()
        return render(request, 'basic_data/fix_guardian_counters.html', {
            'guardians_count': guardians_count
        })
    
    elif request.method == 'POST':
        try:
            fixed_count = 0
            
            # إعادة حساب العدادات لجميع أولياء الأمور
            guardians = Guardian.objects.all()
            
            # تصفية حسب الحي للمندوبين
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                guardians = guardians.filter(district=request.user.representative.district)
            
            for guardian in guardians:
                # إعادة حساب عدد الأطفال
                children_count = guardian.children.count()
                
                # إعادة حساب عدد الزوجات
                wives_count = guardian.wives.count()
                
                # إعادة حساب إجمالي أفراد العائلة (ولي الأمر + الأطفال + الزوجات)
                family_members_count = 1 + children_count + wives_count
                
                # تحديث العدادات إذا كانت مختلفة
                if (guardian.children_count != children_count or 
                    guardian.wives_count != wives_count or 
                    guardian.family_members_count != family_members_count):
                    
                    guardian.children_count = children_count
                    guardian.wives_count = wives_count
                    guardian.family_members_count = family_members_count
                    guardian.save()
                    fixed_count += 1
            
            messages.success(request, f'تم إصلاح عدادات {fixed_count} ولي أمر بنجاح')
            
        except Exception as e:
            messages.error(request, f'خطأ في إصلاح العدادات: {str(e)}')
            
            return redirect('basic_data:guardians_list')

@login_required
def update_counters_from_excel(request):
    """تحديث العدادات لتتطابق مع ملف Excel المستورد"""
    if request.method == 'GET':
        guardians = Guardian.objects.all()
        
        # تصفية حسب الحي للمندوبين
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            guardians = guardians.filter(district=request.user.representative.district)
        
        guardians_count = guardians.count()
        return render(request, 'basic_data/update_counters_from_excel.html', {
            'guardians_count': guardians_count
        })
    
    elif request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                messages.error(request, 'يرجى اختيار ملف Excel')
                return redirect('basic_data:update_counters_from_excel')
            
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'يرجى اختيار ملف Excel بصيغة .xlsx')
                return redirect('basic_data:update_counters_from_excel')
            
            import pandas as pd
            from utils.excel_utils import clean_national_id
            
            # قراءة ملف Excel
            df = pd.read_excel(excel_file, sheet_name=0)
            
            updated_count = 0
            not_found_count = 0
            
            for index, row in df.iterrows():
                # التحقق من البيانات المطلوبة
                if pd.isna(row['رقم الهوية']):
                    continue
                
                national_id = clean_national_id(row['رقم الهوية'])
                
                try:
                    guardian = Guardian.objects.get(national_id=national_id)
                    
                    # تصفية حسب الحي للمندوبين
                    if not request.user.is_superuser and hasattr(request.user, 'representative'):
                        if guardian.district != request.user.representative.district:
                            continue
                    
                    # الحصول على العدادات من ملف Excel
                    children_count_excel = int(row['عدد الأبناء']) if not pd.isna(row['عدد الأبناء']) else 0
                    wives_count_excel = int(row['عدد الزوجات']) if not pd.isna(row['عدد الزوجات']) else 0
                    family_members_count_excel = 1 + children_count_excel + wives_count_excel
                    
                    # تحديث العدادات
                    guardian.children_count = children_count_excel
                    guardian.wives_count = wives_count_excel
                    guardian.family_members_count = family_members_count_excel
                    guardian.save()
                    
                    updated_count += 1
                    
                except Guardian.DoesNotExist:
                    not_found_count += 1
                    continue
            
            if updated_count > 0:
                messages.success(request, f'تم تحديث عدادات {updated_count} ولي أمر بنجاح')
            
            if not_found_count > 0:
                messages.warning(request, f'لم يتم العثور على {not_found_count} ولي أمر')
            
        except Exception as e:
            messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    
            return redirect('basic_data:guardians_list')

@login_required
def duplicate_guardians_report(request):
    """تقرير أولياء الأمور المكررين - حيث الرجل وزوجته كلاهما مسجل كولي أمر"""
    duplicates = []
    
    # فلاتر الطلب
    marital_status_filter = request.GET.get('marital_status', '')
    wives_count_filter = request.GET.get('wives_count', '')
    show_data_errors = request.GET.get('show_data_errors', '')
    
    # البحث عن أولياء الأمور الذين لديهم زوجات
    guardians_with_wives = Guardian.objects.filter(wives__isnull=False).distinct()
    
    # تطبيق فلتر الحالة الاجتماعية
    if marital_status_filter:
        guardians_with_wives = guardians_with_wives.filter(marital_status=marital_status_filter)
    
    # تطبيق فلتر عدد الزوجات
    if wives_count_filter:
        try:
            wives_count = int(wives_count_filter)
            guardians_with_wives = guardians_with_wives.filter(wives_count=wives_count)
        except ValueError:
            pass
    
    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians_with_wives = guardians_with_wives.filter(district=request.user.representative.district)
    
    data_errors = []  # لتجميع أخطاء البيانات
    
    for guardian in guardians_with_wives:
        for wife in guardian.wives.all():
            # التحقق من أخطاء البيانات
            has_data_error = False
            error_details = []
            
            # خطأ: نفس رقم الهوية للزوج والزوجة
            if wife.national_id and wife.national_id == guardian.national_id:
                has_data_error = True
                error_details.append('نفس رقم هوية الزوج')
            
            # البحث عن الزوجة في قاعدة بيانات أولياء الأمور
            wife_as_guardian_query = Guardian.objects.all()
            
            # تصفية حسب الحي للمندوبين
            if not request.user.is_superuser and hasattr(request.user, 'representative'):
                wife_as_guardian_query = wife_as_guardian_query.filter(district=request.user.representative.district)
            
            if wife.national_id:
                # البحث برقم الهوية، مع استثناء ولي الأمر الحالي
                wife_as_guardian = wife_as_guardian_query.filter(
                    national_id=wife.national_id
                ).exclude(id=guardian.id).first()
            else:
                # البحث بالاسم إذا لم يكن هناك رقم هوية، مع استثناء ولي الأمر الحالي
                wife_as_guardian = wife_as_guardian_query.filter(
                    name__icontains=wife.name
                ).exclude(id=guardian.id).first()
            
            # تجميع أخطاء البيانات
            if has_data_error:
                data_errors.append({
                    'guardian': guardian,
                    'wife_record': wife,
                    'errors': error_details
                })
            
            # التأكد من أن الزوجة المُوجدة ليست نفس ولي الأمر وأنها أنثى
            if (wife_as_guardian and 
                wife_as_guardian.id != guardian.id and 
                wife_as_guardian.gender == 'أنثى'):
                
                duplicate_entry = {
                    'husband_guardian': guardian,
                    'wife_record': wife,
                    'wife_guardian': wife_as_guardian,
                    'husband_id': guardian.national_id,
                    'wife_id': wife.national_id or 'غير محدد',
                    'match_type': 'رقم الهوية' if wife.national_id else 'الاسم',
                    'has_data_error': has_data_error,
                    'error_details': error_details
                }
                
                # تطبيق فلتر أخطاء البيانات
                if show_data_errors == 'only' and not has_data_error:
                    continue
                elif show_data_errors == 'exclude' and has_data_error:
                    continue
                    
                duplicates.append(duplicate_entry)
    
    # إحصائيات
    total_duplicates = len(duplicates)
    id_matches = len([d for d in duplicates if d['match_type'] == 'رقم الهوية'])
    name_matches = len([d for d in duplicates if d['match_type'] == 'الاسم'])
    
    context = {
        'duplicates': duplicates,
        'total_duplicates': total_duplicates,
        'id_matches': id_matches,
        'name_matches': name_matches,
        'data_errors': data_errors,
        'data_errors_count': len(data_errors),
        'title': 'تقرير أولياء الأمور المكررين',
        'marital_status_filter': marital_status_filter,
        'wives_count_filter': wives_count_filter,
        'show_data_errors': show_data_errors,
    }
    
    return render(request, 'basic_data/duplicate_guardians_report.html', context)

@login_required
def districts_list(request):
    """عرض قائمة المناطق والأحياء"""
    districts = District.objects.all()
    
    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        districts = districts.filter(id=request.user.representative.district.id)
    
    # البحث
    search_query = request.GET.get('search', '')
    if search_query:
        districts = districts.filter(
            models.Q(name__icontains=search_query) |
            models.Q(representative_name__icontains=search_query) |
            models.Q(representative_phone__icontains=search_query)
        )
    
    # الترتيب
    districts = districts.order_by('name')
    
    # Pagination
    paginator = Paginator(districts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_count': districts.count(),
    }
    
    return render(request, 'basic_data/districts_list.html', context)

@login_required
def download_guardians_template(request):
    """تحميل نموذج Excel لأولياء الأمور"""
    try:
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # إنشاء ملف Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="guardians_import_template.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نموذج أولياء الأمور"
        
        # إضافة العناوين
        headers = [
            'اسم ولي الأمر', 'رقم الهوية', 'الجنس', 'الوظيفة', 'رقم الجوال',
            'الحالة الاجتماعية', 'عدد الأبناء', 'عدد الزوجات',
            'حالة الإقامة', 'المحافظة الأصلية', 'المدينة الأصلية', 'عنوان النزوح', 'الحي'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # إضافة بيانات نموذجية
        sample_data = [
            ['محمد أحمد علي', '123456789', 'ذكر', 'مهندس', '0599123456', 'متزوج', 3, 1, 'مقيم', 'غزة', 'غزة', '', 'الرمال'],
            ['فاطمة خالد سالم', '987654321', 'أنثى', 'معلمة', '0599876543', 'أرملة', 2, 0, 'نازح', 'خان يونس', 'خان يونس', 'مخيم النصيرات', 'النصيرات']
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'خطأ في إنشاء النموذج: {str(e)}')
        return redirect('basic_data:guardians_list')

@login_required
def download_children_template(request):
    """تحميل نموذج Excel للأبناء"""
    messages.warning(request, 'هذه الميزة غير متاحة حالياً على الخادم')
    return redirect('basic_data:children_list')

@login_required
def download_wives_template(request):
    """تحميل نموذج Excel للزوجات"""
    messages.warning(request, 'هذه الميزة غير متاحة حالياً على الخادم')
    return redirect('basic_data:wives_list')

@login_required
def download_districts_template(request):
    """تحميل قالب الأحياء"""
    messages.warning(request, 'هذه الميزة غير متاحة حالياً على الخادم')
    return redirect('basic_data:districts_list')

@login_required
def export_guardians_pdf(request):
    """تصدير أولياء الأمور إلى PDF"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object, using the buffer as its "file"
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Get guardians data
    guardians = Guardian.objects.all()
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_user_district(request.user)
        if district:
            guardians = guardians.filter(district=district)
    
    # Apply search filters from request
    search_query = request.GET.get('search', '')
    if search_query:
        from .utils.search_utils import search_guardians
        guardians = search_guardians(search_query, 'exact')
    
    # Write PDF content
    y_position = height - 50
    p.drawString(50, y_position, "Guardians Report")
    y_position -= 30
    
    for guardian in guardians[:50]:  # Limit to first 50 records
        text = f"{guardian.name} - {guardian.national_id} - {guardian.district.name if guardian.district else 'N/A'}"
        p.drawString(50, y_position, text)
        y_position -= 20
        if y_position < 50:
            p.showPage()
            y_position = height - 50
    
    # Close the PDF object cleanly
    p.save()
    
    # Get the value of the BytesIO buffer and write it to the response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="guardians_report.pdf"'
    return response

@login_required
def download_import_errors(request):
    """تحميل ملف أخطاء الاستيراد"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        
        # الحصول على الأخطاء من الجلسة
        detailed_errors = request.session.get('detailed_errors', [])
        
        if not detailed_errors:
            # إنشاء رسالة خطأ افتراضية
            error_data = [{
                'رقم الصف': 1,
                'الاسم': 'لا توجد أخطاء',
                'رقم الهوية': '',
                'نوع الخطأ': 'لا توجد أخطاء',
                'تفاصيل الخطأ': 'لا توجد أخطاء استيراد حالياً'
            }]
        else:
            error_data = detailed_errors
        
        df = pd.DataFrame(error_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='أخطاء الاستيراد', index=False)
        
        output.seek(0)
        
        # إنشاء الاستجابة
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="import_errors.xlsx"'
        
        return response
        
    except Exception as e:
        # نظام بديل في حالة الخطأ
        from django.http import HttpResponse
        import json
        
        errors = {
            'message': 'لا توجد أخطاء استيراد حالياً',
            'timestamp': str(timezone.now()),
            'user': request.user.username
        }
        
        response = HttpResponse(
            json.dumps(errors, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="import_errors.json"'
        return response

@login_required 
def import_guardians_excel(request):
    """استيراد أولياء الأمور من Excel"""
    messages.warning(request, 'ميزة الاستيراد غير متاحة حالياً على الخادم')
    return redirect('basic_data:guardians_list')

@login_required
def export_wives_excel(request):
    """تصدير الزوجات إلى Excel"""
    try:
        # Get all wives based on filters
        wives = Wife.objects.select_related('guardian', 'guardian__district').all()
        
        # Apply user district filter if not superuser
        if not request.user.is_superuser and hasattr(request.user, 'representative'):
            wives = wives.filter(guardian__district=request.user.representative.district)
        
        # Apply search filters
        search_query = request.GET.get('search', '')
        if search_query:
            wives = wives.filter(
                models.Q(name__icontains=search_query) |
                models.Q(national_id__icontains=search_query) |
                models.Q(guardian__name__icontains=search_query) |
                models.Q(guardian__national_id__icontains=search_query)
            )
        
        # Export using utils function
        return export_wives_to_excel(wives)
        
    except Exception as e:
        messages.error(request, f'حدث خطأ في التصدير: {str(e)}')
        return redirect('basic_data:wives_list')

@login_required
def import_wives_excel(request):
    """استيراد الزوجات من Excel"""
    messages.warning(request, 'ميزة الاستيراد غير متاحة حالياً على الخادم')
    return redirect('basic_data:wives_list')

@login_required
def cancel_migration(request):
    """إلغاء عملية الترحيل"""
    messages.info(request, 'تم إلغاء عملية الترحيل')
    return redirect('basic_data:guardians_list')

@login_required
def migrate_wife_to_guardian(request, wife_id):
    """ترحيل زوجة واحدة إلى ولي أمر جديد"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بترحيل البيانات')
        return redirect('basic_data:wives_list')
    
    wife = get_object_or_404(Wife, pk=wife_id)
    
    if request.method == 'POST':
        new_guardian_id = request.POST.get('new_guardian_id')
        if new_guardian_id:
            try:
                new_guardian = Guardian.objects.get(pk=new_guardian_id)
                old_guardian = wife.guardian
                wife.guardian = new_guardian
                wife.save()
                
                # Update counters
                old_guardian.update_counters()
                new_guardian.update_counters()
                
                messages.success(request, f'تم ترحيل الزوجة {wife.name} بنجاح')
                return redirect('basic_data:guardian_detail', pk=new_guardian.pk)
            except Guardian.DoesNotExist:
                messages.error(request, 'ولي الأمر المختار غير موجود')
        else:
            messages.error(request, 'يرجى اختيار ولي أمر')
    
    # Get available guardians for migration
    guardians = Guardian.objects.exclude(pk=wife.guardian.pk)
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = guardians.filter(district=request.user.representative.district)
    
    context = {
        'wife': wife,
        'guardians': guardians[:50],  # Limit for performance
        'title': f'ترحيل الزوجة {wife.name}'
    }
    return render(request, 'basic_data/migrate_wife.html', context)

@login_required
def migrate_all_wives_to_guardian(request, guardian_id):
    """ترحيل جميع زوجات ولي أمر إلى ولي أمر جديد"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بترحيل البيانات')
        return redirect('basic_data:guardians_list')
    
    guardian = get_object_or_404(Guardian, pk=guardian_id)
    
    if request.method == 'POST':
        new_guardian_id = request.POST.get('new_guardian_id')
        if new_guardian_id:
            try:
                new_guardian = Guardian.objects.get(pk=new_guardian_id)
                wives_count = guardian.wives.count()
                
                # Migrate all wives
                guardian.wives.all().update(guardian=new_guardian)
                
                # Update counters
                guardian.update_counters()
                new_guardian.update_counters()
                
                messages.success(request, f'تم ترحيل {wives_count} زوجة بنجاح')
                return redirect('basic_data:guardian_detail', pk=new_guardian.pk)
            except Guardian.DoesNotExist:
                messages.error(request, 'ولي الأمر المختار غير موجود')
        else:
            messages.error(request, 'يرجى اختيار ولي أمر')
    
    # Get available guardians for migration
    guardians = Guardian.objects.exclude(pk=guardian.pk)
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = guardians.filter(district=request.user.representative.district)
    
    context = {
        'guardian': guardian,
        'wives_count': guardian.wives.count(),
        'guardians': guardians[:50],  # Limit for performance
        'title': f'ترحيل جميع زوجات {guardian.name}'
    }
    return render(request, 'basic_data/migrate_all_wives.html', context)

@login_required
def migrate_child_to_guardian(request, child_id):
    """ترحيل طفل واحد إلى ولي أمر جديد"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بترحيل البيانات')
        return redirect('basic_data:children_list')
    
    child = get_object_or_404(Child, pk=child_id)
    
    if request.method == 'POST':
        new_guardian_id = request.POST.get('new_guardian_id')
        if new_guardian_id:
            try:
                new_guardian = Guardian.objects.get(pk=new_guardian_id)
                old_guardian = child.guardian
                child.guardian = new_guardian
                child.save()
                
                # Update counters
                old_guardian.update_counters()
                new_guardian.update_counters()
                
                messages.success(request, f'تم ترحيل الطفل {child.name} بنجاح')
                return redirect('basic_data:guardian_detail', pk=new_guardian.pk)
            except Guardian.DoesNotExist:
                messages.error(request, 'ولي الأمر المختار غير موجود')
        else:
            messages.error(request, 'يرجى اختيار ولي أمر')
    
    # Get available guardians for migration
    guardians = Guardian.objects.exclude(pk=child.guardian.pk)
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = guardians.filter(district=request.user.representative.district)
    
    context = {
        'child': child,
        'guardians': guardians[:50],  # Limit for performance
        'title': f'ترحيل الطفل {child.name}'
    }
    return render(request, 'basic_data/migrate_child.html', context)

@login_required
def migrate_all_children_to_guardian(request, guardian_id):
    """ترحيل جميع أطفال ولي أمر إلى ولي أمر جديد"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بترحيل البيانات')
        return redirect('basic_data:guardians_list')
    
    guardian = get_object_or_404(Guardian, pk=guardian_id)
    
    if request.method == 'POST':
        new_guardian_id = request.POST.get('new_guardian_id')
        if new_guardian_id:
            try:
                new_guardian = Guardian.objects.get(pk=new_guardian_id)
                children_count = guardian.children.count()
                
                # Migrate all children
                guardian.children.all().update(guardian=new_guardian)
                
                # Update counters
                guardian.update_counters()
                new_guardian.update_counters()
                
                messages.success(request, f'تم ترحيل {children_count} طفل بنجاح')
                return redirect('basic_data:guardian_detail', pk=new_guardian.pk)
            except Guardian.DoesNotExist:
                messages.error(request, 'ولي الأمر المختار غير موجود')
        else:
            messages.error(request, 'يرجى اختيار ولي أمر')
    
    # Get available guardians for migration
    guardians = Guardian.objects.exclude(pk=guardian.pk)
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        guardians = guardians.filter(district=request.user.representative.district)
    
    context = {
        'guardian': guardian,
        'children_count': guardian.children.count(),
        'guardians': guardians[:50],  # Limit for performance
        'title': f'ترحيل جميع أطفال {guardian.name}'
    }
    return render(request, 'basic_data/migrate_all_children.html', context)

@login_required
def export_children_excel(request):
    """تصدير الأبناء إلى Excel مع تطبيق الفلاتر"""
    search_query = request.GET.get('search', '')
    search_type = request.GET.get('search_type', 'exact')
    district_filter = request.GET.get('district', '')
    guardian_filter = request.GET.get('guardian', '')
    age_group = request.GET.get('age_group', '')
    birth_date_from = request.GET.get('birth_date_from', '')
    birth_date_to = request.GET.get('birth_date_to', '')

    # البحث
    if search_query:
        children = search_children(search_query, search_type)
    else:
        children = Child.objects.select_related('guardian').all()

    # تصفية حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        children = children.filter(guardian__district=request.user.representative.district)

    # فلترة الحي
    if district_filter:
        try:
            children = children.filter(guardian__district_id=int(district_filter))
        except (ValueError, TypeError):
            pass

    # فلترة ولي الأمر
    if guardian_filter:
        try:
            children = children.filter(guardian_id=int(guardian_filter))
        except (ValueError, TypeError):
            pass

    # فلترة العمر
    from datetime import date, timedelta
    today = date.today()
    if age_group == '0-5':
        min_date = today - timedelta(days=5*365)
        children = children.filter(birth_date__gte=min_date)
    elif age_group == '6-12':
        max_date = today - timedelta(days=6*365)
        min_date = today - timedelta(days=12*365)
        children = children.filter(birth_date__lte=max_date, birth_date__gte=min_date)
    elif age_group == '13-18':
        max_date = today - timedelta(days=13*365)
        min_date = today - timedelta(days=18*365)
        children = children.filter(birth_date__lte=max_date, birth_date__gte=min_date)
    elif age_group == '18+':
        max_date = today - timedelta(days=18*365)
        children = children.filter(birth_date__lte=max_date)

    # فلترة تاريخ الميلاد
    if birth_date_from:
        children = children.filter(birth_date__gte=birth_date_from)
    if birth_date_to:
        children = children.filter(birth_date__lte=birth_date_to)

    children = children.order_by('name')

    # استخدام وظيفة التصدير من utils
    try:
        from utils.excel_utils_simple import export_children_to_excel
        return export_children_to_excel(children)
    except ImportError:
        messages.error(request, 'ميزة التصدير غير متاحة حالياً')
        return redirect('basic_data:children_list')

@login_required
def import_children_excel(request):
    """استيراد الأبناء من ملف Excel"""
    messages.warning(request, 'ميزة الاستيراد غير متاحة حالياً على الخادم')
    return redirect('basic_data:children_list')
