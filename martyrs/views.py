from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from .models import Martyr
from basic_data.models import District
from .forms import MartyrForm, MartyrSearchForm, MartyrImportForm
from .utils import search_martyrs
import pandas as pd
from django.core.exceptions import ValidationError
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def martyrs_list(request):
    """عرض قائمة الشهداء مع البحث والفلترة"""
    form = MartyrSearchForm(request.GET)
    martyrs = Martyr.objects.all()
    
    if form.is_valid():
        # البحث المرن
        search_query = form.cleaned_data.get('search')
        if search_query:
            martyrs = search_martyrs(search_query)
        
        # فلترة حسب السنة
        year = form.cleaned_data.get('year')
        if year:
            martyrs = martyrs.filter(martyrdom_date__year=year)
    
    # ترتيب النتائج
    martyrs = martyrs.order_by('-martyrdom_date', 'name')
    
    # التقسيم إلى صفحات
    paginator = Paginator(martyrs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_count': martyrs.count(),
    }
    
    return render(request, 'martyrs/martyrs_list.html', context)

def martyr_add(request):
    """إضافة شهيد جديد"""
    if request.method == 'POST':
        form = MartyrForm(request.POST)
        if form.is_valid():
            martyr = form.save()
            messages.success(request, f'تم إضافة الشهيد {martyr.name} بنجاح')
            return redirect('martyrs:martyr_detail', pk=martyr.pk)
    else:
        form = MartyrForm()
    
    return render(request, 'martyrs/martyr_form.html', {
        'form': form,
        'title': 'إضافة شهيد جديد'
    })

def martyr_detail(request, pk):
    """عرض تفاصيل شهيد"""
    martyr = get_object_or_404(Martyr, pk=pk)
    return render(request, 'martyrs/martyr_detail.html', {
        'martyr': martyr
    })

def martyr_edit(request, pk):
    """تعديل بيانات شهيد"""
    martyr = get_object_or_404(Martyr, pk=pk)
    
    if request.method == 'POST':
        form = MartyrForm(request.POST, instance=martyr)
        if form.is_valid():
            form.save()
            messages.success(request, f'تم تحديث بيانات الشهيد {martyr.name} بنجاح')
            return redirect('martyrs:martyr_detail', pk=pk)
    else:
        form = MartyrForm(instance=martyr)
    
    return render(request, 'martyrs/martyr_form.html', {
        'form': form,
        'martyr': martyr,
        'title': f'تعديل بيانات الشهيد: {martyr.name}'
    })

def martyr_delete(request, pk):
    """حذف شهيد"""
    martyr = get_object_or_404(Martyr, pk=pk)
    
    if request.method == 'POST':
        martyr_name = martyr.name
        martyr.delete()
        messages.success(request, f'تم حذف الشهيد {martyr_name} بنجاح')
        return redirect('martyrs:martyrs_list')
    
    return render(request, 'martyrs/martyr_delete.html', {
        'martyr': martyr
    })

def martyrs_import(request):
    """استيراد الشهداء من ملف Excel"""
    if request.method == 'POST':
        form = MartyrImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['excel_file']
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['name', 'national_id', 'martyrdom_date', 'agent_name', 
                                  'agent_national_id', 'agent_phone', 'relationship_to_martyr']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'martyrs/martyrs_import.html', {'form': form})
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # البحث عن المنطقة إذا كانت موجودة
                        district = None
                        if 'district' in df.columns and pd.notna(row['district']):
                            district = District.objects.filter(name__icontains=row['district']).first()
                        
                        martyr_data = {
                            'name': row['name'],
                            'national_id': str(row['national_id']).zfill(9),
                            'martyrdom_date': pd.to_datetime(row['martyrdom_date']).date(),
                            'agent_name': row['agent_name'],
                            'agent_national_id': str(row['agent_national_id']).zfill(9),
                            'agent_phone': str(row['agent_phone']),
                            'relationship_to_martyr': row['relationship_to_martyr'],
                            'district': district,
                            'notes': row.get('notes', '')
                        }
                        
                        # التحقق من عدم وجود رقم الهوية مسبقاً
                        if not Martyr.objects.filter(national_id=martyr_data['national_id']).exists():
                            Martyr.objects.create(**martyr_data)
                            success_count += 1
                        else:
                            errors.append(f'الصف {index + 2}: رقم الهوية {martyr_data["national_id"]} موجود مسبقاً')
                            error_count += 1
                            
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                        error_count += 1
                
                if success_count > 0:
                    messages.success(request, f'تم استيراد {success_count} شهيد بنجاح')
                
                if error_count > 0:
                    error_msg = f'فشل في استيراد {error_count} سجل. الأخطاء:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_msg)
                
                return redirect('martyrs:martyrs_list')
                
            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء قراءة الملف: {str(e)}')
    else:
        form = MartyrImportForm()
    
    return render(request, 'martyrs/martyrs_import.html', {'form': form})

def martyrs_ajax_search(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) >= 2:
        martyrs = search_martyrs(query)[:10]
        results = [
            {
                'id': martyr.id,
                'name': martyr.name,
                'national_id': martyr.national_id,
                'relationship': martyr.relationship_to_martyr or ''
            }
            for martyr in martyrs
        ]
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})

def export_martyrs_excel(request):
    """تصدير بيانات الشهداء إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    form = MartyrSearchForm(request.GET)
    martyrs = Martyr.objects.all()
    
    if form.is_valid():
        search_query = form.cleaned_data.get('search')
        if search_query:
            martyrs = search_martyrs(search_query)
        

            
        year = form.cleaned_data.get('year')
        if year:
            martyrs = martyrs.filter(martyrdom_date__year=year)
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الشهداء"
    
    # العناوين - بدون المنطقة
    headers = [
        'اسم الشهيد', 'رقم الهوية', 'تاريخ الاستشهاد',
        'اسم الوكيل', 'رقم هوية الوكيل', 'صلة القرابة', 'رقم هاتف الوكيل',
        'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, martyr in enumerate(martyrs, 2):
        ws.cell(row=row, column=1, value=martyr.name)
        ws.cell(row=row, column=2, value=martyr.national_id)
        ws.cell(row=row, column=3, value=martyr.martyrdom_date.strftime('%Y-%m-%d') if martyr.martyrdom_date else "")
        ws.cell(row=row, column=4, value=martyr.agent_name or "")
        ws.cell(row=row, column=5, value=martyr.agent_national_id or "")
        ws.cell(row=row, column=6, value=martyr.relationship_to_martyr or "")
        ws.cell(row=row, column=7, value=martyr.agent_phone or "")
        ws.cell(row=row, column=8, value=martyr.notes or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = f"الشهداء_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_martyrs_excel(request):
    """استيراد بيانات الشهداء من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات المطلوبة (اسم الشهيد ورقم الهوية فقط)
                    name_key = 'اسم الشهيد*' if 'اسم الشهيد*' in row else 'اسم الشهيد'
                    id_key = 'رقم الهوية*' if 'رقم الهوية*' in row else 'رقم الهوية'
                    
                    if pd.isna(row.get(name_key)) or pd.isna(row.get(id_key)):
                        errors.append(f"الصف {index + 2}: اسم الشهيد ورقم الهوية مطلوبان")
                        error_count += 1
                        continue
                    
                    # إنشاء الشهيد (بدون منطقة)
                    martyr = Martyr.objects.create(
                        name=str(row[name_key]).strip(),
                        national_id=str(row[id_key]).strip(),
                        martyrdom_date=pd.to_datetime(row['تاريخ الاستشهاد']).date() if not pd.isna(row.get('تاريخ الاستشهاد')) else None,
                        agent_name=str(row.get('اسم الوكيل', '')).strip() if not pd.isna(row.get('اسم الوكيل')) else None,
                        agent_national_id=str(row.get('رقم هوية الوكيل', '')).strip() if not pd.isna(row.get('رقم هوية الوكيل')) else None,
                        relationship_to_martyr=str(row.get('صلة القرابة', '')).strip() if not pd.isna(row.get('صلة القرابة')) else None,
                        agent_phone=str(row.get('رقم هاتف الوكيل', '')).strip() if not pd.isna(row.get('رقم هاتف الوكيل')) else None,
                        notes=str(row.get('ملاحظات', '')).strip() if not pd.isna(row.get('ملاحظات')) else None,
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} شهيد بنجاح")
            
            if error_count > 0:
                messages.warning(request, f"فشل في استيراد {error_count} صف. الأخطاء: {'; '.join(errors[:5])}")
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('martyrs:martyrs_list')

def download_martyrs_template(request):
    """تحميل نموذج Excel فارغ للشهداء"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج الشهداء"
    
    # العناوين - بدون المنطقة، مع تحديد الحقول الإجبارية
    headers = [
        'اسم الشهيد*', 'رقم الهوية*', 'تاريخ الاستشهاد',
        'اسم الوكيل', 'رقم هوية الوكيل', 'صلة القرابة', 'رقم هاتف الوكيل',
        'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'أحمد محمد علي', '123456789', '2024-01-15',
        'محمد أحمد علي', '987654321', 'والد', '0591234567',
        'شهيد أثناء الدفاع عن الوطن'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # إضافة ملاحظات توضيحية
    ws.cell(row=4, column=1, value="ملاحظات:")
    ws.cell(row=5, column=1, value="• الحقول المؤشرة بـ (*) إجبارية")
    ws.cell(row=6, column=1, value="• صلة القرابة: أدخل النص باللغة العربية (مثل: والد، أخ، ابن، عم...)")
    ws.cell(row=7, column=1, value="• التاريخ: استخدم تنسيق YYYY-MM-DD")
    ws.cell(row=8, column=1, value="• رقم الهوية: 9 أرقام بالضبط")
    ws.cell(row=9, column=1, value="• رقم الجوال: مثال 0599123456")
    ws.cell(row=10, column=1, value="• جميع الحقول الأخرى اختيارية")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = "نموذج_الشهداء.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
