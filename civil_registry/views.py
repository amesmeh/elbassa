from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
import pandas as pd
import io
from .models import CivilRegistry
from .forms import CivilRecordForm, CivilRecordSearchForm, CivilRecordImportForm
from django.contrib.auth.decorators import login_required

# إضافة استيراد نظام التصدير
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.pdf_export import export_queryset_to_pdf
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

def civil_records_list(request):
    """عرض قائمة السجلات المدنية مع البحث والتصفية"""
    civil_records = CivilRegistry.objects.all().order_by('-created_at')
    search_form = CivilRecordSearchForm(request.GET)
    
    # تطبيق البحث والتصفية
    if search_form.is_valid():
        if search_form.cleaned_data['search_query']:
            query = search_form.cleaned_data['search_query']
            civil_records = civil_records.filter(
                Q(name__icontains=query) |
                Q(national_id__icontains=query) |
                Q(city__icontains=query) |
                Q(neighborhood__icontains=query)
            )
        
        if search_form.cleaned_data['gender']:
            civil_records = civil_records.filter(gender=search_form.cleaned_data['gender'])
        
        if search_form.cleaned_data['governorate']:
            civil_records = civil_records.filter(governorate=search_form.cleaned_data['governorate'])
        
        if search_form.cleaned_data['age_from']:
            from datetime import date
            birth_year_to = date.today().year - search_form.cleaned_data['age_from']
            civil_records = civil_records.filter(birth_date__year__lte=birth_year_to)
        
        if search_form.cleaned_data['age_to']:
            from datetime import date
            birth_year_from = date.today().year - search_form.cleaned_data['age_to']
            civil_records = civil_records.filter(birth_date__year__gte=birth_year_from)
    
    # التصفح
    paginator = Paginator(civil_records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_count': civil_records.count(),
        'title': 'السجل المدني',
        'gender_choices': CivilRegistry.GENDER_CHOICES,
        'governorate_choices': CivilRegistry.GOVERNORATE_CHOICES,
        'search_query': request.GET.get('search_query', ''),
        'gender_filter': request.GET.get('gender', ''),
        'governorate_filter': request.GET.get('governorate', ''),
    }
    return render(request, 'civil_registry/civil_registry_list.html', context)

def civil_record_detail(request, pk):
    """عرض تفاصيل سجل مدني"""
    civil_record = get_object_or_404(CivilRegistry, pk=pk)
    context = {
        'civil_record': civil_record,
        'title': f'تفاصيل السجل - {civil_record.name}'
    }
    return render(request, 'civil_registry/civil_registry_detail.html', context)

def civil_record_add(request):
    """إضافة سجل مدني جديد"""
    if request.method == 'POST':
        form = CivilRecordForm(request.POST)
        if form.is_valid():
            civil_record = form.save()
            messages.success(request, f'تم إضافة السجل المدني لـ {civil_record.name} بنجاح.')
            return redirect('civil_registry:civil_record_detail', pk=civil_record.pk)
    else:
        form = CivilRecordForm()
    
    context = {
        'form': form,
        'title': 'إضافة سجل مدني جديد'
    }
    return render(request, 'civil_registry/civil_record_form.html', context)

def civil_record_edit(request, pk):
    """تعديل سجل مدني"""
    civil_record = get_object_or_404(CivilRegistry, pk=pk)
    
    if request.method == 'POST':
        form = CivilRecordForm(request.POST, instance=civil_record)
        if form.is_valid():
            civil_record = form.save()
            messages.success(request, f'تم تحديث بيانات السجل المدني لـ {civil_record.name} بنجاح.')
            return redirect('civil_registry:civil_record_detail', pk=civil_record.pk)
    else:
        form = CivilRecordForm(instance=civil_record)
    
    context = {
        'form': form,
        'civil_record': civil_record,
        'title': f'تعديل السجل المدني - {civil_record.name}'
    }
    return render(request, 'civil_registry/civil_record_form.html', context)

@login_required
def civil_record_delete(request, pk):
    """حذف سجل مدني"""
    record = get_object_or_404(CivilRegistry, pk=pk)
    record.delete()
    messages.success(request, 'تم حذف السجل بنجاح')
    return redirect('civil_registry:civil_registry_list')

def civil_record_import(request):
    """استيراد السجلات المدنية من ملف Excel"""
    if request.method == 'POST':
        form = CivilRecordImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                
                # قراءة ملف Excel
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['name', 'national_id', 'gender', 'birth_date', 'governorate', 'city']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'civil_registry/civil_record_import.html', {'form': form})
                
                # استيراد البيانات
                imported_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # التحقق من عدم وجود السجل مسبقاً
                        if CivilRegistry.objects.filter(national_id=row['national_id']).exists():
                            errors.append(f'الصف {index + 2}: رقم الهوية موجود مسبقاً')
                            continue
                        
                        # إنشاء السجل
                        civil_record = CivilRegistry(
                            name=row['name'],
                            national_id=row['national_id'],
                            gender=row['gender'],
                            birth_date=pd.to_datetime(row['birth_date']).date(),
                            governorate=row['governorate'],
                            city=row['city'],
                            family_members_count=row.get('family_members_count', None),
                            neighborhood=row.get('neighborhood', ''),
                            address=row.get('address', ''),
                            notes=row.get('notes', '')
                        )
                        
                        civil_record.full_clean()
                        civil_record.save()
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                
                # عرض النتائج
                if imported_count > 0:
                    messages.success(request, f'تم استيراد {imported_count} سجل بنجاح.')
                
                if errors:
                    error_message = f'حدثت أخطاء في {len(errors)} صف:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_message += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_message)
                
                if imported_count > 0:
                    return redirect('civil_registry:civil_records_list')
                    
            except Exception as e:
                messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    else:
        form = CivilRecordImportForm()
    
    context = {
        'form': form,
        'title': 'استيراد السجلات المدنية'
    }
    return render(request, 'civil_registry/civil_record_import.html', context)

def civil_record_search_ajax(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    civil_records = CivilRegistry.objects.filter(
        Q(name__icontains=query) |
        Q(national_id__icontains=query) |
        Q(city__icontains=query)
    )[:10]
    
    results = []
    for record in civil_records:
        results.append({
            'id': record.id,
            'text': f'{record.name} - {record.national_id} ({record.get_governorate_display()})',
            'name': record.name,
            'national_id': record.national_id,
            'governorate': record.governorate
        })
    
    return JsonResponse({'results': results})

def search_civil_registry(query):
    """البحث المرن في السجل المدني"""
    if not query:
        return CivilRegistry.objects.none()
    
    return CivilRegistry.objects.filter(
        Q(name__icontains=query) |
        Q(national_id__icontains=query) |
        Q(city__icontains=query) |
        Q(neighborhood__icontains=query)
    )

def export_civil_registry_excel(request):
    """تصدير بيانات السجل المدني إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    form = CivilRecordSearchForm(request.GET)
    records = CivilRegistry.objects.all()
    
    if form.is_valid():
        search_query = form.cleaned_data.get('search_query')
        if search_query:
            records = search_civil_registry(search_query)
        
        if form.cleaned_data.get('gender'):
            records = records.filter(gender=form.cleaned_data['gender'])
        
        if form.cleaned_data.get('governorate'):
            records = records.filter(governorate=form.cleaned_data['governorate'])
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "السجل المدني"
    
    # العناوين الجديدة حسب الحقول المطلوبة
    headers = [
        'الاسم', 'رقم الهوية', 'الجنس', 'تاريخ الميلاد', 
        'المحافظة', 'المدينة', 'عدد أفراد الأسرة'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تحويل الجنس إلى عربي
    gender_choices = dict(CivilRegistry.GENDER_CHOICES)
    governorate_choices = dict(CivilRegistry.GOVERNORATE_CHOICES)
    
    # كتابة البيانات
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.name)
        ws.cell(row=row, column=2, value=record.national_id)
        ws.cell(row=row, column=3, value=gender_choices.get(record.gender, record.gender))
        ws.cell(row=row, column=4, value=record.birth_date.strftime('%Y-%m-%d') if record.birth_date else "")
        ws.cell(row=row, column=5, value=governorate_choices.get(record.governorate, record.governorate))
        ws.cell(row=row, column=6, value=record.city)
        ws.cell(row=row, column=7, value=record.family_members_count or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = f"السجل_المدني_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_civil_registry_excel(request):
    """استيراد بيانات السجل المدني من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            
            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['name', 'national_id', 'gender', 'birth_date', 'governorate', 'city']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                return redirect('civil_registry:civil_records_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات المطلوبة
                    if pd.isna(row.get('name')) or pd.isna(row.get('national_id')):
                        errors.append(f"الصف {index + 2}: الاسم ورقم الهوية مطلوبان")
                        error_count += 1
                        continue
                    
                    # التحقق من عدم وجود السجل مسبقاً
                    national_id = str(row['national_id']).strip()
                    if CivilRegistry.objects.filter(national_id=national_id).exists():
                        errors.append(f"الصف {index + 2}: رقم الهوية {national_id} موجود مسبقاً")
                        error_count += 1
                        continue
                    
                    # إنشاء السجل المدني
                    record = CivilRegistry.objects.create(
                        name=str(row['name']).strip(),
                        national_id=national_id,
                        gender=str(row.get('gender', 'M')).strip().upper(),
                        birth_date=pd.to_datetime(row['birth_date']).date() if not pd.isna(row.get('birth_date')) else None,
                        governorate=str(row.get('governorate', '')).strip() if not pd.isna(row.get('governorate')) else '',
                        city=str(row.get('city', '')).strip() if not pd.isna(row.get('city')) else '',
                        family_members_count=int(row.get('family_members_count', 0)) if not pd.isna(row.get('family_members_count')) and str(row.get('family_members_count', '')).strip() else None,
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} سجل بنجاح")
            
            if error_count > 0:
                error_message = f"فشل في استيراد {error_count} صف:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    error_message += f"\n... و {len(errors) - 10} أخطاء أخرى"
                messages.warning(request, error_message)
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('civil_registry:civil_records_list')

def download_civil_registry_template(request):
    """تحميل نموذج Excel فارغ للسجل المدني"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج السجل المدني"
    
    # العناوين الجديدة حسب الحقول المطلوبة
    headers = [
        'name', 'national_id', 'gender', 'birth_date', 
        'governorate', 'city', 'family_members_count'
    ]
    
    # العناوين العربية للمرجع
    arabic_headers = [
        'الاسم', 'رقم الهوية', 'الجنس (M/F)', 'تاريخ الميلاد (YYYY-MM-DD)', 
        'المحافظة', 'المدينة', 'عدد أفراد الأسرة'
    ]
    
    # كتابة العناوين الإنجليزية (للاستيراد)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة العناوين العربية كتوضيح
    for col, header in enumerate(arabic_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="333333")
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'أحمد محمد علي', '123456789', 'M', '1990-01-15', 
        'gaza', 'غزة', '5'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=3, column=col, value=data)
    
    # إضافة ملاحظات
    ws.cell(row=5, column=1, value="ملاحظات:")
    ws.cell(row=6, column=1, value="• الجنس: استخدم M للذكر و F للأنثى")
    ws.cell(row=7, column=1, value="• التاريخ: استخدم تنسيق YYYY-MM-DD")
    ws.cell(row=8, column=1, value="• المحافظة: gaza, north_gaza, middle_area, khan_younis, rafah, west_bank, jerusalem, other")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = "نموذج_السجل_المدني.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
