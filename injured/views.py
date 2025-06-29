from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from .models import Injured
from basic_data.models import District, Guardian
from .forms import InjuredForm, InjuredSearchForm, InjuredImportForm
from .utils import search_injured
import pandas as pd
from django.core.exceptions import ValidationError
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def injured_list(request):
    """عرض قائمة الجرحى مع البحث والفلترة"""
    form = InjuredSearchForm(request.GET)
    injured = Injured.objects.all().select_related('district', 'guardian')
    
    if form.is_valid():
        # البحث المرن
        search_query = form.cleaned_data.get('search')
        if search_query:
            injured = search_injured(search_query)
        
        # فلترة حسب الحي
        district = form.cleaned_data.get('district')
        if district:
            injured = injured.filter(district=district)
        
        # فلترة حسب نوع الإصابة
        injury_type = form.cleaned_data.get('injury_type')
        if injury_type:
            injured = injured.filter(injury_type=injury_type)
        
        # فلترة حسب السنة
        year = form.cleaned_data.get('year')
        if year:
            injured = injured.filter(injury_date__year=year)
    
    # ترتيب النتائج
    injured = injured.order_by('-injury_date', 'name')
    
    # التقسيم إلى صفحات
    paginator = Paginator(injured, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_count': injured.count(),
    }
    
    return render(request, 'injured/injured_list.html', context)

def injured_add(request):
    """إضافة جريح جديد"""
    if request.method == 'POST':
        form = InjuredForm(request.POST)
        if form.is_valid():
            injured = form.save()
            messages.success(request, f'تم إضافة الجريح {injured.name} بنجاح')
            return redirect('injured:injured_detail', pk=injured.pk)
    else:
        form = InjuredForm()
    
    return render(request, 'injured/injured_form.html', {
        'form': form,
        'title': 'إضافة جريح جديد'
    })

def injured_detail(request, pk):
    """عرض تفاصيل جريح"""
    injured = get_object_or_404(Injured, pk=pk)
    return render(request, 'injured/injured_detail.html', {
        'injured': injured
    })

def injured_edit(request, pk):
    """تعديل بيانات جريح"""
    injured = get_object_or_404(Injured, pk=pk)
    
    if request.method == 'POST':
        form = InjuredForm(request.POST, instance=injured)
        if form.is_valid():
            form.save()
            messages.success(request, f'تم تحديث بيانات الجريح {injured.name} بنجاح')
            return redirect('injured:injured_detail', pk=pk)
    else:
        form = InjuredForm(instance=injured)
    
    return render(request, 'injured/injured_form.html', {
        'form': form,
        'injured': injured,
        'title': f'تعديل بيانات الجريح: {injured.name}'
    })

def injured_delete(request, pk):
    """حذف جريح"""
    injured = get_object_or_404(Injured, pk=pk)
    
    if request.method == 'POST':
        injured_name = injured.name
        injured.delete()
        messages.success(request, f'تم حذف الجريح {injured_name} بنجاح')
        return redirect('injured:injured_list')
    
    return render(request, 'injured/injured_delete.html', {
        'injured': injured
    })

def injured_import(request):
    """استيراد الجرحى من ملف Excel"""
    if request.method == 'POST':
        form = InjuredImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['excel_file']
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['name', 'national_id', 'guardian_national_id', 
                                  'injury_date', 'injury_type']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'injured/injured_import.html', {'form': form})
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        # البحث عن ولي الأمر
                        guardian = None
                        if pd.notna(row['guardian_national_id']):
                            try:
                                guardian = Guardian.objects.get(national_id=str(row['guardian_national_id']).zfill(9))
                            except Guardian.DoesNotExist:
                                pass
                        
                        injured_data = {
                            'name': row['name'],
                            'national_id': str(row['national_id']).zfill(9),
                            'guardian_national_id': str(row['guardian_national_id']).zfill(9),
                            'injury_date': pd.to_datetime(row['injury_date']).date(),
                            'injury_type': row['injury_type'],
                            'guardian': guardian,
                            'guardian_name': guardian.name if guardian else '',
                            'phone_number': row.get('phone_number', guardian.phone_number if guardian else ''),
                            'injury_description': row.get('injury_description', ''),
                            'district': guardian.district if guardian else None,
                            'notes': row.get('notes', '')
                        }
                        
                        # التحقق من عدم وجود رقم الهوية مسبقاً
                        if not Injured.objects.filter(national_id=injured_data['national_id']).exists():
                            Injured.objects.create(**injured_data)
                            success_count += 1
                        else:
                            errors.append(f'الصف {index + 2}: رقم الهوية {injured_data["national_id"]} موجود مسبقاً')
                            error_count += 1
                            
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                        error_count += 1
                
                if success_count > 0:
                    messages.success(request, f'تم استيراد {success_count} جريح بنجاح')
                
                if error_count > 0:
                    error_msg = f'فشل في استيراد {error_count} سجل. الأخطاء:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_msg)
                
                return redirect('injured:injured_list')
                
            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء قراءة الملف: {str(e)}')
    else:
        form = InjuredImportForm()
    
    return render(request, 'injured/injured_import.html', {'form': form})

def injured_ajax_search(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) >= 2:
        injured = search_injured(query)[:10]
        results = [
            {
                'id': person.id,
                'name': person.name,
                'national_id': person.national_id,
                'district': person.district.name if person.district else ''
            }
            for person in injured
        ]
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})

def guardian_lookup_ajax(request):
    """البحث عن ولي الأمر عبر AJAX"""
    national_id = request.GET.get('national_id', '')
    if len(national_id) == 9:
        try:
            guardian = Guardian.objects.get(national_id=national_id)
            return JsonResponse({
                'found': True,
                'name': guardian.name,
                'phone': guardian.phone_number or '',
                'district_id': guardian.district.id if guardian.district else None,
                'district_name': guardian.district.name if guardian.district else ''
            })
        except Guardian.DoesNotExist:
            return JsonResponse({'found': False})
    return JsonResponse({'found': False})

def export_injured_excel(request):
    """تصدير بيانات المصابين إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    form = InjuredSearchForm(request.GET)
    injured = Injured.objects.all().select_related('district', 'guardian')
    
    if form.is_valid():
        search_query = form.cleaned_data.get('search')
        if search_query:
            injured = search_injured(search_query)
        
        district = form.cleaned_data.get('district')
        if district:
            injured = injured.filter(district=district)
            
        injury_type = form.cleaned_data.get('injury_type')
        if injury_type:
            injured = injured.filter(injury_type=injury_type)
            
        year = form.cleaned_data.get('year')
        if year:
            injured = injured.filter(injury_date__year=year)
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات المصابين"
    
    # العناوين
    headers = [
        'اسم المصاب', 'رقم الهوية', 'ولي الأمر', 'رقم هاتف ولي الأمر',
        'تاريخ الإصابة', 'نوع الإصابة', 'وصف الإصابة', 'رقم الهاتف',
        'الحي', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, injured_item in enumerate(injured, 2):
        ws.cell(row=row, column=1, value=injured_item.name)
        ws.cell(row=row, column=2, value=injured_item.national_id)
        ws.cell(row=row, column=3, value=injured_item.guardian.name if injured_item.guardian else injured_item.guardian_name)
        ws.cell(row=row, column=4, value=injured_item.guardian.phone_number if injured_item.guardian else "")
        ws.cell(row=row, column=5, value=injured_item.injury_date.strftime('%Y-%m-%d') if injured_item.injury_date else "")
        ws.cell(row=row, column=6, value=injured_item.get_injury_type_display())
        ws.cell(row=row, column=7, value=injured_item.injury_description or "")
        ws.cell(row=row, column=8, value=injured_item.phone_number or "")
        ws.cell(row=row, column=9, value=injured_item.district.name if injured_item.district else "")
        ws.cell(row=row, column=10, value=injured_item.notes or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = f"المصابين_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_injured_excel(request):
    """استيراد بيانات المصابين من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # تحديد الأعمدة التي يجب قراءتها كنص لتجنب مشكلة الأرقام العشرية
            dtype_dict = {
                'رقم الهوية': str,
                'رقم هوية ولي الأمر': str,
                'رقم الهاتف': str
            }
            
            df = pd.read_excel(excel_file, dtype=dtype_dict)
            
            def clean_national_id(value):
                """تنظيف رقم الهوية وإضافة الأصفار إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                if cleaned.isdigit() and len(cleaned) < 9:
                    cleaned = cleaned.zfill(9)
                
                return cleaned
            
            def clean_phone_number(value):
                """تنظيف رقم الجوال وإضافة الصفر إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                if cleaned.isdigit() and len(cleaned) == 9 and cleaned.startswith('5'):
                    cleaned = '0' + cleaned
                
                return cleaned
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # تنظيف البيانات
                    national_id = clean_national_id(row.get('رقم الهوية', ''))
                    phone_number = clean_phone_number(row.get('رقم الهاتف', ''))
                    guardian_national_id = clean_national_id(row.get('رقم هوية ولي الأمر', ''))
                    
                    # التحقق من البيانات المطلوبة
                    if pd.isna(row.get('اسم المصاب')) or not national_id or len(national_id) != 9:
                        errors.append(f"الصف {index + 2}: اسم المصاب ورقم الهوية مطلوبان وصحيحان")
                        error_count += 1
                        continue
                    
                    # البحث عن ولي الأمر إذا كان موجود
                    guardian = None
                    if guardian_national_id and len(guardian_national_id) == 9:
                        try:
                            guardian = Guardian.objects.get(national_id=guardian_national_id)
                        except Guardian.DoesNotExist:
                            pass
                    
                    # البحث عن الحي
                    district = None
                    if not pd.isna(row.get('الحي')) or not pd.isna(row.get('المنطقة')):
                        district_name = str(row.get('الحي', row.get('المنطقة', ''))).strip()
                        if district_name:
                            try:
                                district = District.objects.get(name=district_name)
                            except District.DoesNotExist:
                                pass
                    
                    # إنشاء المصاب
                    injured = Injured.objects.create(
                        name=str(row['اسم المصاب']).strip(),
                        national_id=national_id,
                        guardian_national_id=guardian_national_id,
                        guardian=guardian,
                        guardian_name=str(row.get('ولي الأمر', '')).strip() if not pd.isna(row.get('ولي الأمر')) else '',
                        injury_date=pd.to_datetime(row['تاريخ الإصابة']).date() if not pd.isna(row.get('تاريخ الإصابة')) else None,
                        injury_type=str(row.get('نوع الإصابة', 'gunshot')).strip().lower() if not pd.isna(row.get('نوع الإصابة')) else 'gunshot',
                        injury_description=str(row.get('وصف الإصابة', '')).strip() if not pd.isna(row.get('وصف الإصابة')) else '',
                        phone_number=phone_number,
                        district=district,
                        notes=str(row.get('ملاحظات', '')).strip() if not pd.isna(row.get('ملاحظات')) else '',
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} مصاب بنجاح")
            
            if error_count > 0:
                messages.warning(request, f"فشل في استيراد {error_count} صف. الأخطاء: {'; '.join(errors[:5])}")
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('injured:injured_list')

def download_injured_template(request):
    """تحميل نموذج Excel فارغ للمصابين"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج المصابين"
    
    # العناوين
    headers = [
        'اسم المصاب', 'رقم الهوية', 'رقم هوية ولي الأمر', 'ولي الأمر',
        'تاريخ الإصابة', 'نوع الإصابة', 'وصف الإصابة', 'رقم الهاتف',
        'الحي', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'محمد أحمد', '123456789', '987654321', 'أحمد محمد',
        '2024-01-15', 'gunshot', 'إصابة في الرأس', '0591234567',
        'الرمال', 'يحتاج متابعة'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # إضافة ملاحظات
    ws.cell(row=4, column=1, value="ملاحظات:")
    ws.cell(row=5, column=1, value="• أنواع الإصابة: gunshot, shrapnel, explosion, burn, fracture, amputation, other")
    ws.cell(row=6, column=1, value="• التاريخ: استخدم تنسيق YYYY-MM-DD")
    ws.cell(row=7, column=1, value="• رقم الهوية: 9 أرقام بالضبط")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = "نموذج_المصابين.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
