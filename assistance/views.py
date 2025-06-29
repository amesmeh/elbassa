from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from datetime import datetime, date
import json
import pandas as pd
import io
import csv
from basic_data.models import District, Guardian, Representative
from basic_data.permissions import get_user_district
from .forms import AssistanceForm, AssistanceSearchForm, AssistanceImportForm
from .utils import search_assistance
import tempfile
import os

from .models import Assistance

@login_required
def assistance_list(request):
    """عرض قائمة المساعدات مع البحث والتصفية والتأكد من تطابق الأسماء"""
    
    # 1. جلب كل المساعدات وتصفيتها
    assistances_qs = Assistance.objects.all().order_by('-assistance_date', '-created_at')
    
    # تصفية المساعدات حسب الحي للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_user_district(request.user)
        if district:
            assistances_qs = assistances_qs.filter(
                Q(district_name=district.name) |
                Q(district_name='') |
                Q(district_name__isnull=True)
            )
    
    search_form = AssistanceSearchForm(request.GET)
    
    if search_form.is_valid():
        query = search_form.cleaned_data.get('search_query')
        assistance_type = search_form.cleaned_data.get('assistance_type')
        district = search_form.cleaned_data.get('district')
        date_from = search_form.cleaned_data.get('date_from')
        date_to = search_form.cleaned_data.get('date_to')

        if query:
            query = query.strip()
            search_terms = query.split()
            name_q = Q()
            for term in search_terms:
                name_q &= Q(beneficiary_name__icontains=term)
            
            assistances_qs = assistances_qs.filter(
                Q(national_id__icontains=query) |
                name_q |
                Q(district_name__icontains=query) |
                Q(assistance_type__icontains=query) |
                Q(notes__icontains=query)
            )
        
        if assistance_type:
            assistances_qs = assistances_qs.filter(assistance_type__icontains=assistance_type)
        if district:
            assistances_qs = assistances_qs.filter(district_name__icontains=district)
        if date_from:
            assistances_qs = assistances_qs.filter(assistance_date__gte=date_from)
        if date_to:
            assistances_qs = assistances_qs.filter(assistance_date__lte=date_to)

    # 2. تحويل QuerySet إلى قائمة لمعالجتها
    assistances_list = list(assistances_qs)
    
    # 3. جلب أرقام الهويات وتحديث الأسماء
    national_ids = [a.national_id for a in assistances_list if a.national_id]
    if national_ids:
        guardians = Guardian.objects.filter(national_id__in=national_ids).values('national_id', 'name', 'residence_status')
        guardian_info_map = {
            g['national_id']: {
                'name': g['name'],
                'residence_status': g['residence_status']
            } for g in guardians
        }
        
        for assistance in assistances_list:
            if assistance.national_id in guardian_info_map:
                info = guardian_info_map[assistance.national_id]
                assistance.beneficiary_name = info['name']
                assistance.residence_status = info['residence_status']
            else:
                assistance.residence_status = "غير محدد"

    # 4. حساب الإحصائيات بعد التصفية
    total_records = len(assistances_list)
    total_records_in_db = Assistance.objects.count()
    
    # حساب الإحصائيات الفريدة من QuerySet الأصلي قبل تحويله إلى قائمة
    distribution_dates = assistances_qs.values('assistance_date').distinct().count()
    unique_beneficiaries = assistances_qs.values('national_id').distinct().count()
    assistance_types_count = assistances_qs.values('assistance_type').distinct().count()
    
    districts = Assistance.objects.exclude(district_name='').values_list('district_name', flat=True).distinct().order_by('district_name')
    assistance_types = Assistance.objects.exclude(assistance_type='').values_list('assistance_type', flat=True).distinct().order_by('assistance_type')
    
    # 5. التصفح باستخدام القائمة المحدثة
    paginator = Paginator(assistances_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_records': total_records,
        'total_records_in_db': total_records_in_db,
        'distribution_dates': distribution_dates,
        'unique_beneficiaries': unique_beneficiaries,
        'assistance_types_count': assistance_types_count,
        'districts': districts,
        'assistance_types': assistance_types,
        'title': 'المساعدات',
        'can_edit': request.user.is_superuser,
        'can_delete': request.user.is_superuser,
    }
    return render(request, 'assistance/assistance_list.html', context)

@login_required
def assistance_detail(request, pk):
    """عرض تفاصيل مساعدة"""
    assistance = get_object_or_404(Assistance, pk=pk)

    # -- START: Logic to sync name --
    if assistance.national_id:
        try:
            guardian = Guardian.objects.get(national_id=assistance.national_id)
            assistance.beneficiary_name = guardian.name
        except Guardian.DoesNotExist:
            pass  # Keep the original name if no guardian is found
    # -- END: Logic to sync name --
    
    # التحقق من صلاحية الوصول للمندوبين
    if not request.user.is_superuser and hasattr(request.user, 'representative'):
        district = get_user_district(request.user)
        if district and assistance.district_name and assistance.district_name != district.name:
            messages.error(request, 'غير مصرح لك بالوصول إلى هذه المساعدة')
            return redirect('assistance:assistance_list')
    
    # الحصول على تاريخ المستفيد
    beneficiary_history = Assistance.get_beneficiary_history(assistance.national_id)
    
    context = {
        'assistance': assistance,
        'beneficiary_history': beneficiary_history,
        'title': f'تفاصيل المساعدة - {assistance.beneficiary_name}',
        'can_edit': request.user.is_superuser,
        'can_delete': request.user.is_superuser,
    }
    return render(request, 'assistance/assistance_detail.html', context)

@login_required
@permission_required('assistance.add_assistance', raise_exception=True)
def assistance_create(request):
    """إضافة مساعدة جديدة - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بإضافة مساعدات جديدة')
        return redirect('assistance:assistance_list')
    
    # التحقق من وجود معامل copy لتكرار السجل
    copy_id = request.GET.get('copy')
    initial_data = {}
    
    if copy_id:
        try:
            source_assistance = Assistance.objects.get(pk=copy_id)
            initial_data = {
                'national_id': source_assistance.national_id,
                'beneficiary_name': source_assistance.beneficiary_name,
                'assistance_type': source_assistance.assistance_type,
                'quantity': source_assistance.quantity,
                'assistance_date': date.today(),  # تاريخ اليوم
                'notes': source_assistance.notes,
            }
        except Assistance.DoesNotExist:
            messages.warning(request, 'المساعدة المحددة للتكرار غير موجودة.')
    
    if request.method == 'POST':
        form = AssistanceForm(request.POST)
        if form.is_valid():
            assistance = form.save(commit=False)
            
            # البحث عن معلومات إضافية للمستفيد
            national_id = assistance.national_id
            
            # البحث في أولياء الأمور أولاً
            try:
                guardian = Guardian.objects.get(national_id=national_id)
                assistance.district_name = guardian.district.name if guardian.district else ''
                assistance.family_members_count = guardian.family_members_count if hasattr(guardian, 'family_members_count') else 1
            except:
                # البحث في السجل المدني
                try:
                    civil_record = CivilRegistry.objects.get(national_id=national_id)
                    assistance.district_name = civil_record.city
                    assistance.family_members_count = 1
                except:
                    pass
            
            assistance.save()
            
            if copy_id:
                messages.success(request, f'تم تكرار المساعدة وإضافتها لـ {assistance.beneficiary_name} بنجاح.')
                # العودة إلى قائمة المساعدات بعد التكرار
                return redirect('assistance:assistance_list')
            else:
                messages.success(request, f'تم إضافة مساعدة لـ {assistance.beneficiary_name} بنجاح.')
                # الذهاب إلى تفاصيل المساعدة للإضافة العادية
                return redirect('assistance:assistance_detail', pk=assistance.pk)
    else:
        form = AssistanceForm(initial=initial_data)
    
    # تحديد العنوان حسب نوع العملية
    if copy_id:
        title = 'تكرار مساعدة'
    else:
        title = 'إضافة مساعدة جديدة'
    
    # الحصول على قائمة أنواع المساعدات المتاحة للاقتراحات
    assistance_types = Assistance.objects.exclude(assistance_type='').values_list('assistance_type', flat=True).distinct().order_by('assistance_type')
    
    context = {
        'form': form,
        'title': title,
        'is_copy': bool(copy_id),
        'assistance_types': assistance_types
    }
    return render(request, 'assistance/assistance_form.html', context)

@login_required
@permission_required('assistance.change_assistance', raise_exception=True)
def assistance_edit(request, pk):
    """تعديل مساعدة - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بتعديل المساعدات')
        return redirect('assistance:assistance_list')
    
    assistance = get_object_or_404(Assistance, pk=pk)

    # -- START: Logic to sync name on GET request --
    if request.method == 'GET' and assistance.national_id:
        try:
            guardian = Guardian.objects.get(national_id=assistance.national_id)
            assistance.beneficiary_name = guardian.name
        except Guardian.DoesNotExist:
            pass  # Keep the original name if no guardian is found
    # -- END: Logic to sync name on GET request --
    
    if request.method == 'POST':
        form = AssistanceForm(request.POST, instance=assistance)
        if form.is_valid():
            assistance = form.save()
            messages.success(request, f'تم تحديث بيانات المساعدة لـ {assistance.beneficiary_name} بنجاح.')
            return redirect('assistance:assistance_detail', pk=assistance.pk)
    else:
        form = AssistanceForm(instance=assistance)
    
    # الحصول على قائمة أنواع المساعدات المتاحة للاقتراحات
    assistance_types = Assistance.objects.exclude(assistance_type='').values_list('assistance_type', flat=True).distinct().order_by('assistance_type')
    
    context = {
        'form': form,
        'assistance': assistance,
        'title': f'تعديل المساعدة - {assistance.beneficiary_name}',
        'assistance_types': assistance_types
    }
    return render(request, 'assistance/assistance_form.html', context)

@login_required
@permission_required('assistance.delete_assistance', raise_exception=True)
def assistance_delete(request, pk):
    """حذف مساعدة - للمشرفين فقط"""
    if not request.user.is_superuser:
        messages.error(request, 'غير مصرح لك بحذف المساعدات')
        return redirect('assistance:assistance_list')
    
    assistance = get_object_or_404(Assistance, pk=pk)
    beneficiary_name = assistance.beneficiary_name
    assistance.delete()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'تم حذف المساعدة لـ {beneficiary_name} بنجاح.'})
    
    messages.success(request, f'تم حذف المساعدة لـ {beneficiary_name} بنجاح.')
    return redirect('assistance:assistance_list')

def beneficiary_search_ajax(request):
    """البحث السريع عن المستفيدين عبر AJAX"""
    national_id = request.GET.get('national_id', '')
    
    if len(national_id) < 3:
        return JsonResponse({'found': False, 'message': 'أدخل 3 أرقام على الأقل'})
    
    # البحث في قاعدة بيانات أولياء الأمور أولاً
    try:
        guardian = Guardian.objects.get(national_id=national_id)
        return JsonResponse({
            'found': True,
            'name': guardian.name,
            'district': guardian.district.name if guardian.district else '',
            'family_members': guardian.family_members_count if hasattr(guardian, 'family_members_count') else 1,
            'source': 'guardian'
        })
    except:
        pass
    
    # البحث في سجلات المساعدات السابقة
    last_assistance = Assistance.objects.filter(national_id=national_id).first()
    if last_assistance:
        return JsonResponse({
            'found': True,
            'name': last_assistance.beneficiary_name,
            'district': last_assistance.district_name or '',
            'family_members': last_assistance.family_members_count if hasattr(last_assistance, 'family_members_count') else 1,
            'source': 'assistance_history'
        })
    
    return JsonResponse({'found': False, 'message': 'لم يتم العثور على المستفيد'})

@require_http_methods(["POST"])
def assistance_quick_edit(request, pk):
    """التعديل السريع للمساعدة عبر AJAX"""
    assistance = get_object_or_404(Assistance, pk=pk)
    
    try:
        data = json.loads(request.body)
        field = list(data.keys())[0]
        value = data[field]
        
        # التحقق من صحة البيانات
        if field == 'beneficiary_name':
            if not value or len(value) < 2:
                return JsonResponse({'success': False, 'message': 'اسم المستفيد يجب أن يكون أكثر من حرفين'})
            assistance.beneficiary_name = value
            
        elif field == 'national_id':
            # التحقق من صحة رقم الهوية
            if not value or len(value) != 9 or not value.isdigit():
                return JsonResponse({'success': False, 'message': 'رقم الهوية يجب أن يكون 9 أرقام'})
            
            # البحث عن ولي الأمر برقم الهوية الجديد
            from basic_data.models import Guardian
            try:
                guardian = Guardian.objects.get(national_id=value)
                # تحديث الاسم تلقائياً من قاعدة بيانات أولياء الأمور
                assistance.beneficiary_name = guardian.name
                assistance.district_name = guardian.district.name if guardian.district else ''
                assistance.family_members_count = guardian.family_members_count if hasattr(guardian, 'family_members_count') else 1
                assistance.guardian = guardian
            except Guardian.DoesNotExist:
                # إذا لم يوجد ولي الأمر، نحتفظ بالاسم الحالي
                pass
            
            assistance.national_id = value
            
        elif field == 'assistance_type':
            if not value or len(value.strip()) < 2:
                return JsonResponse({'success': False, 'message': 'نوع المساعدة يجب أن يكون أكثر من حرفين'})
            assistance.assistance_type = value.strip()
            
        elif field == 'quantity':
            try:
                quantity = float(value)
                if quantity <= 0:
                    return JsonResponse({'success': False, 'message': 'الكمية يجب أن تكون أكبر من صفر'})
                assistance.quantity = quantity
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'message': 'الكمية يجب أن تكون رقماً صحيحاً'})
                
        elif field == 'assistance_date':
            try:
                assistance.assistance_date = datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'success': False, 'message': 'تنسيق التاريخ غير صحيح'})
                
        elif field == 'district_name':
            assistance.district_name = value
            
        else:
            return JsonResponse({'success': False, 'message': 'حقل غير مسموح بتعديله'})
        
        assistance.save()
        
        # إرسال القيمة المعروضة
        display_value = value
        if field == 'assistance_type':
            display_value = assistance.assistance_type
        elif field == 'quantity':
            display_value = str(int(float(value)))
        
        # إضافة معلومات إضافية إذا تم تعديل رقم الهوية
        response_data = {
            'success': True,
            'display_value': display_value,
            'message': 'تم التحديث بنجاح'
        }
        
        if field == 'national_id':
            response_data['beneficiary_name'] = assistance.beneficiary_name
            response_data['family_members_count'] = assistance.family_members_count if hasattr(assistance, 'family_members_count') and assistance.family_members_count else 1
            response_data['district_name'] = assistance.district_name or ''
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'حدث خطأ: {str(e)}'})

def assistance_get_field_value(request, pk):
    """الحصول على قيمة حقل معين من المساعدة"""
    assistance = get_object_or_404(Assistance, pk=pk)
    field = request.GET.get('field')
    
    if field == 'beneficiary_name':
        value = assistance.beneficiary_name
        display_value = assistance.beneficiary_name
    elif field == 'national_id':
        value = assistance.national_id
        display_value = assistance.national_id
    elif field == 'assistance_type':
        value = assistance.assistance_type
        display_value = assistance.assistance_type
    elif field == 'quantity':
        value = str(assistance.quantity)
        display_value = str(int(assistance.quantity))
    elif field == 'assistance_date':
        value = assistance.assistance_date.strftime('%Y-%m-%d')
        display_value = value
    elif field == 'district_name':
        value = assistance.district_name or ''
        display_value = assistance.district_name or ''
    else:
        value = ''
        display_value = ''
    
    return JsonResponse({
        'value': value,
        'display_value': display_value
    })

def export_assistance_excel(request):
    """تصدير المساعدات إلى Excel مع تطبيق الفلاتر"""
    search_query = request.GET.get('search', '')
    assistance_type_filter = request.GET.get('assistance_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    district_filter = request.GET.get('district', '')
    
    # تطبيق نفس منطق الفلترة المستخدم في assistance_list
    assistances = Assistance.objects.all().order_by('-assistance_date', '-created_at')
    
    if search_query:
        query = search_query.strip()
        search_terms = query.split()
        name_q = Q()
        for term in search_terms:
            name_q &= Q(beneficiary_name__icontains=term)
        
        assistances = assistances.filter(
            Q(national_id__icontains=query) |
            name_q |
            Q(district_name__icontains=query) |
            Q(assistance_type__icontains=query) |
            Q(notes__icontains=query)
        )
    
    if assistance_type_filter:
        assistances = assistances.filter(assistance_type__icontains=assistance_type_filter)
    
    if district_filter:
        assistances = assistances.filter(district_name__icontains=district_filter)
    
    if date_from:
        assistances = assistances.filter(assistance_date__gte=date_from)
    
    if date_to:
        assistances = assistances.filter(assistance_date__lte=date_to)
    
    try:
        # محاولة استخدام pandas و openpyxl
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # إنشاء ملف Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="assistance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # إنشاء DataFrame
        data = []
        for assistance in assistances:
            # البحث عن معلومات الزوجات من خلال رقم الهوية
            wives_info = {}
            
            # تهيئة جميع مفاتيح الزوجات أولاً
            for i in range(1, 5):
                wives_info[f'wife_{i}_name'] = ''
                wives_info[f'wife_{i}_id'] = ''
            
            try:
                guardian = Guardian.objects.filter(national_id=assistance.national_id).first()
                if guardian:
                    wives = guardian.wives.all()[:4]  # حتى 4 زوجات
                    for i, wife in enumerate(wives, 1):
                        wives_info[f'wife_{i}_name'] = wife.name
                        wives_info[f'wife_{i}_id'] = wife.national_id or ''
            except Exception as e:
                # في حالة الخطأ، المفاتيح موجودة بالفعل بقيم فارغة
                pass
            
            row_data = {
                'رقم الهوية': assistance.national_id,
                'اسم المستفيد': assistance.beneficiary_name,
                'نوع المساعدة': assistance.assistance_type,
                'الكمية': assistance.quantity,
                'تاريخ المساعدة': assistance.assistance_date.strftime('%Y-%m-%d') if assistance.assistance_date else '',
                'اسم الحي': assistance.district_name or '',
                'عدد أفراد العائلة': assistance.family_members_count if hasattr(assistance, 'family_members_count') and assistance.family_members_count else 1,
                'اسم الزوجة 1': wives_info['wife_1_name'],
                'رقم هوية الزوجة 1': wives_info['wife_1_id'],
                'اسم الزوجة 2': wives_info['wife_2_name'],
                'رقم هوية الزوجة 2': wives_info['wife_2_id'],
                'اسم الزوجة 3': wives_info['wife_3_name'],
                'رقم هوية الزوجة 3': wives_info['wife_3_id'],
                'اسم الزوجة 4': wives_info['wife_4_name'],
                'رقم هوية الزوجة 4': wives_info['wife_4_id'],
                'الملاحظات': assistance.notes or '',
                'تاريخ الإضافة': assistance.created_at.strftime('%Y-%m-%d %H:%M:%S') if assistance.created_at else '',
            }
            data.append(row_data)
        
        df = pd.DataFrame(data)
        
        # إنشاء workbook وورقة عمل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "المساعدات"
        
        # إضافة البيانات
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # تنسيق الخط العربي Simplified Arabic
        arabic_font = Font(name='Simplified Arabic', size=12)
        
        # تنسيق العناوين
        header_font = Font(name='Simplified Arabic', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, readingOrder=2)
        
        # تطبيق التنسيق على العناوين
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # إضافة حدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, readingOrder=2)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = arabic_font
                cell.alignment = data_alignment
                # إضافة حدود
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # تعيين اتجاه الورقة من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True
        
        # تعديل عرض الأعمدة
        column_widths = {
            'A': 12,  # رقم الهوية
            'B': 20,  # اسم المستفيد
            'C': 15,  # نوع المساعدة
            'D': 10,  # الكمية
            'E': 15,  # تاريخ المساعدة
            'F': 15,  # اسم الحي
            'G': 12,  # عدد أفراد العائلة
            'H': 18,  # اسم الزوجة 1
            'I': 15,  # رقم هوية الزوجة 1
            'J': 18,  # اسم الزوجة 2
            'K': 15,  # رقم هوية الزوجة 2
            'L': 18,  # اسم الزوجة 3
            'M': 15,  # رقم هوية الزوجة 3
            'N': 18,  # اسم الزوجة 4
            'O': 15,  # رقم هوية الزوجة 4
            'P': 25,  # الملاحظات
            'Q': 18,  # تاريخ الإضافة
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # تجميد الصف الأول
        ws.freeze_panes = ws['A2']
        
        # حفظ إلى buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except ImportError:
        # نظام بديل باستخدام CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="assistance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'رقم الهوية', 'اسم المستفيد', 'نوع المساعدة', 'الكمية', 
            'تاريخ المساعدة', 'اسم الحي', 'عدد أفراد العائلة',
            'اسم الزوجة 1', 'رقم هوية الزوجة 1',
            'اسم الزوجة 2', 'رقم هوية الزوجة 2',
            'اسم الزوجة 3', 'رقم هوية الزوجة 3',
            'اسم الزوجة 4', 'رقم هوية الزوجة 4',
            'الملاحظات', 'تاريخ الإضافة'
        ])
        
        for assistance in assistances:
            # البحث عن معلومات الزوجات
            wives_info = ['', '', '', '', '', '', '', '']  # 4 زوجات × 2 (اسم + رقم هوية)
            try:
                guardian = Guardian.objects.filter(national_id=assistance.national_id).first()
                if guardian:
                    wives = guardian.wives.all()[:4]
                    for i, wife in enumerate(wives):
                        wives_info[i*2] = wife.name
                        wives_info[i*2+1] = wife.national_id or ''
            except:
                pass
            
            writer.writerow([
                assistance.national_id,
                assistance.beneficiary_name,
                assistance.get_assistance_type_display(),
                assistance.quantity,
                assistance.assistance_date.strftime('%Y-%m-%d') if assistance.assistance_date else '',
                assistance.district_name or '',
                assistance.family_members_count if hasattr(assistance, 'family_members_count') and assistance.family_members_count else 1,
                wives_info[0], wives_info[1],  # الزوجة 1
                wives_info[2], wives_info[3],  # الزوجة 2
                wives_info[4], wives_info[5],  # الزوجة 3
                wives_info[6], wives_info[7],  # الزوجة 4
                assistance.notes or '',
                assistance.created_at.strftime('%Y-%m-%d %H:%M:%S') if assistance.created_at else '',
            ])
        
        return response

def create_assistance_template(request):
    """تحميل نموذج استيراد المساعدات"""
    try:
        # محاولة استخدام pandas و openpyxl
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="assistance_import_template.xlsx"'
        
        # إنشاء نموذج بيانات
        template_data = [
            {
                'رقم الهوية': '123456789',
                'اسم المستفيد': 'أحمد محمد علي',
                'نوع المساعدة': 'مواد غذائية',
                'الكمية': 50,
                'تاريخ المساعدة': '2024-01-15',
            },
            {
                'رقم الهوية': '987654321',
                'اسم المستفيد': 'محمد خالد حسن',
                'نوع المساعدة': 'مساعدة نقدية',
                'الكمية': 500,
                'تاريخ المساعدة': '2024-01-20',
            },
            {
                'رقم الهوية': '456789123',
                'اسم المستفيد': 'خالد أحمد سالم',
                'نوع المساعدة': 'أدوية',
                'الكمية': 10,
                'تاريخ المساعدة': '2024-01-25',
            }
        ]
        
        df = pd.DataFrame(template_data)
        
        # إنشاء workbook وورقة عمل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نموذج المساعدات"
        
        # إضافة البيانات
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # تنسيق الخط العربي Simplified Arabic
        arabic_font = Font(name='Simplified Arabic', size=12)
        
        # تنسيق العناوين
        header_font = Font(name='Simplified Arabic', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, readingOrder=2)
        
        # تطبيق التنسيق على العناوين
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # إضافة حدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, readingOrder=2)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = arabic_font
                cell.alignment = data_alignment
                # إضافة حدود
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # تعيين اتجاه الورقة من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True
        
        # تعديل عرض الأعمدة
        column_widths = {
            'A': 15,  # رقم الهوية
            'B': 25,  # اسم المستفيد
            'C': 20,  # نوع المساعدة
            'D': 12,  # الكمية
            'E': 18,  # تاريخ المساعدة
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # تجميد الصف الأول
        ws.freeze_panes = ws['A2']
        
        # حفظ إلى buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except ImportError:
        # نظام بديل باستخدام CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="assistance_import_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'رقم الهوية', 'اسم المستفيد', 'نوع المساعدة', 'الكمية', 'تاريخ المساعدة'
        ])
        
        # إضافة بيانات نموذجية
        writer.writerow([
            '123456789', 'أحمد محمد علي', 'مواد غذائية', '50', '2024-01-15'
        ])
        writer.writerow([
            '987654321', 'محمد خالد حسن', 'مساعدة نقدية', '500', '2024-01-20'
        ])
        writer.writerow([
            '456789123', 'خالد أحمد سالم', 'أدوية', '10', '2024-01-25'
        ])
        
        return response

def import_assistance_excel(request):
    """استيراد المساعدات من ملف Excel"""
    if request.method == 'POST':
        form = AssistanceImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                
                # محاولة استخدام pandas أولاً
                try:
                    import pandas as pd
                    
                    # قراءة ملف Excel
                    df = pd.read_excel(excel_file)
                    
                    # التحقق من وجود الأعمدة المطلوبة
                    required_columns = ['رقم الهوية', 'اسم المستفيد', 'نوع المساعدة', 'الكمية', 'تاريخ المساعدة']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                        return render(request, 'assistance/assistance_import.html', {'form': form})
                    
                    # استيراد البيانات
                    imported_count = 0
                    errors = []
                    
                    for index, row in df.iterrows():
                        try:
                            # التحقق من البيانات المطلوبة
                            if pd.isna(row['رقم الهوية']) or pd.isna(row['اسم المستفيد']):
                                errors.append(f'الصف {index + 2}: رقم الهوية واسم المستفيد مطلوبان')
                                continue
                            
                            national_id = str(int(row['رقم الهوية'])) if not pd.isna(row['رقم الهوية']) else ''
                            if len(national_id) != 9:
                                errors.append(f'الصف {index + 2}: رقم الهوية يجب أن يكون 9 أرقام')
                                continue
                            
                            # البحث في قاعدة البيانات أولاً، وإذا لم يوجد نستخدم الاسم من الملف
                            beneficiary_name = ''
                            try:
                                guardian = Guardian.objects.filter(national_id=national_id).first()
                                if guardian:
                                    # إذا وجد في قاعدة البيانات، استخدم اسمه من القاعدة
                                    beneficiary_name = guardian.name
                                else:
                                    # إذا لم يوجد في القاعدة، استخدم الاسم من الملف
                                    beneficiary_name = str(row['اسم المستفيد']).strip() if not pd.isna(row['اسم المستفيد']) else f'مستفيد - {national_id}'
                            except Exception:
                                # في حالة الخطأ، استخدم الاسم من الملف
                                beneficiary_name = str(row['اسم المستفيد']).strip() if not pd.isna(row['اسم المستفيد']) else f'مستفيد - {national_id}'
                            
                            # نوع المساعدة كما هو (حقل نصي حر)
                            assistance_type = str(row['نوع المساعدة']).strip() if not pd.isna(row['نوع المساعدة']) else 'غير محدد'
                            
                            # تحويل التاريخ (مع تحديد أن اليوم يأتي أولاً)
                            assistance_date = pd.to_datetime(row['تاريخ المساعدة'], dayfirst=True).date()
                            
                            # إنشاء المساعدة
                            assistance = Assistance(
                                national_id=national_id,
                                beneficiary_name=beneficiary_name,
                                assistance_type=assistance_type,
                                quantity=float(row['الكمية']) if not pd.isna(row['الكمية']) else 1,
                                assistance_date=assistance_date,
                                district_name='',  # سيتم إضافة هذا لاحقاً من بيانات الولي
                                family_members_count=1,  # سيتم حسابه تلقائياً من بيانات الولي
                                notes=''
                            )
                            
                            assistance.full_clean()
                            assistance.save()
                            
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append(f'الصف {index + 2}: {str(e)}')
                
                except ImportError:
                    messages.error(request, 'مكتبة pandas غير مثبتة. يرجى استخدام ملف CSV بدلاً من Excel أو تثبيت pandas.')
                    return render(request, 'assistance/assistance_import.html', {'form': form})
                
                # عرض النتائج
                if imported_count > 0:
                    messages.success(request, f'تم استيراد {imported_count} مساعدة بنجاح.')
                
                if errors:
                    error_message = f'حدثت أخطاء في {len(errors)} صف:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_message += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_message)
                
                if imported_count > 0:
                    return redirect('assistance:assistance_list')
                    
            except Exception as e:
                messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = AssistanceImportForm()
    
    context = {
        'form': form,
        'title': 'استيراد المساعدات'
    }
    return render(request, 'assistance/assistance_import.html', context)

@require_POST
def delete_selected_assistance(request):
    """حذف المساعدات المحددة"""
    try:
        import json
        data = json.loads(request.body)
        assistance_ids = data.get('assistance_ids', [])
        
        if not assistance_ids:
            return JsonResponse({'success': False, 'message': 'لم يتم تحديد أي مساعدات للحذف'})
        
        # حذف المساعدات المحددة
        deleted_count = Assistance.objects.filter(id__in=assistance_ids).delete()[0]
        
        return JsonResponse({
            'success': True, 
            'message': f'تم حذف {deleted_count} مساعدة بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@require_POST 
def delete_all_assistance(request):
    """حذف جميع المساعدات"""
    try:
        # حذف جميع المساعدات
        deleted_count = Assistance.objects.all().delete()[0]
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف جميع المساعدات ({deleted_count} سجل) بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@require_POST
def update_assistance_date(request, pk):
    """تحديث تاريخ المساعدة"""
    try:
        import json
        from datetime import datetime
        
        # الحصول على المساعدة
        assistance = get_object_or_404(Assistance, pk=pk)
        
        # قراءة البيانات من الطلب
        data = json.loads(request.body)
        new_date_str = data.get('new_date')
        reason = data.get('reason', '')
        
        if not new_date_str:
            return JsonResponse({'success': False, 'error': 'التاريخ الجديد مطلوب'})
        
        # تحويل التاريخ
        try:
            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'تنسيق التاريخ غير صحيح'})
        
        # حفظ التاريخ القديم لأغراض التسجيل
        old_date = assistance.assistance_date
        
        # التحقق من أن التاريخ مختلف
        if old_date == new_date:
            return JsonResponse({'success': False, 'error': 'التاريخ الجديد نفس التاريخ القديم'})
        
        # تحديث التاريخ
        assistance.assistance_date = new_date
        assistance.save()
        
        # يمكن إضافة سجل تدقيق هنا إذا كان مطلوباً
        # مثلاً: تسجيل عملية التغيير في جدول منفصل
        
        return JsonResponse({
            'success': True,
            'message': f'تم تحديث التاريخ من {old_date} إلى {new_date}',
            'new_date': new_date_str
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})

@require_POST
def bulk_update_date_preview(request):
    """معاينة التحديث الجماعي للتواريخ"""
    try:
        import json
        from datetime import datetime
        from django.db.models import Q
        from urllib.parse import parse_qs
        
        # قراءة البيانات من الطلب
        data = json.loads(request.body)
        old_date_str = data.get('old_date')
        new_date_str = data.get('new_date')
        include_filters = data.get('include_filters', False)
        search_params = data.get('search_params', '')
        
        if not old_date_str or not new_date_str:
            return JsonResponse({'success': False, 'error': 'التواريخ مطلوبة'})
        
        # تحويل التواريخ
        try:
            old_date = datetime.strptime(old_date_str, '%Y-%m-%d').date()
            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'تنسيق التاريخ غير صحيح'})
        
        # بناء الاستعلام
        assistances = Assistance.objects.filter(assistance_date=old_date)
        
        # تطبيق الفلاتر إذا كان مطلوباً
        if include_filters and search_params:
            from urllib.parse import parse_qs
            params = parse_qs(search_params)
            
            # تطبيق نفس منطق البحث من assistance_list
            if 'search_query' in params and params['search_query'][0]:
                query = params['search_query'][0].strip()
                search_terms = query.split()
                
                name_q = Q()
                for term in search_terms:
                    name_q &= Q(beneficiary_name__icontains=term)
                
                assistances = assistances.filter(
                    Q(national_id__icontains=query) |
                    name_q |
                    Q(district_name__icontains=query) |
                    Q(assistance_type__icontains=query) |
                    Q(notes__icontains=query)
                )
            
            if 'assistance_type' in params and params['assistance_type'][0]:
                assistances = assistances.filter(assistance_type__icontains=params['assistance_type'][0])
            
            if 'district' in params and params['district'][0]:
                assistances = assistances.filter(district_name__icontains=params['district'][0])
        
        count = assistances.count()
        
        # الحصول على عينة من الأسماء
        sample_names = ''
        if count > 0:
            sample_assistances = assistances[:3]
            names = [a.beneficiary_name for a in sample_assistances]
            sample_names = ', '.join(names)
            if count > 3:
                sample_names += f' و {count - 3} آخرين'
        
        return JsonResponse({
            'success': True,
            'count': count,
            'sample_names': sample_names
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})

@require_POST
def bulk_update_date(request):
    """التحديث الجماعي للتواريخ"""
    try:
        import json
        from datetime import datetime
        from django.db.models import Q
        
        # قراءة البيانات من الطلب
        data = json.loads(request.body)
        old_date_str = data.get('old_date')
        new_date_str = data.get('new_date')
        reason = data.get('reason', '')
        include_filters = data.get('include_filters', False)
        search_params = data.get('search_params', '')
        
        if not old_date_str or not new_date_str or not reason.strip():
            return JsonResponse({'success': False, 'error': 'جميع الحقول مطلوبة'})
        
        # تحويل التواريخ
        try:
            old_date = datetime.strptime(old_date_str, '%Y-%m-%d').date()
            new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'تنسيق التاريخ غير صحيح'})
        
        if old_date == new_date:
            return JsonResponse({'success': False, 'error': 'التاريخ الجديد نفس التاريخ القديم'})
        
        # بناء الاستعلام
        assistances = Assistance.objects.filter(assistance_date=old_date)
        
        # تطبيق الفلاتر إذا كان مطلوباً
        if include_filters and search_params:
            from urllib.parse import parse_qs
            params = parse_qs(search_params)
            
            # تطبيق نفس منطق البحث من assistance_list
            if 'search_query' in params and params['search_query'][0]:
                query = params['search_query'][0].strip()
                search_terms = query.split()
                
                name_q = Q()
                for term in search_terms:
                    name_q &= Q(beneficiary_name__icontains=term)
                
                assistances = assistances.filter(
                    Q(national_id__icontains=query) |
                    name_q |
                    Q(district_name__icontains=query) |
                    Q(assistance_type__icontains=query) |
                    Q(notes__icontains=query)
                )
            
            if 'assistance_type' in params and params['assistance_type'][0]:
                assistances = assistances.filter(assistance_type__icontains=params['assistance_type'][0])
            
            if 'district' in params and params['district'][0]:
                assistances = assistances.filter(district_name__icontains=params['district'][0])
        
        # تنفيذ التحديث
        updated_count = assistances.update(assistance_date=new_date)
        
        # يمكن إضافة سجل تدقيق هنا إذا كان مطلوباً
        # مثلاً: تسجيل عملية التغيير الجماعي في جدول منفصل
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'message': f'تم تحديث {updated_count} مساعدة من {old_date} إلى {new_date}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'حدث خطأ: {str(e)}'})
