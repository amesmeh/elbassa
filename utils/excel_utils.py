import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from basic_data.models import Guardian, District, Wife, Child
from datetime import datetime

def create_excel_response(filename):
    """إنشاء استجابة HTTP لملف Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def style_excel_header(worksheet, header_row=1):
    """تنسيق رأس الجدول في Excel"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[header_row]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

def auto_adjust_column_width(worksheet):
    """ضبط عرض الأعمدة تلقائياً مع مراعاة المحتوى العربي"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    # حساب طول النص مع مراعاة النصوص العربية
                    cell_length = len(str(cell.value))
                    # إضافة مساحة إضافية للنصوص العربية
                    if any('\u0600' <= char <= '\u06FF' for char in str(cell.value)):
                        cell_length = int(cell_length * 1.3)  # زيادة 30% للنصوص العربية
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # تحديد عرض مناسب (بحد أدنى 8 وحد أقصى 50)
        adjusted_width = max(min(max_length + 3, 50), 8)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_guardians_to_excel(guardians_queryset=None):
    """تصدير أولياء الأمور إلى Excel مع التنسيق العربي المحسن"""
    if guardians_queryset is None:
        guardians = Guardian.objects.select_related('district').prefetch_related('wives').all()
    else:
        guardians = guardians_queryset.select_related('district').prefetch_related('wives')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # إنشاء ملف Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="guardians_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # إعداد الترجمات العربية الشاملة
        gender_translation = {
            'male': 'ذكر',
            'female': 'أنثى',
            'M': 'ذكر',
            'F': 'أنثى',
            'ذكر': 'ذكر',
            'أنثى': 'أنثى'
        }
        
        marital_status_translation = {
            'single': 'أعزب',
            'married': 'متزوج',
            'divorced': 'مطلق',
            'widowed': 'أرمل',
            'أعزب': 'أعزب',
            'متزوج': 'متزوج',
            'مطلق': 'مطلق',
            'أرمل': 'أرمل'
        }
        
        residence_status_translation = {
            'resident': 'مقيم',
            'displaced': 'نازح',
            'refugee': 'لاجئ',
            'مقيم': 'مقيم',
            'نازح': 'نازح',
            'لاجئ': 'لاجئ'
        }
        
        # إنشاء DataFrame
        data = []
        for guardian in guardians:
            # البحث عن معلومات الزوجات
            wives_info = {}
            
            # تهيئة جميع مفاتيح الزوجات أولاً
            for i in range(1, 5):
                wives_info[f'wife_{i}_name'] = ''
                wives_info[f'wife_{i}_id'] = ''
            
            # ملء معلومات الزوجات
            wives = guardian.wives.all()[:4]  # حتى 4 زوجات
            for i, wife in enumerate(wives, 1):
                wives_info[f'wife_{i}_name'] = wife.name
                wives_info[f'wife_{i}_id'] = wife.national_id or ''
            
            # تحويل القيم للعربية
            gender_arabic = gender_translation.get(guardian.gender, guardian.gender)
            marital_status_arabic = marital_status_translation.get(guardian.marital_status, guardian.marital_status)
            residence_status_arabic = residence_status_translation.get(guardian.residence_status, guardian.residence_status)
            
            row_data = {
                'اسم ولي الأمر': guardian.name,
                'رقم الهوية': guardian.national_id,
                'الجنس': gender_arabic,
                'الوظيفة': guardian.current_job or '',
                'رقم الجوال': guardian.phone_number,
                'الحالة الاجتماعية': marital_status_arabic,
                'عدد الأبناء': guardian.children_count,
                'عدد الزوجات': guardian.wives_count,
                'اسم الزوجة 1': wives_info['wife_1_name'],
                'رقم هوية الزوجة 1': wives_info['wife_1_id'],
                'اسم الزوجة 2': wives_info['wife_2_name'],
                'رقم هوية الزوجة 2': wives_info['wife_2_id'],
                'اسم الزوجة 3': wives_info['wife_3_name'],
                'رقم هوية الزوجة 3': wives_info['wife_3_id'],
                'اسم الزوجة 4': wives_info['wife_4_name'],
                'رقم هوية الزوجة 4': wives_info['wife_4_id'],
                'عدد أفراد العائلة': guardian.family_members_count,
                'حالة الإقامة': residence_status_arabic,
                'المحافظة الأصلية': guardian.original_governorate or '',
                'المدينة الأصلية': guardian.original_city or '',
                'عنوان النزوح': guardian.displacement_address or '',
                'الحي الحالي': guardian.district.name if guardian.district else '',
                'تاريخ الإنشاء': guardian.created_at.strftime('%Y-%m-%d %H:%M:%S') if guardian.created_at else '',
            }
            data.append(row_data)
        
        df = pd.DataFrame(data)
        
        # إنشاء workbook وورقة عمل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أولياء الأمور"
        
        # إضافة البيانات
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # تنسيق الخط العربي Simplified Arabic
        arabic_font = Font(name='Simplified Arabic', size=12)
        
        # تنسيق العناوين
        header_font = Font(name='Simplified Arabic', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='17A2B8', end_color='17A2B8', fill_type='solid')  # لون أزرق فيروزي
        header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, readingOrder=2)
        
        # تطبيق التنسيق على العناوين
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # إضافة حدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, readingOrder=2)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = arabic_font
                cell.alignment = data_alignment
                # إضافة حدود
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # تعيين اتجاه الورقة من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True
        
        # ضبط عرض الأعمدة تلقائياً
        auto_adjust_column_width(ws)
        
        # تجميد الصف الأول
        ws.freeze_panes = ws['A2']
        
        # حفظ إلى buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except ImportError:
        # نظام بديل باستخدام pandas فقط
        data = []
        for guardian in guardians:
            # البحث عن معلومات الزوجات للنظام البديل
            wives_info = ['', '', '', '', '', '', '', '']  # 4 زوجات × 2 (اسم + رقم هوية)
            wives = guardian.wives.all()[:4]
            for i, wife in enumerate(wives):
                wives_info[i*2] = wife.name
                wives_info[i*2+1] = wife.national_id or ''
            
            # ترجمة الحالات للنظام البديل مع شمولية أكبر
            gender_arabic = gender_translation.get(guardian.gender, guardian.gender)
            if not gender_arabic or gender_arabic not in ['ذكر', 'أنثى']:
                gender_arabic = 'ذكر' if guardian.gender in ['male', 'M', 'ذكر'] else 'أنثى'
            marital_status_arabic = {
                'single': 'أعزب', 'married': 'متزوج', 'divorced': 'مطلق', 'widowed': 'أرمل'
            }.get(guardian.marital_status, guardian.marital_status)
            residence_status_arabic = {
                'resident': 'مقيم', 'displaced': 'نازح', 'refugee': 'لاجئ'
            }.get(guardian.residence_status, guardian.residence_status)
            
            data.append({
                'اسم ولي الأمر': guardian.name,
                'رقم الهوية': guardian.national_id,
                'الجنس': gender_arabic,
                'الوظيفة': guardian.current_job or '',
                'رقم الجوال': guardian.phone_number,
                'الحالة الاجتماعية': marital_status_arabic,
                'عدد الأبناء': guardian.children_count,
                'عدد الزوجات': guardian.wives_count,
                'اسم الزوجة 1': wives_info[0],
                'رقم هوية الزوجة 1': wives_info[1],
                'اسم الزوجة 2': wives_info[2],
                'رقم هوية الزوجة 2': wives_info[3],
                'اسم الزوجة 3': wives_info[4],
                'رقم هوية الزوجة 3': wives_info[5],
                'اسم الزوجة 4': wives_info[6],
                'رقم هوية الزوجة 4': wives_info[7],
                'عدد أفراد العائلة': guardian.family_members_count,
                'حالة الإقامة': residence_status_arabic,
                'المحافظة الأصلية': guardian.original_governorate or '',
                'المدينة الأصلية': guardian.original_city or '',
                'عنوان النزوح': guardian.displacement_address or '',
                'الحي الحالي': guardian.district.name if guardian.district else '',
                'تاريخ الإنشاء': guardian.created_at.strftime('%Y-%m-%d %H:%M:%S') if guardian.created_at else ''
            })
        
        df = pd.DataFrame(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='أولياء الأمور', index=False)
            
            # تنسيق الملف
            worksheet = writer.sheets['أولياء الأمور']
            style_excel_header(worksheet)
            auto_adjust_column_width(worksheet)
        
        output.seek(0)
        
        # إنشاء الاستجابة
        filename = f"guardians_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = create_excel_response(filename)
        response.write(output.getvalue())
        
        return response

def export_children_to_excel(children_queryset=None):
    """تصدير الأبناء إلى Excel"""
    if children_queryset is None:
        children = Child.objects.select_related('guardian', 'guardian__district').all()
    else:
        children = children_queryset.select_related('guardian', 'guardian__district')
    
    data = []
    for child in children:
        data.append({
            'اسم الطفل': child.name,
            'رقم الهوية': child.national_id or '',
            'تاريخ الميلاد': child.birth_date.strftime('%Y-%m-%d'),
            'العمر': child.age,
            'اسم ولي الأمر': child.guardian.name,
            'رقم هوية ولي الأمر': child.guardian.national_id,
            'الحي': child.guardian.district.name if child.guardian.district else '',
            'تاريخ الإنشاء': child.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='الأبناء', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['الأبناء']
        style_excel_header(worksheet)
        auto_adjust_column_width(worksheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = f"children_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def export_districts_to_excel(districts_queryset=None):
    """تصدير الأحياء إلى Excel"""
    if districts_queryset is None:
        districts = District.objects.all()
    else:
        districts = districts_queryset
    
    data = []
    for district in districts:
        guardians_count = Guardian.objects.filter(district=district).count()
        data.append({
            'اسم الحي': district.name,
            'اسم المندوب': district.representative_name,
            'رقم جوال المندوب': district.representative_phone,
            'عدد أولياء الأمور': guardians_count,
            'تاريخ الإنشاء': district.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='الأحياء', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['الأحياء']
        style_excel_header(worksheet)
        auto_adjust_column_width(worksheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = f"districts_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def export_wives_to_excel(wives_queryset=None):
    """تصدير الزوجات إلى Excel"""
    if wives_queryset is None:
        wives = Wife.objects.select_related('guardian', 'guardian__district').all()
    else:
        wives = wives_queryset.select_related('guardian', 'guardian__district')
    
    data = []
    for wife in wives:
        data.append({
            'اسم الزوجة': wife.name,
            'رقم الهوية': wife.national_id or '',
            'اسم ولي الأمر': wife.guardian.name,
            'رقم هوية ولي الأمر': wife.guardian.national_id,
            'رقم جوال ولي الأمر': wife.guardian.phone_number,
            'الحي': wife.guardian.district.name if wife.guardian.district else '',
            'تاريخ الإنشاء': wife.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='الزوجات', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['الزوجات']
        style_excel_header(worksheet)
        auto_adjust_column_width(worksheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = f"wives_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def create_guardians_template():
    """إنشاء نموذج Excel لاستيراد أولياء الأمور"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # إنشاء ملف Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="guardians_import_template.xlsx"'
        
        # إنشاء نموذج بيانات
        template_data = [
            {
                'اسم ولي الأمر': 'محمد أحمد علي',
                'رقم الهوية': '123456789',
                'الجنس': 'ذكر',
                'الوظيفة': 'مهندس',
                'رقم الجوال': '0599123456',
                'الحالة الاجتماعية': 'متزوج',
                'عدد الأبناء': 3,
                'عدد الزوجات': 1,
                'حالة الإقامة': 'مقيم',
                'المحافظة الأصلية': 'غزة',
                'المدينة الأصلية': 'غزة',
                'عنوان النزوح': '',
                'الحي الحالي': 'الرمال'
            },
            {
                'اسم ولي الأمر': 'أحمد خالد سالم',
                'رقم الهوية': '987654321',
                'الجنس': 'ذكر',
                'الوظيفة': 'مدرس',
                'رقم الجوال': '0597888999',
                'الحالة الاجتماعية': 'متزوج',
                'عدد الأبناء': 4,
                'عدد الزوجات': 2,
                'حالة الإقامة': 'نازح',
                'المحافظة الأصلية': 'شمال غزة',
                'المدينة الأصلية': 'بيت لاهيا',
                'عنوان النزوح': 'مدرسة الأونروا - رفح',
                'الحي الحالي': 'الشجاعية'
            }
        ]
        
        df = pd.DataFrame(template_data)
        
        # إنشاء workbook وورقة عمل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نموذج أولياء الأمور"
        
        # إضافة البيانات
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # تنسيق الخط العربي Simplified Arabic
        arabic_font = Font(name='Simplified Arabic', size=12)
        
        # تنسيق العناوين
        header_font = Font(name='Simplified Arabic', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # لون أخضر
        header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, readingOrder=2)
        
        # تطبيق التنسيق على العناوين
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # إضافة حدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, readingOrder=2)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = arabic_font
                cell.alignment = data_alignment
                # تنسيق أرقام الهوية كنص
                if cell.column == 2:  # عمود رقم الهوية فقط
                    cell.number_format = '@'
                # إضافة حدود
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # تعيين اتجاه الورقة من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True
        
        # ضبط عرض الأعمدة تلقائياً
        auto_adjust_column_width(ws)
        
        # تجميد الصف الأول
        ws.freeze_panes = ws['A2']
        
        # إضافة ورقة التعليمات
        ws_instructions = wb.create_sheet('تعليمات الاستيراد')
        
        # محتوى التعليمات
        instructions = [
            ['تعليمات استيراد أولياء الأمور', ''],
            ['', ''],
            ['البيانات المطلوبة:', ''],
            ['• اسم ولي الأمر (مطلوب)', ''],
            ['• رقم الهوية (مطلوب - 9 أرقام)', ''],
            ['• الجنس: ذكر أو أنثى (مطلوب)', ''],
            ['• رقم الجوال (مطلوب)', ''],
            ['', ''],
            ['الحالات المقبولة:', ''],
            ['الحالة الاجتماعية:', 'متزوج، أرمل، مطلق، أعزب'],
            ['حالة الإقامة:', 'مقيم، نازح، لاجئ'],
            ['', ''],
            ['ملاحظات مهمة:', ''],
            ['• أرقام الهوية يجب أن تكون 9 أرقام بالضبط', ''],
            ['• تأكد من وجود الحي في النظام', ''],
            ['• البيانات الاختيارية يمكن تركها فارغة', ''],
            ['• احفظ الملف بصيغة .xlsx قبل الاستيراد', '']
        ]
        
        # إضافة التعليمات
        for row_num, (col1, col2) in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=col1)
            if col2:
                ws_instructions.cell(row=row_num, column=2, value=col2)
        
        # تنسيق ورقة التعليمات
        title_font = Font(name='Simplified Arabic', size=14, bold=True, color='000080')
        header_font_inst = Font(name='Simplified Arabic', size=12, bold=True)
        normal_font = Font(name='Simplified Arabic', size=11)
        
        # تنسيق العنوان
        ws_instructions.cell(row=1, column=1).font = title_font
        
        # تنسيق باقي النص
        for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
            for cell in row:
                if cell.value and ':' in str(cell.value):
                    cell.font = header_font_inst
                else:
                    cell.font = normal_font
                cell.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        
        # تعيين اتجاه ورقة التعليمات
        ws_instructions.sheet_view.rightToLeft = True
        
        # تعديل عرض أعمدة التعليمات
        ws_instructions.column_dimensions['A'].width = 35
        ws_instructions.column_dimensions['B'].width = 25
        
        # حفظ إلى buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except ImportError:
        # نظام بديل باستخدام pandas فقط
        template_data = {
            'اسم ولي الأمر': ['محمد أحمد علي', 'أحمد خالد سالم'],
            'رقم الهوية': ['123456789', '987654321'],
            'الجنس': ['ذكر', 'ذكر'],
            'الوظيفة': ['مهندس', 'مدرس'],
            'رقم الجوال': ['0599123456', '0597888999'],
            'الحالة الاجتماعية': ['متزوج', 'متزوج'],
            'عدد الأبناء': [3, 4],
            'عدد الزوجات': [1, 2],
            'حالة الإقامة': ['مقيم', 'نازح'],
            'المحافظة الأصلية': ['غزة', 'شمال غزة'],
            'المدينة الأصلية': ['غزة', 'بيت لاهيا'],
            'عنوان النزوح': ['', 'مدرسة الأونروا - رفح'],
            'الحي الحالي': ['الرمال', 'الشجاعية']
        }
        
        df = pd.DataFrame(template_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='نموذج أولياء الأمور', index=False)
            
            # تنسيق الملف
            worksheet = writer.sheets['نموذج أولياء الأمور']
            style_excel_header(worksheet)
            
            # تنسيق أعمدة أرقام الهوية كنص
            for row in range(2, worksheet.max_row + 1):
                # عمود رقم الهوية الأساسي فقط
                cell = worksheet.cell(row=row, column=2)
                cell.number_format = '@'  # تنسيق نص
            
            auto_adjust_column_width(worksheet)
            
            # إضافة تعليمات
            instructions_sheet = writer.book.create_sheet('تعليمات الاستيراد')
            instructions = [
                ['تعليمات استيراد أولياء الأمور'],
                [''],
                ['1. املأ البيانات في ورقة "نموذج أولياء الأمور"'],
                ['2. تأكد من صحة أرقام الهوية (9 أرقام)'],
                ['3. تأكد من صحة أرقام الجوال'],
                ['4. الجنس: ذكر أو أنثى (مطلوب)'],
                ['5. الحالة الاجتماعية: متزوج، أرمل، مطلق، أعزب'],
                ['6. حالة الإقامة: مقيم، نازح، لاجئ'],
                ['7. تأكد من وجود الحي في النظام'],
                ['8. احفظ الملف واستورده من النظام'],
                [''],
                ['تنبيه مهم: تأكد من أن أرقام الهوية مكتوبة كنص وليس كأرقام'],
                ['لتجنب إضافة .0 في النهاية']
            ]
            
            for row_num, instruction in enumerate(instructions, 1):
                instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
            
            # تنسيق ورقة التعليمات
            instructions_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            auto_adjust_column_width(instructions_sheet)
        
        output.seek(0)
        
        # إنشاء الاستجابة
        filename = "guardians_import_template.xlsx"
        response = create_excel_response(filename)
        response.write(output.getvalue())
        
        return response

def create_children_template():
    """إنشاء نموذج Excel لاستيراد الأبناء"""
    # تحديد ترتيب الأعمدة بوضوح
    columns = [
        'اسم الطفل', 
        'رقم الهوية', 
        'تاريخ الميلاد', 
        'رقم هوية ولي الأمر'
    ]
    
    template_data = [
        ['أحمد محمد علي', '987654321', '2015-05-15', '123456789'],
        ['فاطمة خالد سالم', '876543210', '2017-03-20', '123456789'],
        ['محمد أحمد حسن', '765432109', '2016-08-10', '234567890']
    ]
    
    df = pd.DataFrame(template_data, columns=columns)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='نموذج الأبناء', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['نموذج الأبناء']
        style_excel_header(worksheet)
        
        # تنسيق أعمدة أرقام الهوية كنص
        for row in range(2, worksheet.max_row + 1):
            # عمود رقم الهوية (العمود الثاني)
            cell = worksheet.cell(row=row, column=2)
            cell.number_format = '@'  # تنسيق نص
            # عمود رقم هوية ولي الأمر (العمود الرابع - آخر عمود)
            cell = worksheet.cell(row=row, column=4)
            cell.number_format = '@'  # تنسيق نص
        
        auto_adjust_column_width(worksheet)
        
        # إضافة تعليمات
        instructions_sheet = writer.book.create_sheet('تعليمات الاستيراد')
        instructions = [
            ['تعليمات استيراد الأبناء'],
            [''],
            ['1. املأ البيانات في ورقة "نموذج الأبناء"'],
            ['2. تأكد من صحة أرقام الهوية (9 أرقام)'],
            ['3. تاريخ الميلاد بصيغة: YYYY-MM-DD'],
            ['4. تأكد من وجود ولي الأمر في النظام'],
            ['5. احفظ الملف واستورده من النظام'],
            [''],
            ['الأعمدة المطلوبة:'],
            ['- اسم الطفل (مطلوب)'],
            ['- رقم الهوية (مطلوب)'],
            ['- تاريخ الميلاد (مطلوب)'],
            ['- رقم هوية ولي الأمر (مطلوب)'],
            [''],
            ['تنبيه مهم: تأكد من أن أرقام الهوية مكتوبة كنص وليس كأرقام'],
            ['لتجنب إضافة .0 في النهاية']
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
        
        # تنسيق ورقة التعليمات
        instructions_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        auto_adjust_column_width(instructions_sheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = "children_import_template.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def create_districts_template():
    """إنشاء نموذج Excel لاستيراد الأحياء"""
    template_data = {
        'اسم الحي': ['الرمال'],
        'اسم المندوب': ['خالد أحمد'],
        'رقم جوال المندوب': ['0599123456']
    }
    
    df = pd.DataFrame(template_data)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='نموذج الأحياء', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets['نموذج الأحياء']
        style_excel_header(worksheet)
        auto_adjust_column_width(worksheet)
        
        # إضافة تعليمات
        instructions_sheet = writer.book.create_sheet('تعليمات الاستيراد')
        instructions = [
            ['تعليمات استيراد الأحياء'],
            [''],
            ['1. املأ البيانات في ورقة "نموذج الأحياء"'],
            ['2. تأكد من عدم تكرار أسماء الأحياء'],
            ['3. تأكد من صحة رقم جوال المندوب'],
            ['4. احفظ الملف واستورده من النظام']
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
        
        # تنسيق ورقة التعليمات
        instructions_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        auto_adjust_column_width(instructions_sheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = "districts_import_template.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def create_wives_template():
    """إنشاء نموذج Excel لاستيراد الزوجات"""
    # إنشاء البيانات مع أسماء الأعمدة واضحة
    columns = ['اسم الزوجة', 'رقم الهوية', 'رقم هوية ولي الأمر']
    sample_data = [
        ['فاطمة أحمد علي', '987654321', '123456789'],
        ['مريم محمد سالم', '876543210', '234567890']
    ]
    
    # إنشاء DataFrame مع أسماء الأعمدة
    df = pd.DataFrame(sample_data, columns=columns)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # كتابة البيانات مع التأكد من وجود أسماء الأعمدة
        df.to_excel(writer, sheet_name='نموذج الزوجات', index=False, header=True)
        
        # الحصول على ورقة العمل
        worksheet = writer.sheets['نموذج الزوجات']
        
        # تنسيق رأس الجدول
        style_excel_header(worksheet, header_row=1)
        
        # تنسيق أعمدة أرقام الهوية كنص
        for row in range(2, worksheet.max_row + 1):
            # عمود رقم الهوية (العمود الثاني)
            cell = worksheet.cell(row=row, column=2)
            cell.number_format = '@'  # تنسيق نص
            # عمود رقم هوية ولي الأمر (العمود الثالث)
            cell = worksheet.cell(row=row, column=3)
            cell.number_format = '@'  # تنسيق نص
        
        # تنسيق أعمدة الرأس أيضاً كنص
        worksheet.cell(row=1, column=2).number_format = '@'
        worksheet.cell(row=1, column=3).number_format = '@'
        
        # ضبط عرض الأعمدة تلقائياً
        auto_adjust_column_width(worksheet)
        
        # إضافة تعليمات في ورقة منفصلة
        instructions_sheet = writer.book.create_sheet('تعليمات الاستيراد')
        instructions = [
            ['تعليمات استيراد الزوجات'],
            [''],
            ['1. املأ البيانات في ورقة "نموذج الزوجات"'],
            ['2. تأكد من صحة أرقام الهوية (9 أرقام)'],
            ['3. رقم الهوية للزوجة اختياري (يمكن تركه فارغاً)'],
            ['4. تأكد من وجود ولي الأمر في النظام قبل الاستيراد'],
            ['5. احفظ الملف واستورده من النظام'],
            [''],
            ['الأعمدة المطلوبة:'],
            ['- اسم الزوجة (مطلوب)'],
            ['- رقم الهوية (اختياري)'],
            ['- رقم هوية ولي الأمر (مطلوب)'],
            [''],
            ['تنبيه مهم: تأكد من أن أرقام الهوية مكتوبة كنص وليس كأرقام'],
            ['لتجنب إضافة .0 في النهاية']
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
        
        # تنسيق ورقة التعليمات
        instructions_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        for row_num in range(9, 12):  # تنسيق عناوين الأعمدة
            instructions_sheet.cell(row=row_num, column=1).font = Font(bold=True)
        
        auto_adjust_column_width(instructions_sheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    filename = "wives_import_template.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response

def clean_national_id(value):
    """تنظيف رقم الهوية من المشاكل الناتجة عن Excel"""
    if pd.isna(value):
        return ''
    
    # تحويل إلى نص أولاً
    str_value = str(value)
    
    # إزالة .0 إذا كانت موجودة في النهاية
    if str_value.endswith('.0'):
        str_value = str_value[:-2]
    
    # إزالة أي مسافات
    str_value = str_value.strip()
    
    # التأكد من أن الرقم يحتوي على 9 أرقام فقط
    if str_value.isdigit() and len(str_value) == 9:
        return str_value
    
    return str_value

def import_guardians_from_excel(file):
    """استيراد أولياء الأمور من ملف Excel"""
    try:
        df = pd.read_excel(file, sheet_name=0)
        
        success_count = 0
        error_messages = []
        detailed_errors = []  # لتجميع الأخطاء المفصلة
        
        for index, row in df.iterrows():
            try:
                # التحقق من البيانات المطلوبة
                if pd.isna(row['اسم ولي الأمر']) or pd.isna(row['رقم الهوية']):
                    error_msg = "الاسم ورقم الهوية مطلوبان"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'الاسم': row.get('اسم ولي الأمر', 'غير محدد'),
                        'رقم الهوية': clean_national_id(row.get('رقم الهوية', '')),
                        'الجنس': row.get('الجنس', 'غير محدد'),
                        'رقم الجوال': row.get('رقم الجوال', 'غير محدد'),
                        'الحي': row.get('الحي الحالي', 'غير محدد'),
                        'نوع الخطأ': 'بيانات مفقودة',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # التحقق من الجنس
                if pd.isna(row['الجنس']) or row['الجنس'] not in ['ذكر', 'أنثى']:
                    error_msg = "الجنس مطلوب ويجب أن يكون 'ذكر' أو 'أنثى'"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'الاسم': row.get('اسم ولي الأمر', 'غير محدد'),
                        'رقم الهوية': clean_national_id(row.get('رقم الهوية', '')),
                        'الجنس': row.get('الجنس', 'غير محدد'),
                        'رقم الجوال': row.get('رقم الجوال', 'غير محدد'),
                        'الحي': row.get('الحي الحالي', 'غير محدد'),
                        'نوع الخطأ': 'جنس غير صحيح',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # البحث عن الحي
                district = None
                if not pd.isna(row['الحي الحالي']):
                    try:
                        district = District.objects.get(name=row['الحي الحالي'])
                    except District.DoesNotExist:
                        error_msg = f"الحي '{row['الحي الحالي']}' غير موجود"
                        error_messages.append(f"الصف {index + 2}: {error_msg}")
                        detailed_errors.append({
                            'رقم الصف': index + 2,
                            'الاسم': row.get('اسم ولي الأمر', 'غير محدد'),
                            'رقم الهوية': clean_national_id(row.get('رقم الهوية', '')),
                            'الجنس': row.get('الجنس', 'غير محدد'),
                            'رقم الجوال': row.get('رقم الجوال', 'غير محدد'),
                            'الحي': row.get('الحي الحالي', 'غير محدد'),
                            'نوع الخطأ': 'حي غير موجود',
                            'تفاصيل الخطأ': error_msg
                        })
                        continue
                
                # تحويل الحالة الاجتماعية - إزالة التحويل واستخدام القيم مباشرة
                marital_status = row.get('الحالة الاجتماعية', 'أعزب')
                # التحقق من صحة القيمة
                valid_marital_statuses = ['متزوج', 'أرمل', 'مطلق', 'أعزب']
                if marital_status not in valid_marital_statuses:
                    marital_status = 'أعزب'  # قيمة افتراضية
                
                # تحويل حالة الإقامة - إزالة التحويل واستخدام القيم مباشرة
                residence_status = row.get('حالة الإقامة', 'مقيم')
                # التحقق من صحة القيمة
                valid_residence_statuses = ['مقيم', 'نازح']
                if residence_status not in valid_residence_statuses:
                    residence_status = 'مقيم'  # قيمة افتراضية
                
                # تحديد عدد الأبناء والزوجات من ملف Excel
                children_count_excel = int(row['عدد الأبناء']) if not pd.isna(row['عدد الأبناء']) else 0
                wives_count_excel = int(row['عدد الزوجات']) if not pd.isna(row['عدد الزوجات']) else 0
                family_members_count_excel = 1 + children_count_excel + wives_count_excel
                
                # التحقق من وجود ولي الأمر مسبقاً
                national_id = clean_national_id(row['رقم الهوية'])
                try:
                    guardian = Guardian.objects.get(national_id=national_id)
                    # تحديث البيانات الموجودة
                    guardian.name = row['اسم ولي الأمر']
                    guardian.current_job = row['الوظيفة'] if not pd.isna(row['الوظيفة']) else ''
                    guardian.phone_number = str(row['رقم الجوال']) if not pd.isna(row['رقم الجوال']) else ''
                    guardian.gender = row['الجنس']
                    guardian.marital_status = marital_status
                    guardian.residence_status = residence_status
                    guardian.original_governorate = row['المحافظة الأصلية'] if not pd.isna(row['المحافظة الأصلية']) else ''
                    guardian.original_city = row['المدينة الأصلية'] if not pd.isna(row['المدينة الأصلية']) else ''
                    guardian.displacement_address = row['عنوان النزوح'] if not pd.isna(row['عنوان النزوح']) else ''
                    guardian.district = district
                    # تحديث العدادات من ملف Excel
                    guardian.children_count = children_count_excel
                    guardian.wives_count = wives_count_excel
                    guardian.family_members_count = family_members_count_excel
                    guardian.save()
                except Guardian.DoesNotExist:
                    # إنشاء ولي أمر جديد
                    guardian = Guardian.objects.create(
                        name=row['اسم ولي الأمر'],
                        national_id=national_id,
                        current_job=row['الوظيفة'] if not pd.isna(row['الوظيفة']) else '',
                        phone_number=str(row['رقم الجوال']) if not pd.isna(row['رقم الجوال']) else '',
                        gender=row['الجنس'],
                        marital_status=marital_status,
                        residence_status=residence_status,
                        original_governorate=row['المحافظة الأصلية'] if not pd.isna(row['المحافظة الأصلية']) else '',
                        original_city=row['المدينة الأصلية'] if not pd.isna(row['المدينة الأصلية']) else '',
                        displacement_address=row['عنوان النزوح'] if not pd.isna(row['عنوان النزوح']) else '',
                        district=district,
                        # تعيين العدادات من ملف Excel
                        children_count=children_count_excel,
                        wives_count=wives_count_excel,
                        family_members_count=family_members_count_excel
                    )
                success_count += 1
                
            except Exception as e:
                error_msg = str(e)
                error_messages.append(f"الصف {index + 2}: {error_msg}")
                detailed_errors.append({
                    'رقم الصف': index + 2,
                    'الاسم': row.get('اسم ولي الأمر', 'غير محدد'),
                    'رقم الهوية': clean_national_id(row.get('رقم الهوية', '')),
                    'الجنس': row.get('الجنس', 'غير محدد'),
                    'رقم الجوال': row.get('رقم الجوال', 'غير محدد'),
                    'الحي': row.get('الحي الحالي', 'غير محدد'),
                    'نوع الخطأ': 'خطأ عام',
                    'تفاصيل الخطأ': error_msg
                })
        
        return {
            'success': True,
            'success_count': success_count,
            'error_messages': error_messages,
            'detailed_errors': detailed_errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'error_message': f"خطأ في قراءة الملف: {str(e)}",
            'detailed_errors': []
        }

def import_children_from_excel(file):
    """استيراد الأبناء من ملف Excel"""
    try:
        df = pd.read_excel(file, sheet_name=0)
        
        success_count = 0
        error_messages = []
        detailed_errors = []  # لتجميع الأخطاء المفصلة
        
        for index, row in df.iterrows():
            try:
                # التحقق من البيانات المطلوبة
                if pd.isna(row['اسم الطفل']) or pd.isna(row['رقم هوية ولي الأمر']):
                    error_msg = "اسم الطفل ورقم هوية ولي الأمر مطلوبان"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الطفل': str(row.get('اسم الطفل', 'غير محدد')),
                        'رقم هوية الطفل': clean_national_id(str(row.get('رقم الهوية', ''))),
                        'رقم هوية ولي الأمر': clean_national_id(str(row.get('رقم هوية ولي الأمر', ''))),
                        'تاريخ الميلاد': str(row.get('تاريخ الميلاد', 'غير محدد')),
                        'نوع الخطأ': 'بيانات مفقودة',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # البحث عن ولي الأمر
                try:
                    guardian_id = clean_national_id(row['رقم هوية ولي الأمر'])
                    guardian = Guardian.objects.get(national_id=guardian_id)
                except Guardian.DoesNotExist:
                    error_msg = f"ولي الأمر برقم الهوية '{guardian_id}' غير موجود"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الطفل': str(row.get('اسم الطفل', 'غير محدد')),
                        'رقم هوية الطفل': clean_national_id(str(row.get('رقم الهوية', ''))),
                        'رقم هوية ولي الأمر': guardian_id,
                        'تاريخ الميلاد': str(row.get('تاريخ الميلاد', 'غير محدد')),
                        'نوع الخطأ': 'ولي أمر غير موجود',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # تحويل تاريخ الميلاد
                birth_date = None
                if not pd.isna(row['تاريخ الميلاد']):
                    try:
                        birth_date = pd.to_datetime(row['تاريخ الميلاد']).date()
                    except:
                        error_msg = "تاريخ الميلاد غير صحيح"
                        error_messages.append(f"الصف {index + 2}: {error_msg}")
                        detailed_errors.append({
                            'رقم الصف': index + 2,
                            'اسم الطفل': str(row.get('اسم الطفل', 'غير محدد')),
                            'رقم هوية الطفل': clean_national_id(str(row.get('رقم الهوية', ''))),
                            'رقم هوية ولي الأمر': guardian_id,
                            'اسم ولي الأمر': guardian.name,
                            'تاريخ الميلاد': str(row.get('تاريخ الميلاد', 'غير محدد')),
                            'نوع الخطأ': 'تاريخ ميلاد غير صحيح',
                            'تفاصيل الخطأ': error_msg
                        })
                        continue
                
                # تحديد الجنس
                gender = 'ذكر'  # افتراضي
                if not pd.isna(row.get('الجنس', '')):
                    gender_value = str(row['الجنس']).strip()
                    if gender_value in ['أنثى', 'انثى', 'female', 'F', 'f']:
                        gender = 'أنثى'
                    else:
                        gender = 'ذكر'
                
                # إنشاء الطفل
                child = Child.objects.create(
                    name=row['اسم الطفل'],
                    national_id=clean_national_id(row['رقم الهوية']) if not pd.isna(row['رقم الهوية']) else '',
                    birth_date=birth_date,
                    gender=gender,
                    guardian=guardian
                )
                
                # لا نحدث العدادات - نحافظ على القيم من ملف Excel الأصلي
                
                success_count += 1
                
            except Exception as e:
                error_msg = str(e)
                error_messages.append(f"الصف {index + 2}: {error_msg}")
                detailed_errors.append({
                    'رقم الصف': index + 2,
                    'اسم الطفل': str(row.get('اسم الطفل', 'غير محدد')),
                    'رقم هوية الطفل': clean_national_id(str(row.get('رقم الهوية', ''))),
                    'رقم هوية ولي الأمر': clean_national_id(str(row.get('رقم هوية ولي الأمر', ''))),
                    'تاريخ الميلاد': str(row.get('تاريخ الميلاد', 'غير محدد')),
                    'نوع الخطأ': 'خطأ عام',
                    'تفاصيل الخطأ': error_msg
                })
        
        return {
            'success': True,
            'success_count': success_count,
            'error_messages': error_messages,
            'detailed_errors': detailed_errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'error_message': f"خطأ في قراءة الملف: {str(e)}",
            'detailed_errors': []
        }

def import_districts_from_excel(file):
    """استيراد الأحياء من ملف Excel"""
    try:
        df = pd.read_excel(file, sheet_name=0)
        
        success_count = 0
        error_messages = []
        detailed_errors = []  # لتجميع الأخطاء المفصلة
        
        for index, row in df.iterrows():
            try:
                # التحقق من البيانات المطلوبة
                if pd.isna(row['اسم الحي']):
                    error_msg = "اسم الحي مطلوب"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الحي': str(row.get('اسم الحي', 'غير محدد')),
                        'اسم المندوب': str(row.get('اسم المندوب', 'غير محدد')),
                        'رقم جوال المندوب': str(row.get('رقم جوال المندوب', 'غير محدد')),
                        'نوع الخطأ': 'بيانات مفقودة',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # التحقق من عدم تكرار الحي
                if District.objects.filter(name=row['اسم الحي']).exists():
                    error_msg = f"الحي '{row['اسم الحي']}' موجود مسبقاً"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الحي': str(row.get('اسم الحي', 'غير محدد')),
                        'اسم المندوب': str(row.get('اسم المندوب', 'غير محدد')),
                        'رقم جوال المندوب': str(row.get('رقم جوال المندوب', 'غير محدد')),
                        'نوع الخطأ': 'حي مكرر',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # إنشاء الحي
                district = District.objects.create(
                    name=row['اسم الحي'],
                    representative_name=row['اسم المندوب'] if not pd.isna(row['اسم المندوب']) else '',
                    representative_phone=str(row['رقم جوال المندوب']) if not pd.isna(row['رقم جوال المندوب']) else ''
                )
                success_count += 1
                
            except Exception as e:
                error_msg = str(e)
                error_messages.append(f"الصف {index + 2}: {error_msg}")
                detailed_errors.append({
                    'رقم الصف': index + 2,
                    'اسم الحي': str(row.get('اسم الحي', 'غير محدد')),
                    'اسم المندوب': str(row.get('اسم المندوب', 'غير محدد')),
                    'رقم جوال المندوب': str(row.get('رقم جوال المندوب', 'غير محدد')),
                    'نوع الخطأ': 'خطأ عام',
                    'تفاصيل الخطأ': error_msg
                })
        
        return {
            'success': True,
            'success_count': success_count,
            'error_messages': error_messages,
            'detailed_errors': detailed_errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'error_message': f"خطأ في قراءة الملف: {str(e)}",
            'detailed_errors': []
        }

def import_wives_from_excel(file):
    """استيراد الزوجات من ملف Excel"""
    try:
        df = pd.read_excel(file, sheet_name=0)
        
        success_count = 0
        error_messages = []
        detailed_errors = []  # لتجميع الأخطاء المفصلة
        
        for index, row in df.iterrows():
            try:
                # التحقق من البيانات المطلوبة
                if pd.isna(row['اسم الزوجة']) or pd.isna(row['رقم هوية ولي الأمر']):
                    error_msg = "اسم الزوجة ورقم هوية ولي الأمر مطلوبان"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الزوجة': str(row.get('اسم الزوجة', 'غير محدد')),
                        'رقم هوية الزوجة': clean_national_id(str(row.get('رقم الهوية', ''))),
                        'رقم هوية ولي الأمر': clean_national_id(str(row.get('رقم هوية ولي الأمر', ''))),
                        'نوع الخطأ': 'بيانات مفقودة',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # البحث عن ولي الأمر
                try:
                    guardian_id = clean_national_id(row['رقم هوية ولي الأمر'])
                    guardian = Guardian.objects.get(national_id=guardian_id)
                except Guardian.DoesNotExist:
                    error_msg = f"ولي الأمر برقم الهوية '{guardian_id}' غير موجود"
                    error_messages.append(f"الصف {index + 2}: {error_msg}")
                    detailed_errors.append({
                        'رقم الصف': index + 2,
                        'اسم الزوجة': str(row.get('اسم الزوجة', 'غير محدد')),
                        'رقم هوية الزوجة': clean_national_id(str(row.get('رقم الهوية', ''))),
                        'رقم هوية ولي الأمر': guardian_id,
                        'نوع الخطأ': 'ولي أمر غير موجود',
                        'تفاصيل الخطأ': error_msg
                    })
                    continue
                
                # التحقق من رقم هوية الزوجة إذا تم توفيره
                national_id = ''
                if not pd.isna(row['رقم الهوية']):
                    national_id = clean_national_id(row['رقم الهوية'])
                    # التحقق من عدم تكرار رقم الهوية للزوجة مع نفس ولي الأمر
                    if Wife.objects.filter(guardian=guardian, national_id=national_id).exists():
                        error_msg = f"زوجة برقم الهوية '{national_id}' موجودة مسبقاً لنفس ولي الأمر"
                        error_messages.append(f"الصف {index + 2}: {error_msg}")
                        detailed_errors.append({
                            'رقم الصف': index + 2,
                            'اسم الزوجة': str(row.get('اسم الزوجة', 'غير محدد')),
                            'رقم هوية الزوجة': national_id,
                            'رقم هوية ولي الأمر': guardian_id,
                            'اسم ولي الأمر': guardian.name,
                            'نوع الخطأ': 'زوجة مكررة',
                            'تفاصيل الخطأ': error_msg
                        })
                        continue
                
                # إنشاء الزوجة
                wife = Wife.objects.create(
                    name=row['اسم الزوجة'],
                    national_id=national_id,
                    guardian=guardian
                )
                
                # لا نحدث العدادات - نحافظ على القيم من ملف Excel الأصلي
                
                success_count += 1
                
            except Exception as e:
                error_msg = str(e)
                error_messages.append(f"الصف {index + 2}: {error_msg}")
                detailed_errors.append({
                    'رقم الصف': index + 2,
                    'اسم الزوجة': str(row.get('اسم الزوجة', 'غير محدد')),
                    'رقم هوية الزوجة': clean_national_id(str(row.get('رقم الهوية', ''))),
                    'رقم هوية ولي الأمر': clean_national_id(str(row.get('رقم هوية ولي الأمر', ''))),
                    'نوع الخطأ': 'خطأ عام',
                    'تفاصيل الخطأ': error_msg
                })
        
        return {
            'success': True,
            'success_count': success_count,
            'error_messages': error_messages,
            'detailed_errors': detailed_errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'error_message': f"خطأ في قراءة الملف: {str(e)}",
            'detailed_errors': []
        }

def create_import_errors_excel(errors_data, import_type):
    """إنشاء ملف Excel يحتوي على أخطاء الاستيراد"""
    if not errors_data:
        return None
    
    # إنشاء DataFrame مع أخطاء الاستيراد
    df = pd.DataFrame(errors_data)
    
    # إنشاء ملف Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'أخطاء استيراد {import_type}', index=False)
        
        # تنسيق الملف
        worksheet = writer.sheets[f'أخطاء استيراد {import_type}']
        style_excel_header(worksheet)
        
        # تلوين صفوف الأخطاء
        error_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = error_fill
        
        auto_adjust_column_width(worksheet)
        
        # إضافة ملخص الأخطاء
        summary_sheet = writer.book.create_sheet('ملخص الأخطاء')
        summary_data = [
            [f'ملخص أخطاء استيراد {import_type}'],
            [''],
            [f'إجمالي الأخطاء: {len(errors_data)}'],
            [''],
            ['أنواع الأخطاء:']
        ]
        
        # تحليل أنواع الأخطاء
        error_types = {}
        for error in errors_data:
            error_msg = error.get('نوع الخطأ', 'غير محدد')
            error_types[error_msg] = error_types.get(error_msg, 0) + 1
        
        for error_type, count in error_types.items():
            summary_data.append([f'- {error_type}: {count}'])
        
        for row_num, data in enumerate(summary_data, 1):
            summary_sheet.cell(row=row_num, column=1, value=data[0])
        
        # تنسيق ورقة الملخص
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        auto_adjust_column_width(summary_sheet)
    
    output.seek(0)
    
    # إنشاء الاستجابة
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"import_errors_{import_type}_{timestamp}.xlsx"
    response = create_excel_response(filename)
    response.write(output.getvalue())
    
    return response 