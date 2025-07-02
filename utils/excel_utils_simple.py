from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from basic_data.models import Guardian, District, Wife, Child
from datetime import datetime

def create_excel_response(filename):
    """إنشاء استجابة HTTP لملف Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def style_excel_header(worksheet, header_row=1):
    """تنسيق رأس الجدول في Excel"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[header_row]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

def auto_adjust_column_width(worksheet):
    """ضبط عرض الأعمدة تلقائياً مع مراعاة المحتوى العربي"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    # حساب طول النص مع مراعاة النصوص العربية
                    cell_length = len(str(cell.value))
                    # إضافة مساحة إضافية للنصوص العربية
                    if any('\u0600' <= char <= '\u06FF' for char in str(cell.value)):
                        cell_length = int(cell_length * 1.3)  # زيادة 30% للنصوص العربية
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # تحديد عرض مناسب (بحد أدنى 8 وحد أقصى 50)
        adjusted_width = max(min(max_length + 3, 50), 8)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_guardians_to_excel(guardians_queryset=None):
    """تصدير أولياء الأمور إلى Excel مع التنسيق العربي المحسن"""
    if guardians_queryset is None:
        guardians = Guardian.objects.select_related('district').prefetch_related('wives').all()
    else:
        guardians = guardians_queryset.select_related('district').prefetch_related('wives')
    
    try:
        # إنشاء ملف Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="guardians_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # إعداد الترجمات العربية الشاملة
        gender_translation = {
            'male': 'ذكر',
            'female': 'أنثى',
            'M': 'ذكر',
            'F': 'أنثى',
            'ذكر': 'ذكر',
            'أنثى': 'أنثى'
        }
        
        marital_status_translation = {
            'single': 'أعزب',
            'married': 'متزوج',
            'divorced': 'مطلق',
            'widowed': 'أرمل',
            'أعزب': 'أعزب',
            'متزوج': 'متزوج',
            'مطلق': 'مطلق',
            'أرمل': 'أرمل'
        }
        
        residence_status_translation = {
            'resident': 'مقيم',
            'displaced': 'نازح',
            'refugee': 'لاجئ',
            'مقيم': 'مقيم',
            'نازح': 'نازح',
            'لاجئ': 'لاجئ'
        }
        
        # إنشاء workbook وورقة عمل
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أولياء الأمور"
        
        # إضافة العناوين
        headers = [
            'اسم ولي الأمر', 'رقم الهوية', 'الجنس', 'الوظيفة', 'رقم الجوال',
            'الحالة الاجتماعية', 'عدد الأبناء', 'عدد الزوجات',
            'اسم الزوجة 1', 'رقم هوية الزوجة 1',
            'اسم الزوجة 2', 'رقم هوية الزوجة 2',
            'اسم الزوجة 3', 'رقم هوية الزوجة 3',
            'اسم الزوجة 4', 'رقم هوية الزوجة 4',
            'عدد أفراد العائلة', 'حالة الإقامة', 'المحافظة الأصلية',
            'المدينة الأصلية', 'عنوان النزوح', 'الحي الحالي', 'تاريخ الإنشاء'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # إضافة البيانات
        row_num = 2
        for guardian in guardians:
            # البحث عن معلومات الزوجات
            wives_info = {}
            
            # تهيئة جميع مفاتيح الزوجات أولاً
            for i in range(1, 5):
                wives_info[f'wife_{i}_name'] = ''
                wives_info[f'wife_{i}_id'] = ''
            
            # ملء معلومات الزوجات
            wives = guardian.wives.all()[:4]  # حتى 4 زوجات
            for i, wife in enumerate(wives, 1):
                wives_info[f'wife_{i}_name'] = wife.name
                wives_info[f'wife_{i}_id'] = wife.national_id or ''
            
            # تحويل القيم للعربية
            gender_arabic = gender_translation.get(guardian.gender, guardian.gender)
            marital_status_arabic = marital_status_translation.get(guardian.marital_status, guardian.marital_status)
            residence_status_arabic = residence_status_translation.get(guardian.residence_status, guardian.residence_status)
            
            # إضافة البيانات للصف
            data_row = [
                guardian.name,
                guardian.national_id,
                gender_arabic,
                guardian.current_job or '',
                guardian.phone_number,
                marital_status_arabic,
                guardian.children_count,
                guardian.wives_count,
                wives_info['wife_1_name'],
                wives_info['wife_1_id'],
                wives_info['wife_2_name'],
                wives_info['wife_2_id'],
                wives_info['wife_3_name'],
                wives_info['wife_3_id'],
                wives_info['wife_4_name'],
                wives_info['wife_4_id'],
                guardian.family_members_count,
                residence_status_arabic,
                guardian.original_governorate or '',
                guardian.original_city or '',
                guardian.displacement_address or '',
                guardian.district.name if guardian.district else '',
                guardian.created_at.strftime('%Y-%m-%d %H:%M:%S') if guardian.created_at else '',
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # تنسيق الخط العربي Simplified Arabic
        arabic_font = Font(name='Simplified Arabic', size=12)
        
        # تنسيق العناوين
        header_font = Font(name='Simplified Arabic', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='17A2B8', end_color='17A2B8', fill_type='solid')  # لون أزرق فيروزي
        header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, readingOrder=2)
        
        # تطبيق التنسيق على العناوين
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # إضافة حدود
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, readingOrder=2)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = arabic_font
                cell.alignment = data_alignment
                # إضافة حدود
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # ضبط عرض الأعمدة
        auto_adjust_column_width(ws)
        
        # حفظ الملف
        wb.save(response)
        return response
        
    except Exception as e:
        # في حالة الخطأ، إرجاع رسالة خطأ
        error_response = HttpResponse(f"خطأ في تصدير البيانات: {str(e)}", content_type='text/plain')
        return error_response

def export_children_to_excel(children_queryset=None):
    """تصدير الأبناء إلى Excel"""
    if children_queryset is None:
        children = Child.objects.select_related('guardian').all()
    else:
        children = children_queryset.select_related('guardian')
    
    try:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="children_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأبناء"
        
        # إضافة العناوين
        headers = ['اسم الطفل', 'رقم الهوية', 'رقم هوية ولي الأمر', 'اسم ولي الأمر', 'تاريخ الميلاد', 'الجنس']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # إضافة البيانات
        row_num = 2
        for child in children:
            data_row = [
                child.name,
                child.national_id or '',
                child.guardian.national_id if child.guardian else '',
                child.guardian.name if child.guardian else '',
                child.birth_date.strftime('%Y-%m-%d') if child.birth_date else '',
                child.gender or ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # تطبيق التنسيق
        style_excel_header(ws)
        auto_adjust_column_width(ws)
        
        wb.save(response)
        return response
        
    except Exception as e:
        error_response = HttpResponse(f"خطأ في تصدير البيانات: {str(e)}", content_type='text/plain')
        return error_response

def export_districts_to_excel(districts_queryset=None):
    """تصدير الأحياء إلى Excel"""
    if districts_queryset is None:
        districts = District.objects.all()
    else:
        districts = districts_queryset
    
    try:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="districts_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأحياء"
        
        # إضافة العناوين
        headers = ['اسم الحي', 'اسم المندوب', 'رقم جوال المندوب']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # إضافة البيانات
        row_num = 2
        for district in districts:
            data_row = [
                district.name,
                district.representative_name or '',
                district.representative_phone or ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # تطبيق التنسيق
        style_excel_header(ws)
        auto_adjust_column_width(ws)
        
        wb.save(response)
        return response
        
    except Exception as e:
        error_response = HttpResponse(f"خطأ في تصدير البيانات: {str(e)}", content_type='text/plain')
        return error_response

def export_wives_to_excel(wives_queryset=None):
    """تصدير الزوجات إلى Excel"""
    if wives_queryset is None:
        wives = Wife.objects.select_related('guardian').all()
    else:
        wives = wives_queryset.select_related('guardian')
    
    try:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="wives_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الزوجات"
        
        # إضافة العناوين
        headers = ['اسم الزوجة', 'رقم الهوية', 'رقم هوية ولي الأمر', 'اسم ولي الأمر']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # إضافة البيانات
        row_num = 2
        for wife in wives:
            data_row = [
                wife.name,
                wife.national_id or '',
                wife.guardian.national_id if wife.guardian else '',
                wife.guardian.name if wife.guardian else ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # تطبيق التنسيق
        style_excel_header(ws)
        auto_adjust_column_width(ws)
        
        wb.save(response)
        return response
        
    except Exception as e:
        error_response = HttpResponse(f"خطأ في تصدير البيانات: {str(e)}", content_type='text/plain')
        return error_response 