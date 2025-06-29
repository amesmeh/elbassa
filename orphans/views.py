from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.urls import reverse
from .models import Orphan
from basic_data.models import District, Guardian
from .forms import OrphanForm, OrphanSearchForm, OrphanImportForm
from .utils import search_orphans
import pandas as pd
from django.core.exceptions import ValidationError
from datetime import date
from utils.excel_utils import create_excel_response, auto_adjust_column_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

def orphans_list(request):
    """عرض قائمة الأيتام مع البحث والفلترة"""
    form = OrphanSearchForm(request.GET)
    orphans = Orphan.objects.all().select_related('district')
    
    if form.is_valid():
        # البحث المرن
        search_query = form.cleaned_data.get('search')
        if search_query:
            orphans = search_orphans(search_query)
        
        # فلترة حسب الجنس
        gender = form.cleaned_data.get('gender')
        if gender:
            orphans = orphans.filter(gender=gender)
        
        # فلترة حسب الحي
        district = form.cleaned_data.get('district')
        if district:
            orphans = orphans.filter(district=district)
        
        # فلترة حسب الحالة الصحية
        health_status = form.cleaned_data.get('health_status')
        if health_status:
            orphans = orphans.filter(health_status=health_status)
        
        # فلترة حسب المرحلة التعليمية
        education_level = form.cleaned_data.get('education_level')
        if education_level:
            orphans = orphans.filter(education_level=education_level)
        
        # فلترة حسب الفئة العمرية
        age_range = form.cleaned_data.get('age_range')
        if age_range:
            today = date.today()
            if age_range == '0-5':
                start_date = today.replace(year=today.year - 5)
                orphans = orphans.filter(birth_date__gte=start_date)
            elif age_range == '6-12':
                start_date = today.replace(year=today.year - 12)
                end_date = today.replace(year=today.year - 6)
                orphans = orphans.filter(birth_date__gte=start_date, birth_date__lte=end_date)
            elif age_range == '13-18':
                start_date = today.replace(year=today.year - 18)
                end_date = today.replace(year=today.year - 13)
                orphans = orphans.filter(birth_date__gte=start_date, birth_date__lte=end_date)
            elif age_range == '18+':
                end_date = today.replace(year=today.year - 18)
                orphans = orphans.filter(birth_date__lte=end_date)
    
    # ترتيب النتائج
    orphans = orphans.order_by('-created_at', 'name')
    
    # التقسيم إلى صفحات
    paginator = Paginator(orphans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_count': orphans.count(),
    }
    
    return render(request, 'orphans/orphans_list.html', context)

def orphan_add(request):
    """إضافة يتيم جديد"""
    if request.method == 'POST':
        form = OrphanForm(request.POST)
        if form.is_valid():
            orphan = form.save()
            messages.success(request, f'تم إضافة اليتيم {orphan.name} بنجاح')
            return redirect('orphans:orphan_detail', pk=orphan.pk)
    else:
        form = OrphanForm()
    
    return render(request, 'orphans/orphan_form.html', {
        'form': form,
        'title': 'إضافة يتيم جديد'
    })

def orphan_detail(request, pk):
    """عرض تفاصيل يتيم"""
    orphan = get_object_or_404(Orphan, pk=pk)
    return render(request, 'orphans/orphan_detail.html', {
        'orphan': orphan
    })

def orphan_edit(request, pk):
    """تعديل بيانات يتيم"""
    orphan = get_object_or_404(Orphan, pk=pk)
    
    if request.method == 'POST':
        form = OrphanForm(request.POST, instance=orphan)
        if form.is_valid():
            form.save()
            messages.success(request, f'تم تحديث بيانات اليتيم {orphan.name} بنجاح')
            return redirect('orphans:orphan_detail', pk=pk)
    else:
        form = OrphanForm(instance=orphan)
    
    return render(request, 'orphans/orphan_form.html', {
        'form': form,
        'orphan': orphan,
        'title': f'تعديل بيانات اليتيم: {orphan.name}'
    })

def orphan_delete(request, pk):
    """حذف يتيم"""
    orphan = get_object_or_404(Orphan, pk=pk)
    
    if request.method == 'POST':
        orphan_name = orphan.name
        orphan.delete()
        messages.success(request, f'تم حذف اليتيم {orphan_name} بنجاح')
        return redirect('orphans:orphans_list')
    
    return render(request, 'orphans/orphan_delete.html', {
        'orphan': orphan
    })

def orphans_import(request):
    """استيراد الأيتام من ملف Excel"""
    if request.method == 'POST':
        form = OrphanImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['excel_file']
                df = pd.read_excel(excel_file)
                
                # التحقق من وجود الأعمدة المطلوبة
                required_columns = ['name', 'national_id', 'birth_date', 'gender', 
                                  'deceased_national_id', 'deceased_name', 'guardian_name']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'الأعمدة التالية مفقودة في الملف: {", ".join(missing_columns)}')
                    return render(request, 'orphans/orphans_import.html', {'form': form})
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        orphan_data = {
                            'name': row['name'],
                            'national_id': str(row['national_id']).zfill(9),
                            'birth_date': pd.to_datetime(row['birth_date']).date(),
                            'gender': row['gender'],
                            'deceased_national_id': str(row['deceased_national_id']).zfill(9),
                            'deceased_name': row['deceased_name'],
                            'guardian_name': row['guardian_name'],
                            'guardian_relationship': row.get('guardian_relationship', 'other'),
                            'phone_number': row.get('phone_number', ''),
                            'health_status': row.get('health_status', 'good'),
                            'education_level': row.get('education_level', 'not_enrolled'),
                            'male_siblings_count': int(row.get('male_siblings_count', 0)),
                            'female_siblings_count': int(row.get('female_siblings_count', 0)),
                            'notes': row.get('notes', '')
                        }
                        
                        # التحقق من عدم وجود رقم الهوية مسبقاً
                        if not Orphan.objects.filter(national_id=orphan_data['national_id']).exists():
                            Orphan.objects.create(**orphan_data)
                            success_count += 1
                        else:
                            errors.append(f'الصف {index + 2}: رقم الهوية {orphan_data["national_id"]} موجود مسبقاً')
                            error_count += 1
                            
                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')
                        error_count += 1
                
                if success_count > 0:
                    messages.success(request, f'تم استيراد {success_count} يتيم بنجاح')
                
                if error_count > 0:
                    error_msg = f'فشل في استيراد {error_count} سجل. الأخطاء:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'\n... و {len(errors) - 10} أخطاء أخرى'
                    messages.warning(request, error_msg)
                
                return redirect('orphans:orphans_list')
                
            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء قراءة الملف: {str(e)}')
    else:
        form = OrphanImportForm()
    
    return render(request, 'orphans/orphans_import.html', {'form': form})

def orphans_ajax_search(request):
    """البحث السريع عبر AJAX"""
    query = request.GET.get('q', '')
    if len(query) >= 2:
        orphans = search_orphans(query)[:10]
        results = [
            {
                'id': orphan.id,
                'name': orphan.name,
                'national_id': orphan.national_id,
                'age': orphan.age if orphan.age else 'غير محدد',
                'district': orphan.district.name if orphan.district else ''
            }
            for orphan in orphans
        ]
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})

def export_orphans_excel(request):
    """تصدير بيانات الأيتام إلى Excel"""
    # تطبيق نفس الفلاتر المستخدمة في القائمة
    form = OrphanSearchForm(request.GET)
    orphans = Orphan.objects.all().select_related('district')  # إزالة guardian لأنه غير موجود
    
    if form.is_valid():
        search_query = form.cleaned_data.get('search')
        if search_query:
            orphans = search_orphans(search_query)
        
        gender = form.cleaned_data.get('gender')
        if gender:
            orphans = orphans.filter(gender=gender)
            
        education_level = form.cleaned_data.get('education_level')
        if education_level:
            orphans = orphans.filter(education_level=education_level)
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الأيتام"
    
    # العناوين
    headers = [
        'اسم اليتيم', 'رقم الهوية', 'الجنس', 'تاريخ الميلاد', 'العمر',
        'عدد الأخوة الذكور', 'عدد الأخوة الإناث', 'اسم الوصي', 'صلة القرابة', 
        'المرحلة التعليمية', 'الحالة الصحية', 'الحي', 'رقم الهاتف', 
        'اسم الشهيد/المتوفي', 'رقم هوية الشهيد', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, orphan in enumerate(orphans, 2):
        ws.cell(row=row, column=1, value=orphan.name)
        ws.cell(row=row, column=2, value=orphan.national_id)
        ws.cell(row=row, column=3, value=orphan.get_gender_display())
        ws.cell(row=row, column=4, value=orphan.birth_date.strftime('%Y-%m-%d') if orphan.birth_date else "")
        ws.cell(row=row, column=5, value=orphan.age if orphan.age else "")
        ws.cell(row=row, column=6, value=orphan.male_siblings_count or 0)
        ws.cell(row=row, column=7, value=orphan.female_siblings_count or 0)
        ws.cell(row=row, column=8, value=orphan.guardian_name or "")
        ws.cell(row=row, column=9, value=orphan.get_guardian_relationship_display() if orphan.guardian_relationship else "")
        ws.cell(row=row, column=10, value=orphan.get_education_level_display())
        ws.cell(row=row, column=11, value=orphan.get_health_status_display())
        ws.cell(row=row, column=12, value=orphan.district.name if orphan.district else "")
        ws.cell(row=row, column=13, value=orphan.phone_number or "")
        ws.cell(row=row, column=14, value=orphan.deceased_name or "")
        ws.cell(row=row, column=15, value=orphan.deceased_national_id or "")
        ws.cell(row=row, column=16, value=orphan.notes or "")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = f"الأيتام_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response

def import_orphans_excel(request):
    """استيراد بيانات الأيتام من Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            
            # تحديد الأعمدة التي يجب قراءتها كنص لتجنب مشكلة الأرقام العشرية
            dtype_dict = {
                'رقم الهوية': str,
                'رقم هوية الشهيد': str,
                'رقم الهاتف': str,
                'رقم هوية ولي الأمر': str
            }
            
            df = pd.read_excel(excel_file, dtype=dtype_dict)
            
            def clean_national_id(value):
                """تنظيف رقم الهوية وإضافة الأصفار إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                # تحويل إلى string وإزالة المسافات والنقاط العشرية
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                # إضافة الأصفار في البداية إذا كان الرقم أقل من 9 خانات
                if cleaned.isdigit() and len(cleaned) < 9:
                    cleaned = cleaned.zfill(9)
                
                return cleaned
            
            def clean_phone_number(value):
                """تنظيف رقم الجوال وإضافة الصفر إذا لزم الأمر"""
                if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                    return ''
                
                # تحويل إلى string وإزالة المسافات والنقاط العشرية
                cleaned = str(value).strip().replace('.0', '').replace('.', '')
                
                # إضافة الصفر في البداية إذا كان الرقم يبدأ بـ 5
                if cleaned.isdigit() and len(cleaned) == 9 and cleaned.startswith('5'):
                    cleaned = '0' + cleaned
                
                return cleaned
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات المطلوبة
                    if pd.isna(row.get('اسم اليتيم')) or pd.isna(row.get('رقم الهوية')):
                        errors.append(f"الصف {index + 2}: اسم اليتيم ورقم الهوية مطلوبان")
                        error_count += 1
                        continue
                    
                    # تنظيف البيانات
                    national_id = clean_national_id(row.get('رقم الهوية', ''))
                    deceased_national_id = clean_national_id(row.get('رقم هوية الشهيد', ''))
                    phone_number = clean_phone_number(row.get('رقم الهاتف', ''))
                    guardian_national_id = clean_national_id(row.get('رقم هوية ولي الأمر', ''))
                    
                    # التحقق من صحة رقم الهوية
                    if not national_id or len(national_id) != 9 or not national_id.isdigit():
                        errors.append(f"الصف {index + 2}: رقم هوية غير صحيح '{national_id}'")
                        error_count += 1
                        continue
                    
                    # البحث عن ولي الأمر إذا كان موجود
                    guardian = None
                    if guardian_national_id and len(guardian_national_id) == 9:
                        try:
                            guardian = Guardian.objects.get(national_id=guardian_national_id)
                        except Guardian.DoesNotExist:
                            pass
                    
                    # تحديد عدد الأخوة الذكور
                    male_siblings_count = 0
                    if not pd.isna(row.get('عدد الأخوة الذكور', '')):
                        try:
                            male_siblings_count = int(row['عدد الأخوة الذكور'])
                            if male_siblings_count < 0:
                                male_siblings_count = 0
                        except (ValueError, TypeError):
                            male_siblings_count = 0
                    
                    # تحديد عدد الأخوة الإناث
                    female_siblings_count = 0
                    if not pd.isna(row.get('عدد الأخوة الإناث', '')):
                        try:
                            female_siblings_count = int(row['عدد الأخوة الإناث'])
                            if female_siblings_count < 0:
                                female_siblings_count = 0
                        except (ValueError, TypeError):
                            female_siblings_count = 0
                    
                    # البحث عن الحي إذا كان موجود
                    district = None
                    if not pd.isna(row.get('الحي')) or not pd.isna(row.get('المنطقة')):
                        district_name = str(row.get('الحي', row.get('المنطقة', ''))).strip()
                        if district_name:
                            try:
                                district = District.objects.get(name=district_name)
                            except District.DoesNotExist:
                                pass
                    
                    # إنشاء اليتيم
                    orphan = Orphan.objects.create(
                        name=str(row['اسم اليتيم']).strip(),
                        national_id=national_id,
                        gender=str(row.get('الجنس', 'ذكر')).strip() if not pd.isna(row.get('الجنس')) else 'ذكر',
                        birth_date=pd.to_datetime(row['تاريخ الميلاد']).date() if not pd.isna(row.get('تاريخ الميلاد')) else None,
                        male_siblings_count=male_siblings_count,
                        female_siblings_count=female_siblings_count,
                        guardian_name=str(row.get('اسم الوصي', '')).strip() if not pd.isna(row.get('اسم الوصي')) else '',
                        guardian_relationship=str(row.get('صلة القرابة', '')).strip() if not pd.isna(row.get('صلة القرابة')) else '',
                        health_status=str(row.get('الحالة الصحية', 'جيدة')).strip() if not pd.isna(row.get('الحالة الصحية')) else 'جيدة',
                        education_level=str(row.get('المرحلة الدراسية', '')).strip() if not pd.isna(row.get('المرحلة الدراسية')) else '',
                        deceased_name=str(row.get('اسم الشهيد/المتوفي', '')).strip() if not pd.isna(row.get('اسم الشهيد/المتوفي')) else '',
                        deceased_national_id=deceased_national_id,
                        phone_number=phone_number,
                        district=district,
                        notes=str(row.get('ملاحظات', '')).strip() if not pd.isna(row.get('ملاحظات')) else ''
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"الصف {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"تم استيراد {success_count} يتيم بنجاح")
            
            if error_count > 0:
                messages.warning(request, f"فشل في استيراد {error_count} صف. الأخطاء: {'; '.join(errors[:5])}")
                
        except Exception as e:
            messages.error(request, f"خطأ في قراءة الملف: {str(e)}")
    
    return redirect('orphans:orphans_list')

def download_orphans_template(request):
    """تحميل نموذج Excel فارغ للأيتام"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نموذج الأيتام"
    
    # العناوين
    headers = [
        'اسم اليتيم', 'رقم الهوية', 'الجنس', 'تاريخ الميلاد', 'عدد الأخوة الذكور',
        'عدد الأخوة الإناث', 'اسم الوصي', 'صلة القرابة', 'المرحلة التعليمية', 
        'الحالة الصحية', 'اسم الشهيد/المتوفي', 'رقم هوية الشهيد', 'رقم الهاتف', 
        'الحي', 'ملاحظات'
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة صف مثال
    example_data = [
        'أحمد محمد', '123456789', 'M', '2010-01-15', 2, 1, 'محمد أحمد',
        'grandfather', 'elementary', 'good', 'فلان الفلاني',
        '987654321', '0591234567', 'رفح', 'يتيم الأب'
    ]
    
    for col, data in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=data)
    
    # إضافة ملاحظات توضيحية
    ws.cell(row=4, column=1, value="ملاحظات:")
    ws.cell(row=5, column=1, value="• الجنس: M للذكر، F للأنثى")
    ws.cell(row=6, column=1, value="• عدد الأخوة الذكور والإناث: أرقام صحيحة (مثال: 2, 1)")
    ws.cell(row=7, column=1, value="• صلة القرابة: grandfather, grandmother, uncle, aunt, brother, sister, other")
    ws.cell(row=8, column=1, value="• المرحلة التعليمية: kindergarten, elementary, middle, high, university, graduated, not_enrolled")
    ws.cell(row=9, column=1, value="• الحالة الصحية: excellent, good, fair, poor, chronic, disabled")
    
    # تعديل عرض الأعمدة
    auto_adjust_column_width(ws)
    
    # حفظ الملف في response
    filename = "نموذج_الأيتام.xlsx"
    response = create_excel_response(filename)
    wb.save(response)
    return response
